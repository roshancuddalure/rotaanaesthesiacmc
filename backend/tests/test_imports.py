from pathlib import Path
from datetime import datetime

from openpyxl import Workbook

from app.services.imports import (
    classify_duty_label,
    clean_person_name,
    detect_import_kind,
    iter_excel_source_cells,
    parse_rota_workbook,
    parse_unitwise_workbook,
    profile_excel_workbook,
    reconstruct_month_from_path,
    reconstruct_rota_date,
    reconstruct_month_from_text,
    is_valid_person_name,
    split_unitwise_names,
)


def create_sample_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "May 2026"
    worksheet["A1"] = "Date"
    worksheet["B1"] = "Main 1st Call"
    worksheet["A2"] = 1
    worksheet["B2"] = "Dr Example (JR)"
    workbook.create_sheet("Unitwise")
    workbook["Unitwise"]["A1"] = "PAC"
    workbook.save(path)


def create_sample_rota(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["A1"] = "DATE"
    worksheet["B1"] = datetime(2026, 1, 5)
    worksheet["C1"] = datetime(2026, 2, 5)
    worksheet["A2"] = "DAY"
    worksheet["B2"] = "Friday"
    worksheet["C2"] = "Saturday"
    worksheet["A3"] = "Main 1st Call"
    worksheet["B3"] = "Dr Example (JR)"
    worksheet["C3"] = "xxxxx"
    worksheet["A4"] = "Caesar A"
    worksheet["B4"] = "Dr Second"
    workbook.save(path)


def create_sample_unitwise(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["A1"] = ""
    worksheet["B1"] = "UNIT I"
    worksheet["C1"] = "UNIT II"
    worksheet["A2"] = "5th calls"
    worksheet["B2"] = "Dr Senior"
    worksheet["C2"] = "Dr Other"
    worksheet["A3"] = ""
    worksheet["B3"] = "Dr More"
    workbook.save(path)


def test_detect_import_kind() -> None:
    assert detect_import_kind(Path("rota.xlsx")) == "excel"
    assert detect_import_kind(Path("unitwise.csv")) == "text"
    assert detect_import_kind(Path("notes.pdf")) == "unknown"


def test_profile_excel_workbook_records_sheet_shape(tmp_path: Path) -> None:
    workbook_path = tmp_path / "May 2026 rota.xlsx"
    create_sample_workbook(workbook_path)

    profile = profile_excel_workbook(workbook_path)

    assert [sheet.name for sheet in profile.sheets] == ["May 2026", "Unitwise"]
    assert profile.sheets[0].non_empty_cells == 4
    assert profile.sheets[0].first_non_empty_row == 1
    assert profile.sheets[0].first_non_empty_column == 1


def test_iter_excel_source_cells_preserves_traceability(tmp_path: Path) -> None:
    workbook_path = tmp_path / "May 2026 rota.xlsx"
    create_sample_workbook(workbook_path)

    cells = iter_excel_source_cells(workbook_path, sheet_name="May 2026")

    assert cells[0].sheet_name == "May 2026"
    assert cells[0].row_index == 1
    assert cells[0].column_label == "A"
    assert cells[-1].cleaned_value == "Dr Example (JR)"


def test_cleaning_and_classification_helpers() -> None:
    assert clean_person_name("  Dr Example (JR)  ") == "Example"
    assert classify_duty_label("Main 1st Call") == "MAIN_1ST_24HR"
    assert classify_duty_label("main_1st_24hr") == "MAIN_1ST_24HR"
    assert classify_duty_label("Cesar call A") == "CAESAR_A_12HR"
    assert classify_duty_label("RC PAC4th call") == "PAC"
    assert classify_duty_label("Schell call + Floating consultant") == "SCHELL_AND_FLOATING"
    assert classify_duty_label("RC Co-call") == "RC_CO_12HR"
    assert is_valid_person_name("Dr Example")
    assert not is_valid_person_name("DATE")
    assert not is_valid_person_name("2025-01-01")
    assert not is_valid_person_name("JULY")
    assert not is_valid_person_name("May 1-15")
    assert not is_valid_person_name("UNIT I")
    assert not is_valid_person_name("2.0")
    assert split_unitwise_names("Angeline Anirutha SICU ONLY / Rohan Chacko (MICU 5-15)") == [
        "Angeline Anirutha",
        "Rohan Chacko",
    ]


def test_reconstruct_month_from_text() -> None:
    reconstructed = reconstruct_month_from_text("Duty Rota - May 2026")

    assert reconstructed is not None
    assert reconstructed.year == 2026
    assert reconstructed.month == 5


def test_reconstruct_month_from_path_handles_short_year() -> None:
    reconstructed = reconstruct_month_from_path(Path("MARCH ROTA 26.xlsx"))

    assert reconstructed is not None
    assert reconstructed.year == 2026
    assert reconstructed.month == 3


def test_reconstruct_rota_date_handles_day_month_swapped_excel_date() -> None:
    reconstructed = reconstruct_month_from_text("May 2026")

    assert reconstructed is not None
    assert reconstruct_rota_date(datetime(2026, 1, 5), reconstructed).day == 1


def test_parse_rota_workbook_returns_traceable_assignments(tmp_path: Path) -> None:
    workbook_path = tmp_path / "May Rota 2026.xlsx"
    create_sample_rota(workbook_path)

    parsed = parse_rota_workbook(workbook_path)

    assert parsed.month.year == 2026
    assert parsed.month.month == 5
    assert len(parsed.assignments) == 2
    assert parsed.assignments[0].duty_date.day == 1
    assert parsed.assignments[0].duty_type == "MAIN_1ST_24HR"
    assert parsed.assignments[0].person_name == "Example"
    assert parsed.assignments[0].column_label == "B"
    assert parsed.warnings == ()


def test_parse_unitwise_workbook_returns_postings(tmp_path: Path) -> None:
    workbook_path = tmp_path / "May 2026.xlsx"
    create_sample_unitwise(workbook_path)

    parsed = parse_unitwise_workbook(workbook_path)

    assert len(parsed.postings) == 3
    assert parsed.postings[0].unit_label == "UNIT I"
    assert parsed.postings[0].posting_label == "5th calls"
    assert parsed.postings[0].person_name == "Senior"
    assert parsed.postings[2].person_name == "More"


def test_parse_rota_workbook_splits_schell_floating_row(tmp_path: Path) -> None:
    workbook_path = tmp_path / "May Rota 2026.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "DATE"
    worksheet["B1"] = datetime(2026, 1, 5)
    worksheet["A2"] = "DAY"
    worksheet["B2"] = "Friday"
    worksheet["A3"] = "Schell call + Floating consultant"
    worksheet["B3"] = "Dr Example"
    workbook.save(workbook_path)

    parsed = parse_rota_workbook(workbook_path)

    assert [assignment.duty_type for assignment in parsed.assignments] == [
        "SCHELL_24HR",
        "FLOATING_24HR",
    ]
