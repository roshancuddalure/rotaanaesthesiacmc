import json
from collections.abc import Generator
from contextlib import contextmanager
from datetime import date
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app import models  # noqa: F401
from app.db.session import Base, get_db
from app.main import app
from app.models import AdminMapping, Person, PersonPosting, Unit
from app.services.auth import seed_superadmin
from app.services.unit_assignment_import import normalize_import_posting, parse_unitwise_excel_upload


@contextmanager
def unit_client() -> Generator[TestClient, None, None]:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    testing_session = sessionmaker(bind=engine)

    with testing_session() as seed_session:
        seed_superadmin(seed_session)
        sujil = Person(canonical_name="Sujil", call_level="3RD_CALL")
        jeenu = Person(canonical_name="Jeenu Ann Jose", call_level="4TH_CALL")
        nasreen = Person(canonical_name="Nasreen Begum K", call_level="2ND_CALL")
        jessica = Person(canonical_name="Jessica Charles", call_level="2ND_CALL")
        renita = Person(canonical_name="Renita J", call_level="2ND_CALL")
        sudharsan = Person(canonical_name="Sudharsan T R", call_level="2ND_CALL")
        calvin = Person(canonical_name="Calvin Lawrence Dalmeida", call_level="2ND_CALL")
        angelin = Person(canonical_name="Angelin Aniruth", call_level="2ND_CALL")
        main = Unit(code="MAIN", name="Main OT", campus="CMC")
        cardiac = Unit(code="CARDIAC", name="Cardiac OT", campus="CMC")
        neuro = Unit(code="NEURO_DEPT", name="Dept of Neuroanaesthesia", campus="CMC")
        unit_five = Unit(code="UNIT_5", name="Unit 5", campus="CMC")
        unit_one = Unit(code="GENERAL_ALPHA", name="General Unit 1", campus="CMC")
        unit_two = Unit(code="GENERAL_BETA", name="General Unit 2", campus="CMC")
        unit_three = Unit(code="GENERAL_GAMMA", name="General Unit 3", campus="CMC")
        seed_session.add_all(
            [
                sujil,
                jeenu,
                nasreen,
                jessica,
                renita,
                sudharsan,
                calvin,
                angelin,
                main,
                cardiac,
                neuro,
                unit_five,
                unit_one,
                unit_two,
                unit_three,
                AdminMapping(
                    mapping_type="unit_label",
                    source_label="UNIT FIVE",
                    target_key="UNIT_5",
                    target_label="Unit 5",
                    status="reviewed",
                    source="test",
                ),
                PersonPosting(
                    person=jeenu,
                    unit=main,
                    posting_type="4TH_CALL",
                    starts_on=date(2026, 6, 1),
                    ends_on=date(2026, 6, 30),
                    source="historical_import",
                ),
            ]
        )
        seed_session.commit()

    def override_get_db() -> Generator[Session, None, None]:
        with testing_session() as session:
            yield session

    app.dependency_overrides[get_db] = override_get_db
    try:
        yield TestClient(app)
    finally:
        app.dependency_overrides.clear()


def sign_in(client: TestClient) -> str:
    response = client.post(
        "/api/v1/auth/sign-in",
        json={"username": "rotachief", "password": "rotateam"},
    )
    assert response.status_code == 200
    return response.json()["token"]


def auth_headers(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def seeded_ids(client: TestClient, token: str) -> tuple[str, str, str]:
    members = client.get("/api/v1/admin/members", headers=auth_headers(token)).json()
    units = client.get("/api/v1/units", headers=auth_headers(token)).json()
    sujil_id = next(member["id"] for member in members if member["canonical_name"] == "Sujil")
    main_id = next(unit["id"] for unit in units if unit["code"] == "MAIN")
    cardiac_id = next(unit["id"] for unit in units if unit["code"] == "CARDIAC")
    return sujil_id, main_id, cardiac_id


def test_unit_management_month_lists_assignment_and_leave_summary() -> None:
    with unit_client() as client:
        token = sign_in(client)
        sujil_id, main_id, _cardiac_id = seeded_ids(client, token)

        response = client.post(
            "/api/v1/unit-management/assignments",
            headers=auth_headers(token),
            json={
                "person_id": sujil_id,
                "unit_id": main_id,
                "posting_type": "3rd call",
                "starts_on": "2026-06-01",
                "ends_on": "2026-06-30",
                "notes": "June main OT posting",
            },
        )
        assert response.status_code == 200
        assignment = response.json()
        assert assignment["posting_type"] == "3RD_CALL"
        assert assignment["person"]["canonical_name"] == "Sujil"
        assert assignment["source"] == "unit_board"

        leave_response = client.post(
            "/api/v1/leave/requests",
            headers=auth_headers(token),
            json={
                "person_id": sujil_id,
                "leave_type": "ANNUAL_LEAVE",
                "leave_slot": "FULL_DAY",
                "starts_on": "2026-06-10",
                "ends_on": "2026-06-12",
                "status": "approved",
            },
        )
        assert leave_response.status_code == 200

        response = client.get(
            "/api/v1/unit-management/month?month=2026-06",
            headers=auth_headers(token),
        )
        assert response.status_code == 200
        month = response.json()
        assert month["month"] == "2026-06"
        assert next(unit for unit in month["units"] if unit["id"] == main_id)["minimum_free_people"] == 1
        assert len(month["assignments"]) == 1
        assert month["assignments"][0]["source"] == "unit_board"
        assert month["validation_issues"] == []

        main_summary = next(summary for summary in month["unit_summaries"] if summary["unit_id"] == main_id)
        assert main_summary["assigned_members"] == 1
        assert main_summary["people_with_leave"] == 1
        assert main_summary["leave_days"] == 3
        assert main_summary["leave_by_call_level"] == {"3RD_CALL": 3}


def test_special_unit_postings_can_be_created_without_unit() -> None:
    with unit_client() as client:
        token = sign_in(client)
        headers = auth_headers(token)
        sujil_id, _main_id, _cardiac_id = seeded_ids(client, token)

        response = client.post(
            "/api/v1/unit-management/assignments",
            headers=headers,
            json={
                "person_id": sujil_id,
                "unit_id": None,
                "posting_type": "PAIN",
                "starts_on": "2026-06-01",
                "ends_on": "2026-06-30",
                "notes": "Pain call posting",
            },
        )

        assert response.status_code == 200
        created = response.json()
        assert created["unit"] is None
        assert created["posting_type"] == "PAIN"

        month = client.get("/api/v1/unit-management/month?month=2026-06", headers=headers).json()
        pain_assignment = next(row for row in month["assignments"] if row["id"] == created["id"])
        assert pain_assignment["unit"] is None
        assert pain_assignment["posting_type"] == "PAIN"
        assert all(issue["code"] != "MISSING_UNIT" for issue in month["validation_issues"])


def test_unit_management_updates_unit_minimum_free_people() -> None:
    with unit_client() as client:
        token = sign_in(client)
        _sujil_id, main_id, _cardiac_id = seeded_ids(client, token)

        response = client.put(
            f"/api/v1/unit-management/units/{main_id}/settings",
            headers=auth_headers(token),
            json={"minimum_free_people": 3},
        )

        assert response.status_code == 200
        assert response.json()["minimum_free_people"] == 3

        month = client.get(
            "/api/v1/unit-management/month?month=2026-06",
            headers=auth_headers(token),
        ).json()
        assert next(unit for unit in month["units"] if unit["id"] == main_id)["minimum_free_people"] == 3


def test_unit_management_updates_and_validates_call_wise_minimums() -> None:
    with unit_client() as client:
        token = sign_in(client)
        sujil_id, main_id, _cardiac_id = seeded_ids(client, token)

        create_response = client.post(
            "/api/v1/unit-management/assignments",
            headers=auth_headers(token),
            json={
                "person_id": sujil_id,
                "unit_id": main_id,
                "posting_type": "3rd call",
                "starts_on": "2026-06-01",
                "ends_on": "2026-06-30",
            },
        )
        assert create_response.status_code == 200

        response = client.put(
            f"/api/v1/unit-management/units/{main_id}/settings?month=2026-06",
            headers=auth_headers(token),
            json={
                "minimum_free_people": 1,
                "call_minimums": [{"call_level": "3rd call", "minimum_free_people": 1}],
            },
        )
        assert response.status_code == 200

        month = client.get(
            "/api/v1/unit-management/month?month=2026-06",
            headers=auth_headers(token),
        ).json()
        call_row = next(
            row
            for row in month["unit_call_minimums"]
            if row["unit_id"] == main_id and row["call_level"] == "3RD_CALL"
        )
        assert call_row["assigned_members"] == 1
        assert call_row["minimum_free_people"] == 1
        assert call_row["max_allowed"] == 1

        blocked = client.put(
            f"/api/v1/unit-management/units/{main_id}/settings?month=2026-06",
            headers=auth_headers(token),
            json={
                "minimum_free_people": 1,
                "call_minimums": [{"call_level": "3rd call", "minimum_free_people": 2}],
            },
        )
        assert blocked.status_code == 400
        assert "cannot exceed assigned members" in blocked.json()["detail"]


def test_unit_management_flags_overlapping_primary_assignment_and_allows_update_delete() -> None:
    with unit_client() as client:
        token = sign_in(client)
        sujil_id, main_id, cardiac_id = seeded_ids(client, token)

        first = client.post(
            "/api/v1/unit-management/assignments",
            headers=auth_headers(token),
            json={
                "person_id": sujil_id,
                "unit_id": main_id,
                "posting_type": "3RD_CALL",
                "starts_on": "2026-06-01",
                "ends_on": "2026-06-30",
            },
        )
        assert first.status_code == 200

        second = client.post(
            "/api/v1/unit-management/assignments",
            headers=auth_headers(token),
            json={
                "person_id": sujil_id,
                "unit_id": cardiac_id,
                "posting_type": "3RD_CALL",
                "starts_on": "2026-06-15",
                "ends_on": "2026-06-20",
            },
        )
        assert second.status_code == 200
        second_id = second.json()["id"]

        response = client.get(
            "/api/v1/unit-management/month?month=2026-06",
            headers=auth_headers(token),
        )
        issues = response.json()["validation_issues"]
        assert any(issue["code"] == "OVERLAPPING_PRIMARY_ASSIGNMENT" for issue in issues)
        assert any(issue["code"] == "MULTIPLE_UNITS_IN_MONTH" for issue in issues)

        updated = client.put(
            f"/api/v1/unit-management/assignments/{second_id}",
            headers=auth_headers(token),
            json={
                "person_id": sujil_id,
                "unit_id": cardiac_id,
                "posting_type": "PAIN",
                "starts_on": "2026-06-15",
                "ends_on": "2026-06-20",
                "notes": "Special posting",
            },
        )
        assert updated.status_code == 200
        assert updated.json()["posting_type"] == "PAIN"

        response = client.get(
            "/api/v1/unit-management/month?month=2026-06",
            headers=auth_headers(token),
        )
        issues = response.json()["validation_issues"]
        assert not any(issue["code"] == "OVERLAPPING_PRIMARY_ASSIGNMENT" for issue in issues)
        assert not any(issue["code"] == "MULTIPLE_UNITS_IN_MONTH" for issue in issues)

        deleted = client.delete(
            f"/api/v1/unit-management/assignments/{second_id}",
            headers=auth_headers(token),
        )
        assert deleted.status_code == 200
        assert deleted.json() == {"status": "deleted"}

        response = client.get(
            "/api/v1/unit-management/month?month=2026-06",
            headers=auth_headers(token),
        )
        assert len(response.json()["assignments"]) == 1


def unitwise_workbook_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "June"
    worksheet["A1"] = ""
    worksheet["B1"] = "Main OT"
    worksheet["C1"] = "Cardiac OT"
    worksheet["A2"] = "3rd calls"
    worksheet["B2"] = "Dr Sujil"
    worksheet["A3"] = "4th calls"
    worksheet["C3"] = "Dr Jeenu Ann Jose"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_unit_management_imports_unitwise_excel_preview_and_apply() -> None:
    with unit_client() as client:
        token = sign_in(client)
        headers = auth_headers(token)

        preview = client.post(
            "/api/v1/unit-management/import-preview?month=2026-06",
            headers=headers,
            files={
                "file": (
                    "June 2026.xlsx",
                    unitwise_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert preview.status_code == 200
        body = preview.json()
        assert body["matched_rows"] == 2
        assert body["rows"][0]["posting_type"] == "3RD_CALL"
        assert body["rows"][0]["unit_name"] == "Main OT"

        applied = client.post(
            "/api/v1/unit-management/import-apply?month=2026-06&replace_existing=true",
            headers=headers,
            files={
                "file": (
                    "June 2026.xlsx",
                    unitwise_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert applied.status_code == 200
        assert applied.json()["created_rows"] == 2
        assert applied.json()["auto_assigned_rows"] == 2

        month = client.get("/api/v1/unit-management/month?month=2026-06", headers=headers).json()
        assert len(month["assignments"]) == 2
        assert {assignment["posting_type"] for assignment in month["assignments"]} == {
            "3RD_CALL",
            "4TH_CALL",
        }

        replaced = client.post(
            "/api/v1/unit-management/import-apply?month=2026-06&replace_existing=true",
            headers=headers,
            files={
                "file": (
                    "June 2026.xlsx",
                    unitwise_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert replaced.status_code == 200
        replaced_body = replaced.json()
        assert replaced_body["created_rows"] == 2
        assert replaced_body["auto_assigned_rows"] == 2
        assert replaced_body["skipped_rows"] == 0
        assert replaced_body["deleted_existing_rows"] == 2

        duplicate_preview = client.post(
            "/api/v1/unit-management/import-preview?month=2026-06&replace_existing=true",
            headers=headers,
            files={
                "file": (
                    "June 2026.xlsx",
                    unitwise_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert duplicate_preview.status_code == 200
        duplicate_body = duplicate_preview.json()
        assert duplicate_body["auto_assignable_rows"] == 2
        assert duplicate_body["needs_review_rows"] == 0


def test_unit_management_imports_notepad_unitwise_list() -> None:
    with unit_client() as client:
        token = sign_in(client)
        headers = auth_headers(token)
        content = b"""
Main OT
3rd calls: Dr Sujil
Cardiac OT
4th calls: Dr Jeenu Ann Jose
"""

        preview = client.post(
            "/api/v1/unit-management/import-preview?month=2026-06",
            headers=headers,
            files={"file": ("unitwise.txt", content, "text/plain")},
        )
        assert preview.status_code == 200
        body = preview.json()
        assert body["matched_rows"] == 2
        assert body["source_formats"] == ["text_unitwise"]


def test_unit_management_imports_official_text_template_format() -> None:
    with unit_client() as client:
        token = sign_in(client)
        headers = auth_headers(token)
        content = b"""
MONTH: 2026-06

[UNIT: UNIT 1]
2ND_CALL:
- Jessica Charles
- Renita J | 2026-06-16 to 2026-06-30

[UNIT: UNIT III]
2ND_CALL:
- Sudharshan

[PAIN]
- Sujil
- Jeenu Ann Jose | 2026-06-01 to 2026-06-15
"""

        preview = client.post(
            "/api/v1/unit-management/import-preview?month=2026-06&replace_existing=true",
            headers=headers,
            files={"file": ("unitwise-template.txt", content, "text/plain")},
        )

        assert preview.status_code == 200
        body = preview.json()
        assert body["matched_rows"] == 5
        assert body["auto_assignable_rows"] == 5
        pain_rows = [row for row in body["rows"] if row["posting_type"] == "PAIN"]
        assert {row["person_name"] for row in pain_rows} == {"Sujil", "Jeenu Ann Jose"}
        assert all(row["special_posting"] is True for row in pain_rows)
        renita = next(row for row in body["rows"] if row["person_name"] == "Renita J")
        assert (renita["starts_on"], renita["ends_on"]) == ("2026-06-16", "2026-06-30")


def fuzzy_unitwise_workbook_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "June"
    worksheet["A1"] = ""
    worksheet["B1"] = "UNIT FIVE"
    worksheet["A2"] = "2nd Call"
    worksheet["B2"] = "Nazreen Begum"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def unresolved_unitwise_workbook_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "June"
    worksheet["A1"] = ""
    worksheet["B1"] = "Unknown unit label"
    worksheet["A2"] = "2nd Call"
    worksheet["B2"] = "Unknown Imported Name"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def fuzzy_unit_label_workbook_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "June"
    worksheet["A1"] = ""
    worksheet["B1"] = "Main O"
    worksheet["A2"] = "3rd Call"
    worksheet["B2"] = "Sujil"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def numbered_units_workbook_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "June"
    worksheet["A1"] = ""
    worksheet["B1"] = "UNIT 1"
    worksheet["C1"] = "UNIT 2"
    worksheet["D1"] = "UNIT III"
    worksheet["A2"] = "2nd Call"
    worksheet["B2"] = "Jassica Charles"
    worksheet["C2"] = "Renita"
    worksheet["D2"] = "Sudharshan"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def messy_excel_unitwise_workbook_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Messy June"
    worksheet["A1"] = "Anaesthesia unitwise allocation"
    worksheet["A2"] = "Prepared by rota board"
    worksheet["A4"] = "Posting"
    worksheet["B4"] = "UNIT 1"
    worksheet["C4"] = "UNIT III"
    worksheet["A5"] = "Date"
    worksheet["B5"] = "1-15"
    worksheet["C5"] = "16-30"
    worksheet["A6"] = "2nd Call"
    worksheet["B6"] = "Jessica Charles & Renita J"
    worksheet["C6"] = "Sudharshan; Nasreen Begum K"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def department_alias_unitwise_workbook_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "June"
    worksheet["A1"] = ""
    worksheet["B1"] = "NEURO"
    worksheet["C1"] = "CARDIAC"
    worksheet["A2"] = "2023 second calls"
    worksheet["B2"] = "Calwin Lawrence"
    worksheet["C2"] = "Angeline Anirutha"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def manual_learning_workbook_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "June"
    worksheet["A1"] = ""
    worksheet["B1"] = "Heart Dept"
    worksheet["A2"] = "2nd Call"
    worksheet["B2"] = "Cal Law"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def special_posting_workbook_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "June"
    worksheet["A1"] = ""
    worksheet["B1"] = "Unknown special column"
    worksheet["A2"] = "Pain Call"
    worksheet["B2"] = "Dr Sujil"
    worksheet["A3"] = "SICU Posting"
    worksheet["B3"] = "Dr Jeenu Ann Jose"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def nested_pain_call_workbook_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "June"
    worksheet["A1"] = ""
    worksheet["B1"] = "UNIT III"
    worksheet["A2"] = "Pain Call"
    worksheet["A3"] = "2025 first calls"
    worksheet["B3"] = "Sujil, Jeenu Ann Jose"
    worksheet["A4"] = "1st Call"
    worksheet["B4"] = "Sujil"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def block_posting_workbook_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "June"
    worksheet["A1"] = ""
    worksheet["B1"] = "Main OT"
    worksheet["A2"] = "3rd Call"
    worksheet["B2"] = "Dr Sujil BP 1-15"
    worksheet["A3"] = "2nd Call"
    worksheet["B3"] = "block posting 16-30 Dr Nasreen Begum K"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_unit_management_import_auto_matches_confident_name_and_mapped_unit() -> None:
    with unit_client() as client:
        token = sign_in(client)
        headers = auth_headers(token)

        preview = client.post(
            "/api/v1/unit-management/import-preview?month=2026-06",
            headers=headers,
            files={
                "file": (
                    "June 2026.xlsx",
                    fuzzy_unitwise_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert preview.status_code == 200
        body = preview.json()
        assert body["matched_rows"] == 1
        assert body["unresolved_rows"] == 0
        assert body["rows"][0]["person_name"] == "Nasreen Begum K"
        assert body["rows"][0]["match_method"] == "fuzzy_auto"
        assert body["rows"][0]["unit_name"] == "Unit 5"
        assert body["rows"][0]["unit_match_method"] == "unit_exact"

        applied = client.post(
            "/api/v1/unit-management/import-apply?month=2026-06&replace_existing=false",
            headers=headers,
            files={
                "file": (
                    "June 2026.xlsx",
                    fuzzy_unitwise_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert applied.status_code == 200
        assert applied.json()["created_rows"] == 1


def test_unit_management_import_auto_resolves_numbered_units_and_name_variants() -> None:
    with unit_client() as client:
        token = sign_in(client)
        headers = auth_headers(token)

        preview = client.post(
            "/api/v1/unit-management/import-preview?month=2026-06",
            headers=headers,
            files={
                "file": (
                    "June 2026.xlsx",
                    numbered_units_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert preview.status_code == 200
        body = preview.json()
        assert body["matched_rows"] == 3
        assert body["auto_assignable_rows"] == 3
        assert body["needs_review_rows"] == 0
        assert {row["person_name"] for row in body["rows"]} == {
            "Jessica Charles",
            "Renita J",
            "Sudharsan T R",
        }
        assert {row["unit_name"] for row in body["rows"]} == {
            "General Unit 1",
            "General Unit 2",
            "General Unit 3",
        }
        assert all(row["unit_match_method"] == "unit_number_exact" for row in body["rows"])
        assert all(row["row_action"] == "auto_assign" for row in body["rows"])
        assert all(row["auto_decision_reason"] == "Ready for auto-assign" for row in body["rows"])

        applied = client.post(
            "/api/v1/unit-management/import-apply?month=2026-06&replace_existing=true",
            headers=headers,
            files={
                "file": (
                    "June 2026.xlsx",
                    numbered_units_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert applied.status_code == 200
        assert applied.json()["auto_assigned_rows"] == 3


def test_unit_management_import_structures_messy_excel_with_late_headers_dates_and_separators() -> None:
    with unit_client() as client:
        token = sign_in(client)
        headers = auth_headers(token)

        preview = client.post(
            "/api/v1/unit-management/import-preview?month=2026-06&replace_existing=true",
            headers=headers,
            files={
                "file": (
                    "June messy.xlsx",
                    messy_excel_unitwise_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert preview.status_code == 200
        body = preview.json()
        assert body["matched_rows"] == 4
        assert body["auto_assignable_rows"] == 4
        assert {row["person_name"] for row in body["rows"]} == {
            "Jessica Charles",
            "Renita J",
            "Sudharsan T R",
            "Nasreen Begum K",
        }
        assert {(row["starts_on"], row["ends_on"]) for row in body["rows"]} == {
            ("2026-06-01", "2026-06-15"),
            ("2026-06-16", "2026-06-30"),
        }
        assert all(row["parser_rule"] == "excel_header_scan" for row in body["rows"])
        assert all("header row 4" in row["source_context"] for row in body["rows"])


def test_unit_management_import_auto_resolves_department_unit_aliases_and_known_name_variants() -> None:
    with unit_client() as client:
        token = sign_in(client)
        headers = auth_headers(token)

        preview = client.post(
            "/api/v1/unit-management/import-preview?month=2026-06&replace_existing=true",
            headers=headers,
            files={
                "file": (
                    "June dept.xlsx",
                    department_alias_unitwise_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert preview.status_code == 200
        body = preview.json()
        assert body["matched_rows"] == 2
        assert body["auto_assignable_rows"] == 2
        assert {row["person_name"] for row in body["rows"]} == {
            "Calvin Lawrence Dalmeida",
            "Angelin Aniruth",
        }
        assert {row["unit_name"] for row in body["rows"]} == {
            "Dept of Neuroanaesthesia",
            "Cardiac OT",
        }
        assert all(row["unit_match_method"] == "unit_exact" for row in body["rows"])


def test_unit_management_import_learns_alias_and_unit_mapping_from_manual_resolution() -> None:
    with unit_client() as client:
        token = sign_in(client)
        headers = auth_headers(token)
        members = client.get("/api/v1/admin/members", headers=headers).json()
        units = client.get("/api/v1/units", headers=headers).json()
        calvin_id = next(member["id"] for member in members if member["canonical_name"] == "Calvin Lawrence Dalmeida")
        cardiac_id = next(unit["id"] for unit in units if unit["code"] == "CARDIAC")

        first_preview = client.post(
            "/api/v1/unit-management/import-preview?month=2026-06&replace_existing=true",
            headers=headers,
            files={
                "file": (
                    "June learn.xlsx",
                    manual_learning_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert first_preview.status_code == 200
        row_key = first_preview.json()["rows"][0]["row_key"]
        resolutions = {row_key: {"person_id": calvin_id, "unit_id": cardiac_id}}

        applied = client.post(
            "/api/v1/unit-management/import-apply?month=2026-06&replace_existing=true",
            headers=headers,
            data={"resolutions_json": json.dumps(resolutions)},
            files={
                "file": (
                    "June learn.xlsx",
                    manual_learning_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert applied.status_code == 200
        assert applied.json()["learned_mappings"] == 2

        second_preview = client.post(
            "/api/v1/unit-management/import-preview?month=2026-07&replace_existing=true",
            headers=headers,
            files={
                "file": (
                    "July learn.xlsx",
                    manual_learning_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert second_preview.status_code == 200
        learned_row = second_preview.json()["rows"][0]
        assert learned_row["person_name"] == "Calvin Lawrence Dalmeida"
        assert learned_row["unit_name"] == "Cardiac OT"
        assert learned_row["auto_assignable"] is True


def test_unit_management_import_auto_resolves_doubtful_unit_above_threshold_with_review_note() -> None:
    with unit_client() as client:
        token = sign_in(client)
        headers = auth_headers(token)

        preview = client.post(
            "/api/v1/unit-management/import-preview?month=2026-06",
            headers=headers,
            files={
                "file": (
                    "June 2026.xlsx",
                    fuzzy_unit_label_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert preview.status_code == 200
        body = preview.json()
        assert body["matched_rows"] == 1
        assert body["auto_resolved_rows"] == 1
        assert body["review_suggested_rows"] == 1
        row = body["rows"][0]
        assert row["preview_status"] == "matched"
        assert row["unit_name"] == "Main OT"
        assert row["unit_match_method"] == "unit_fuzzy_auto"
        assert row["review_suggested"] is True
        assert any("Unit auto-resolved" in note for note in row["resolution_notes"])


def test_unit_management_import_keeps_special_postings_without_unit() -> None:
    with unit_client() as client:
        token = sign_in(client)
        headers = auth_headers(token)

        preview = client.post(
            "/api/v1/unit-management/import-preview?month=2026-06",
            headers=headers,
            files={
                "file": (
                    "June 2026.xlsx",
                    special_posting_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert preview.status_code == 200
        body = preview.json()
        assert body["matched_rows"] == 2
        assert body["unresolved_rows"] == 0
        assert {row["posting_type"] for row in body["rows"]} == {"PAIN", "SICU"}
        assert all(row["special_posting"] is True for row in body["rows"])
        assert all(row["unit_id"] is None for row in body["rows"])
        assert all("Unresolved unit" not in row["issues"] for row in body["rows"])

        applied = client.post(
            "/api/v1/unit-management/import-apply?month=2026-06&replace_existing=true",
            headers=headers,
            files={
                "file": (
                    "June 2026.xlsx",
                    special_posting_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert applied.status_code == 200
        assert applied.json()["created_rows"] == 2

        month = client.get("/api/v1/unit-management/month?month=2026-06", headers=headers).json()
        assert {assignment["posting_type"] for assignment in month["assignments"]} == {"PAIN", "SICU"}
        assert all(assignment["unit"] is None for assignment in month["assignments"])
        assert month["validation_issues"] == []


def test_unit_management_import_keeps_nested_pain_child_rows_as_special_split_people() -> None:
    with unit_client() as client:
        token = sign_in(client)
        headers = auth_headers(token)

        preview = client.post(
            "/api/v1/unit-management/import-preview?month=2026-06&replace_existing=true",
            headers=headers,
            files={
                "file": (
                    "June 2026.xlsx",
                    nested_pain_call_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert preview.status_code == 200
        body = preview.json()
        pain_rows = [row for row in body["rows"] if row["posting_type"] == "PAIN"]
        normal_rows = [row for row in body["rows"] if row["posting_type"] == "1ST_CALL"]
        assert len(pain_rows) == 2
        assert len(normal_rows) == 1
        assert {row["person_name"] for row in pain_rows} == {"Sujil", "Jeenu Ann Jose"}
        assert all(row["special_posting"] is True for row in pain_rows)
        assert all(row["unit_id"] is None for row in pain_rows)
        assert all(row["raw_posting_label"] == "Pain Call" for row in pain_rows)
        assert all(row["child_posting_label"] == "2025 first calls" for row in pain_rows)
        assert all("Person appears more than once" not in " ".join(row["issues"]) for row in body["rows"])
        assert body["auto_assignable_rows"] == 3

        applied = client.post(
            "/api/v1/unit-management/import-apply?month=2026-06&replace_existing=true",
            headers=headers,
            files={
                "file": (
                    "June 2026.xlsx",
                    nested_pain_call_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert applied.status_code == 200
        assert applied.json()["auto_assigned_rows"] == 3


def test_unit_management_import_parses_block_posting_labels_as_normal_unit_assignment() -> None:
    with unit_client() as client:
        token = sign_in(client)
        headers = auth_headers(token)

        preview = client.post(
            "/api/v1/unit-management/import-preview?month=2026-06",
            headers=headers,
            files={
                "file": (
                    "June 2026.xlsx",
                    block_posting_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert preview.status_code == 200
        body = preview.json()
        assert body["matched_rows"] == 2
        assert [row["cleaned_person_name"] for row in body["rows"]] == ["Sujil", "Nasreen Begum K"]
        assert [(row["starts_on"], row["ends_on"]) for row in body["rows"]] == [
            ("2026-06-01", "2026-06-15"),
            ("2026-06-16", "2026-06-30"),
        ]
        assert all(row["unit_name"] == "Main OT" for row in body["rows"])
        assert all(not row["special_posting"] for row in body["rows"])


def test_unit_management_import_applies_manual_row_resolutions() -> None:
    with unit_client() as client:
        token = sign_in(client)
        headers = auth_headers(token)
        sujil_id, main_id, _cardiac_id = seeded_ids(client, token)

        first_preview = client.post(
            "/api/v1/unit-management/import-preview?month=2026-06",
            headers=headers,
            files={
                "file": (
                    "June 2026.xlsx",
                    unresolved_unitwise_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert first_preview.status_code == 200
        unresolved = first_preview.json()
        row_key = unresolved["rows"][0]["row_key"]
        assert unresolved["matched_rows"] == 0

        resolutions = {
            row_key: {
                "person_id": sujil_id,
                "unit_id": main_id,
                "posting_type": "2ND_CALL",
            }
        }
        preview = client.post(
            "/api/v1/unit-management/import-preview?month=2026-06",
            headers=headers,
            data={"resolutions_json": json.dumps(resolutions)},
            files={
                "file": (
                    "June 2026.xlsx",
                    unresolved_unitwise_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert preview.status_code == 200
        body = preview.json()
        assert body["matched_rows"] == 1
        assert body["rows"][0]["person_id"] == sujil_id
        assert body["rows"][0]["unit_id"] == main_id
        assert body["rows"][0]["match_method"] == "manual_override"
        assert body["rows"][0]["unit_match_method"] == "unit_manual_override"

        applied = client.post(
            "/api/v1/unit-management/import-apply?month=2026-06&replace_existing=false",
            headers=headers,
            data={"resolutions_json": json.dumps(resolutions)},
            files={
                "file": (
                    "June 2026.xlsx",
                    unresolved_unitwise_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert applied.status_code == 200
        assert applied.json()["created_rows"] == 1


def test_unit_management_import_correction_can_force_special_posting_and_skip_row() -> None:
    with unit_client() as client:
        token = sign_in(client)
        headers = auth_headers(token)

        first_preview = client.post(
            "/api/v1/unit-management/import-preview?month=2026-06&replace_existing=true",
            headers=headers,
            files={
                "file": (
                    "June 2026.xlsx",
                    numbered_units_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert first_preview.status_code == 200
        rows = first_preview.json()["rows"]
        resolutions = {
            rows[0]["row_key"]: {"posting_type": "PAIN"},
            rows[1]["row_key"]: {"skip": True},
        }

        corrected = client.post(
            "/api/v1/unit-management/import-preview?month=2026-06&replace_existing=true",
            headers=headers,
            data={"resolutions_json": json.dumps(resolutions)},
            files={
                "file": (
                    "June 2026.xlsx",
                    numbered_units_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert corrected.status_code == 200
        body = corrected.json()
        forced = next(row for row in body["rows"] if row["row_key"] == rows[0]["row_key"])
        skipped = next(row for row in body["rows"] if row["row_key"] == rows[1]["row_key"])
        assert forced["posting_type"] == "PAIN"
        assert forced["special_posting"] is True
        assert forced["unit_id"] is None
        assert skipped["skip"] is True
        assert skipped["auto_assignable"] is False
        assert "Row excluded by reviewer" in skipped["issues"]


def split_month_unitwise_workbook_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "June"
    worksheet["A1"] = ""
    worksheet["B1"] = "Main OT"
    worksheet["C1"] = "Cardiac OT"
    worksheet["A2"] = "3rd call SR/APs"
    worksheet["B2"] = "Dr Sujil (June 1-15)"
    worksheet["C2"] = "Dr Sujil (June 16-30)"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_unit_management_import_allows_non_overlapping_split_month_unit_assignments() -> None:
    with unit_client() as client:
        token = sign_in(client)
        headers = auth_headers(token)

        preview = client.post(
            "/api/v1/unit-management/import-preview?month=2026-06",
            headers=headers,
            files={
                "file": (
                    "June 2026.xlsx",
                    split_month_unitwise_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert preview.status_code == 200
        body = preview.json()
        assert body["matched_rows"] == 2
        assert [row["posting_type"] for row in body["rows"]] == ["3RD_CALL", "3RD_CALL"]
        assert [(row["starts_on"], row["ends_on"]) for row in body["rows"]] == [
            ("2026-06-01", "2026-06-15"),
            ("2026-06-16", "2026-06-30"),
        ]
        assert all("overlapping dates" not in " ".join(row["issues"]) for row in body["rows"])

        applied = client.post(
            "/api/v1/unit-management/import-apply?month=2026-06&replace_existing=true",
            headers=headers,
            files={
                "file": (
                    "June 2026.xlsx",
                    split_month_unitwise_workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert applied.status_code == 200
        assert applied.json()["created_rows"] == 2

        month = client.get("/api/v1/unit-management/month?month=2026-06", headers=headers).json()
        assert [(row["starts_on"], row["ends_on"]) for row in month["assignments"]] == [
            ("2026-06-01", "2026-06-15"),
            ("2026-06-16", "2026-06-30"),
        ]


def test_unit_management_import_preview_returns_parser_warnings_without_crashing() -> None:
    with unit_client() as client:
        token = sign_in(client)
        headers = auth_headers(token)

        preview = client.post(
            "/api/v1/unit-management/import-preview?month=2026-06",
            headers=headers,
            files={"file": ("unitwise.txt", b"Unknown heading without unit context\n", "text/plain")},
        )

        assert preview.status_code == 200
        body = preview.json()
        assert body["matched_rows"] == 0
        assert body["parser_warnings"]


def test_unitwise_import_normalizes_real_call_and_special_posting_labels() -> None:
    examples = {
        "3rd call SR/APs": "3RD_CALL",
        "DM/PDF": "3RD_CALL",
        "3rd calls/Paeds call (DM/PDF)": "3RD_CALL",
        "2nd call SRs": "2ND_CALL",
        "2nd call PGs (2024)": "2ND_CALL",
        "2022 3rd calls": "3RD_CALL",
        "1st call PGs (2025)": "1ST_CALL",
        "Pain Call": "PAIN",
        "PAC Main campus": "PAC",
        "ICU POSTINGS": "SICU",
        "DRP (April - June 2026)": "DRP",
    }

    assert {label: normalize_import_posting(label) for label in examples} == examples


def test_unitwise_import_keeps_pain_campus_rows_under_pain_call() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = ""
    worksheet["B1"] = "May 1-15"
    worksheet["C1"] = "16-31"
    worksheet["A2"] = "Pain call"
    worksheet["A3"] = "Main campus"
    worksheet["B3"] = "Mizpah, Jona Samraj"
    worksheet["C3"] = "Shanmuga, Sampanna"
    worksheet["A4"] = "Ranipet"
    worksheet["B4"] = "Shanmuga, Sampanna"
    worksheet["C4"] = "Mizpah, Jona Samraj"
    worksheet["A5"] = "2025 first calls"
    worksheet["B5"] = "Sujil, Jeenu Ann Jose"
    output = BytesIO()
    workbook.save(output)

    parsed = parse_unitwise_excel_upload("May 2026.xlsx", output.getvalue())

    assert {posting.posting_label for posting in parsed.postings} == {"Pain call"}
    assert {normalize_import_posting(posting.posting_label) for posting in parsed.postings} == {"PAIN"}
    child_rows = [posting for posting in parsed.postings if posting.child_posting_label == "2025 first calls"]
    assert [posting.person_name for posting in child_rows] == ["Sujil", "Jeenu Ann Jose"]
