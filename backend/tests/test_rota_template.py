from collections.abc import Generator
from contextlib import contextmanager
from datetime import date, datetime, time, timedelta
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
import xlsxwriter
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app import models  # noqa: F401
from app.db.session import Base, get_db
from app.main import app
from app.models import DutyAssignment, DutySlot, LeaveRequest, Person, PersonPosting, Unit, UnitCallMinimum
from app.services.auth import seed_superadmin
from app.services.rota_setup import monthly_setup, update_scope_units
from app.services.rota_rules import default_phase_one_rules
from app.services.rota_template import (
    TemplateGenerationOptions,
    allocation_statistics,
    call_wise_template_export,
    clear_template_cache,
    generate_empty_template,
    template_month,
    write_call_wise_template_export,
    write_eagle_eye_matrix,
)


def seed_template_context(session: Session) -> tuple[Unit, Unit, Person]:
    included = Unit(code="UNIT_I", name="Unit I", campus="MAIN")
    excluded = Unit(code="UNIT_II", name="Unit II", campus="MAIN")
    person = Person(canonical_name="Template Member", call_level="1ST_CALL")
    cover_one = Person(canonical_name="Template Cover One", call_level="1ST_CALL")
    cover_two = Person(canonical_name="Template Cover Two", call_level="1ST_CALL")
    session.add_all([included, excluded, person, cover_one, cover_two])
    session.flush()
    session.add_all(
        [
            PersonPosting(
                person=person,
                unit=included,
                posting_type="1ST_CALL",
                starts_on=date(2026, 5, 1),
                source="unit_board",
            ),
            PersonPosting(
                person=cover_one,
                unit=included,
                posting_type="1ST_CALL",
                starts_on=date(2026, 5, 1),
                source="unit_board",
            ),
            PersonPosting(
                person=cover_two,
                unit=included,
                posting_type="1ST_CALL",
                starts_on=date(2026, 5, 1),
                source="unit_board",
            ),
        ]
    )
    session.add(
        LeaveRequest(
            person=person,
            leave_type="ANNUAL_LEAVE",
            leave_slot="FULL_DAY",
            starts_on=date(2026, 5, 1),
            ends_on=date(2026, 5, 1),
            status="approved",
        )
    )
    session.commit()
    return included, excluded, person


def lock_month_scope(session: Session, included: Unit, excluded: Unit) -> None:
    _period, scope = monthly_setup(session, "2026-05")
    update_scope_units(
        session,
        scope,
        [included.id],
        [excluded.id],
        True,
        True,
        "Ready for template generation",
    )


def test_generate_empty_template_forces_hard_blocked_slots_for_review() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        included, excluded, _person = seed_template_context(session)
        lock_month_scope(session, included, excluded)

        result = generate_empty_template(
            session,
            "2026-05",
            TemplateGenerationOptions(
                duty_keys=["MAIN_1ST_24HR", "MAIN_PAC_PG"],
                starts_on=date(2026, 5, 1),
                ends_on=date(2026, 5, 1),
            ),
        )

        slots = session.query(DutySlot).all()
        assert result["latest_run"]["created_slots"] == 2
        assert result["latest_run"]["needs_review_slots"] == 2
        assert result["latest_run"]["summary"]["unresolved_slots"] == 0
        assert result["latest_run"]["summary"]["forced_review_slots"] == 2
        assert result["latest_run"]["blocked_slots"] == 2
        assert len(slots) == 2
        assert {slot.unit_id for slot in slots} == {included.id}
        assert {slot.duty_type for slot in slots} == {"MAIN_1ST_24HR", "MAIN_PAC_PG"}
        assert {slot.template_status for slot in slots} == {"needs_review"}
        assert all(str(slot.template_reason).startswith("Forced minimal-damage allocation") for slot in slots)
        assert excluded.id not in {slot.unit_id for slot in slots}


def test_generate_empty_template_requires_locked_scope() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        included, _excluded, _person = seed_template_context(session)
        _period, scope = monthly_setup(session, "2026-05")
        update_scope_units(session, scope, [included.id], [], True, False)

        try:
            generate_empty_template(
                session,
                "2026-05",
                TemplateGenerationOptions(duty_keys=["MAIN_1ST_24HR"]),
            )
        except ValueError as exc:
            assert "scope must be locked" in str(exc)
        else:
            raise AssertionError("Unlocked generation scope should block template generation")


def test_generate_empty_template_balances_duties_across_units() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        unit_a = Unit(code="UNIT_A", name="Unit A", campus="MAIN", minimum_free_people=0)
        unit_b = Unit(code="UNIT_B", name="Unit B", campus="MAIN", minimum_free_people=0)
        people = [
            *[
                Person(canonical_name=f"A First {index}", call_level="1ST_CALL")
                for index in range(3)
            ],
            *[
                Person(canonical_name=f"B First {index}", call_level="1ST_CALL")
                for index in range(3)
            ],
        ]
        session.add_all([unit_a, unit_b, *people])
        session.flush()
        for person in people[:3]:
            session.add(
                PersonPosting(
                    person=person,
                    unit=unit_a,
                    posting_type="1ST_CALL",
                    starts_on=date(2026, 5, 1),
                    source="unit_board",
                )
            )
        for person in people[3:]:
            session.add(
                PersonPosting(
                    person=person,
                    unit=unit_b,
                    posting_type="1ST_CALL",
                    starts_on=date(2026, 5, 1),
                    source="unit_board",
                )
            )
        session.commit()
        _period, scope = monthly_setup(session, "2026-05")
        update_scope_units(session, scope, [unit_a.id, unit_b.id], [], True, True, "Balance test")

        result = generate_empty_template(
            session,
            "2026-05",
            TemplateGenerationOptions(
                duty_keys=["MAIN_1ST_24HR"],
                starts_on=date(2026, 5, 1),
                ends_on=date(2026, 5, 4),
            ),
        )

        counts = {
            unit.code: session.query(DutySlot).filter(DutySlot.unit_id == unit.id).count()
            for unit in [unit_a, unit_b]
        }
        assert result["latest_run"]["created_slots"] == 4
        assert counts == {"UNIT_A": 2, "UNIT_B": 2}


def test_generate_empty_template_balances_saturdays_and_sundays_across_units() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        units = [
            Unit(code="UNIT_A", name="Unit A", campus="MAIN", minimum_free_people=0),
            Unit(code="UNIT_B", name="Unit B", campus="MAIN", minimum_free_people=0),
            Unit(code="UNIT_C", name="Unit C", campus="MAIN", minimum_free_people=0),
        ]
        people = [
            Person(canonical_name=f"{unit_code} First {index}", call_level="1ST_CALL")
            for unit_code in ["A", "B", "C"]
            for index in range(3)
        ]
        session.add_all([*units, *people])
        session.flush()
        for unit_index, unit in enumerate(units):
            for person in people[unit_index * 3 : unit_index * 3 + 3]:
                session.add(
                    PersonPosting(
                        person=person,
                        unit=unit,
                        posting_type="1ST_CALL",
                        starts_on=date(2026, 5, 1),
                        source="unit_board",
                    )
                )
        session.commit()
        _period, scope = monthly_setup(session, "2026-05")
        update_scope_units(
            session,
            scope,
            [unit.id for unit in units],
            [],
            True,
            True,
            "Weekend balance test",
        )

        result = generate_empty_template(
            session,
            "2026-05",
            TemplateGenerationOptions(
                duty_keys=["MAIN_1ST_24HR"],
                starts_on=date(2026, 5, 2),
                ends_on=date(2026, 5, 17),
                include_weekdays=False,
                include_weekends=True,
            ),
        )

        saturday_counts = {
            unit.code: session.query(DutySlot)
            .filter(
                DutySlot.unit_id == unit.id,
                DutySlot.duty_date.in_(
                    [date(2026, 5, 2), date(2026, 5, 9), date(2026, 5, 16)]
                ),
            )
            .count()
            for unit in units
        }
        sunday_counts = {
            unit.code: session.query(DutySlot)
            .filter(
                DutySlot.unit_id == unit.id,
                DutySlot.duty_date.in_(
                    [date(2026, 5, 3), date(2026, 5, 10), date(2026, 5, 17)]
                ),
            )
            .count()
            for unit in units
        }

        assert result["latest_run"]["created_slots"] == 6
        assert saturday_counts == {"UNIT_A": 1, "UNIT_B": 1, "UNIT_C": 1}
        assert sunday_counts == {"UNIT_A": 1, "UNIT_B": 1, "UNIT_C": 1}


def test_generate_empty_template_counts_previous_24hr_duty_as_next_day_unavailable() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        unit_a = Unit(code="UNIT_A", name="Unit A", campus="MAIN", minimum_free_people=0)
        unit_b = Unit(code="UNIT_B", name="Unit B", campus="MAIN", minimum_free_people=0)
        people = [
            Person(canonical_name=f"A First {index}", call_level="1ST_CALL")
            for index in range(3)
        ] + [
            Person(canonical_name=f"B First {index}", call_level="1ST_CALL")
            for index in range(3)
        ]
        session.add_all([unit_a, unit_b, *people])
        session.flush()
        for person in people[:3]:
            session.add(
                PersonPosting(
                    person=person,
                    unit=unit_a,
                    posting_type="1ST_CALL",
                    starts_on=date(2026, 5, 1),
                    source="unit_board",
                )
            )
        for person in people[3:]:
            session.add(
                PersonPosting(
                    person=person,
                    unit=unit_b,
                    posting_type="1ST_CALL",
                    starts_on=date(2026, 5, 1),
                    source="unit_board",
                )
            )
        session.commit()
        _period, scope = monthly_setup(session, "2026-05")
        update_scope_units(
            session,
            scope,
            [unit_a.id, unit_b.id],
            [],
            True,
            True,
            "Post duty balance test",
        )

        generate_empty_template(
            session,
            "2026-05",
            TemplateGenerationOptions(
                duty_keys=["MAIN_1ST_24HR"],
                starts_on=date(2026, 5, 1),
                ends_on=date(2026, 5, 2),
            ),
        )

        slots = session.query(DutySlot).order_by(DutySlot.duty_date).all()
        may_1, may_2 = slots

        assert may_1.unit_id == unit_a.id
        assert may_2.unit_id == unit_b.id
        assert may_2.template_status == "needs_review"


def test_generate_empty_template_forces_and_balances_when_all_units_hard_blocked() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        units = [
            Unit(code="UNIT_A", name="Unit A", campus="MAIN", minimum_free_people=1),
            Unit(code="UNIT_B", name="Unit B", campus="MAIN", minimum_free_people=1),
        ]
        people = [
            Person(canonical_name="A First", call_level="1ST_CALL"),
            Person(canonical_name="B First", call_level="1ST_CALL"),
        ]
        session.add_all([*units, *people])
        session.flush()
        for unit, person in zip(units, people, strict=True):
            session.add(
                PersonPosting(
                    person=person,
                    unit=unit,
                    posting_type="1ST_CALL",
                    starts_on=date(2026, 5, 1),
                    source="unit_board",
                )
            )
        session.commit()
        _period, scope = monthly_setup(session, "2026-05")
        update_scope_units(
            session,
            scope,
            [unit.id for unit in units],
            [],
            True,
            True,
            "Unresolved mandatory slot test",
        )

        result = generate_empty_template(
            session,
            "2026-05",
                TemplateGenerationOptions(
                    duty_keys=["MAIN_1ST_24HR"],
                    starts_on=date(2026, 5, 1),
                    ends_on=date(2026, 5, 2),
                ),
            )

        slots = session.query(DutySlot).order_by(DutySlot.duty_date).all()

        assert result["latest_run"]["created_slots"] == 2
        assert result["latest_run"]["summary"]["unresolved_slots"] == 0
        assert result["latest_run"]["summary"]["forced_review_slots"] == 2
        assert result["latest_run"]["blocked_slots"] == 4
        assert [slot.unit_id for slot in slots] == [units[0].id, units[1].id]
        assert {slot.slot_label for slot in slots} == {"UNIT_A:primary", "UNIT_B:primary"}
        assert {slot.template_status for slot in slots} == {"needs_review"}
        assert all("least damaging and fairest unit" in str(slot.template_reason) for slot in slots)


def test_allocation_statistics_summarizes_units_dates_and_blocked_events() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        units = [
            Unit(code="UNIT_A", name="Unit A", campus="MAIN", minimum_free_people=0),
            Unit(code="UNIT_B", name="Unit B", campus="MAIN", minimum_free_people=0),
        ]
        people = [
            Person(canonical_name=f"A First {index}", call_level="1ST_CALL")
            for index in range(3)
        ] + [
            Person(canonical_name=f"B First {index}", call_level="1ST_CALL")
            for index in range(3)
        ]
        session.add_all([*units, *people])
        session.flush()
        for person in people[:3]:
            session.add(
                PersonPosting(
                    person=person,
                    unit=units[0],
                    posting_type="1ST_CALL",
                    starts_on=date(2026, 5, 1),
                    source="unit_board",
                )
            )
        for person in people[3:]:
            session.add(
                PersonPosting(
                    person=person,
                    unit=units[1],
                    posting_type="1ST_CALL",
                    starts_on=date(2026, 5, 1),
                    source="unit_board",
                )
            )
        session.commit()
        _period, scope = monthly_setup(session, "2026-05")
        update_scope_units(
            session,
            scope,
            [unit.id for unit in units],
            [],
            True,
            True,
            "Statistics test",
        )

        generate_empty_template(
            session,
            "2026-05",
            TemplateGenerationOptions(
                duty_keys=["MAIN_1ST_24HR"],
                starts_on=date(2026, 5, 2),
                ends_on=date(2026, 5, 3),
            ),
        )

        stats = allocation_statistics(session, "2026-05")

        assert stats["summary"]["total_slots"] == 2
        assert stats["summary"]["included_units"] == 2
        assert stats["unit_tallies"][0]["weekend_slots"] == 1
        assert stats["unit_tallies"][1]["weekend_slots"] == 1
        assert [row["total_slots"] for row in stats["date_distribution"]] == [1, 1]
        matrix = {row["unit_name"]: row["counts"]["MAIN_1ST_24HR"] for row in stats["unit_duty_matrix"]}
        assert matrix == {"Unit A": 1, "Unit B": 1, "Unresolved": 0}
        assert stats["blocked_or_skipped_events"]


def test_cluster_specific_duties_show_cluster_suffix_in_unit_assignment_label() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        unit = Unit(code="UNIT_I", name="Unit I", campus="MAIN", minimum_free_people=0)
        people = [
            Person(canonical_name=f"Third Call {index}", call_level="3RD_CALL")
            for index in range(8)
        ]
        session.add_all([unit, *people])
        session.flush()
        for person in people:
            session.add(
                PersonPosting(
                    person=person,
                    unit=unit,
                    posting_type="3RD_CALL",
                    starts_on=date(2026, 5, 1),
                    source="unit_board",
                )
            )
        session.commit()
        _period, scope = monthly_setup(session, "2026-05")
        update_scope_units(session, scope, [unit.id], [], True, True, "Cluster suffix test")

        result = generate_empty_template(
            session,
            "2026-05",
            TemplateGenerationOptions(
                duty_keys=["RC_3RD_CALL_A", "RC_3RD_CALL_B", "MAIN_PAC_PG"],
                starts_on=date(2026, 5, 2),
                ends_on=date(2026, 5, 2),
            ),
        )

        labels = {
            slot["duty_type"]: slot["unit_assignment_label"]
            for slot in result["slots"]
        }

        assert labels["RC_3RD_CALL_A"] == "Unit 1 (3A)"
        assert labels["RC_3RD_CALL_B"] == "Unit 1 (3B)"
        assert labels["MAIN_PAC_PG"] == "Unit 1 (3C)"

        filename, payload = call_wise_template_export(session, "2026-05")
        workbook = load_workbook(BytesIO(payload))

        assert filename == "call-wise-rota-template-2026-05.xlsx"
        sheet_values = [
            workbook["3rd Call"].cell(row=row, column=2).value
            for row in range(2, 5)
        ]
        assert "Unit 1 (3A)" in sheet_values
        assert "Unit 1 (3B)" in sheet_values
        assert "Unit 1 (3C)" in sheet_values


def test_generate_empty_template_uses_unit_minimum_free_people() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        strict = Unit(code="STRICT", name="Strict Unit", campus="MAIN", minimum_free_people=1)
        flexible = Unit(code="FLEX", name="Flexible Unit", campus="MAIN", minimum_free_people=0)
        strict_person = Person(canonical_name="Strict First", call_level="1ST_CALL")
        flexible_people = [
            Person(canonical_name=f"Flexible First {index}", call_level="1ST_CALL")
            for index in range(3)
        ]
        session.add_all([strict, flexible, strict_person, *flexible_people])
        session.flush()
        session.add_all(
            [
                PersonPosting(person=strict_person, unit=strict, posting_type="1ST_CALL", starts_on=date(2026, 5, 1), source="unit_board"),
                *[
                    PersonPosting(
                        person=person,
                        unit=flexible,
                        posting_type="1ST_CALL",
                        starts_on=date(2026, 5, 1),
                        source="unit_board",
                    )
                    for person in flexible_people
                ],
            ]
        )
        session.commit()
        _period, scope = monthly_setup(session, "2026-05")
        update_scope_units(session, scope, [strict.id, flexible.id], [], True, True, "Minimum free test")

        generate_empty_template(
            session,
            "2026-05",
            TemplateGenerationOptions(
                duty_keys=["MAIN_1ST_24HR"],
                starts_on=date(2026, 5, 2),
                ends_on=date(2026, 5, 2),
            ),
        )

        slot = session.query(DutySlot).filter(DutySlot.source == "phase4_template").one()
        assert slot.unit_id == flexible.id


def test_generate_empty_template_uses_call_wise_unit_minimum_free_people() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        strict = Unit(code="STRICT", name="Strict Unit", campus="MAIN", minimum_free_people=0)
        flexible = Unit(code="FLEX", name="Flexible Unit", campus="MAIN", minimum_free_people=0)
        strict_first = Person(canonical_name="Strict First", call_level="1ST_CALL")
        strict_third = Person(canonical_name="Strict Third", call_level="3RD_CALL")
        flexible_firsts = [
            Person(canonical_name=f"Flexible First {index}", call_level="1ST_CALL")
            for index in range(3)
        ]
        session.add_all([strict, flexible, strict_first, strict_third, *flexible_firsts])
        session.flush()
        session.add_all(
            [
                PersonPosting(person=strict_first, unit=strict, posting_type="1ST_CALL", starts_on=date(2026, 5, 1), source="unit_board"),
                PersonPosting(person=strict_third, unit=strict, posting_type="3RD_CALL", starts_on=date(2026, 5, 1), source="unit_board"),
                *[
                    PersonPosting(
                        person=person,
                        unit=flexible,
                        posting_type="1ST_CALL",
                        starts_on=date(2026, 5, 1),
                        source="unit_board",
                    )
                    for person in flexible_firsts
                ],
                UnitCallMinimum(unit=strict, call_level="1ST_CALL", minimum_free_people=1),
            ]
        )
        session.commit()
        _period, scope = monthly_setup(session, "2026-05")
        update_scope_units(session, scope, [strict.id, flexible.id], [], True, True, "Call minimum free test")

        generate_empty_template(
            session,
            "2026-05",
            TemplateGenerationOptions(
                duty_keys=["MAIN_1ST_24HR"],
                starts_on=date(2026, 5, 2),
                ends_on=date(2026, 5, 2),
            ),
        )

        slot = session.query(DutySlot).filter(DutySlot.source == "phase4_template").one()
        assert slot.unit_id == flexible.id


def test_template_month_excludes_historical_import_slots() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        included, excluded, _person = seed_template_context(session)
        lock_month_scope(session, included, excluded)
        period, _scope = monthly_setup(session, "2026-05")
        starts_at = datetime.combine(date(2026, 5, 1), time(hour=8))
        session.add(
            DutySlot(
                rota_period=period,
                unit=None,
                duty_date=date(2026, 5, 1),
                duty_type="HISTORICAL_IMPORTED_CALL",
                slot_label="legacy:row",
                starts_at=starts_at,
                ends_at=starts_at + timedelta(hours=24),
                is_24hr=True,
                source="historical_analysis_import",
            )
        )
        generate_empty_template(
            session,
            "2026-05",
            TemplateGenerationOptions(
                duty_keys=["MAIN_1ST_24HR"],
                starts_on=date(2026, 5, 2),
                ends_on=date(2026, 5, 2),
            ),
        )

        result = template_month(session, "2026-05")

        assert result["summary"]["total_slots"] == 1
        assert [slot["unit_code"] for slot in result["slots"]] == ["UNIT_I"]


def test_clear_template_cache_removes_generated_slots_and_runs_only() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        included, excluded, _person = seed_template_context(session)
        lock_month_scope(session, included, excluded)
        period, _scope = monthly_setup(session, "2026-05")
        starts_at = datetime.combine(date(2026, 5, 1), time(hour=8))
        session.add(
            DutySlot(
                rota_period=period,
                unit=None,
                duty_date=date(2026, 5, 1),
                duty_type="HISTORICAL_IMPORTED_CALL",
                slot_label="legacy:row",
                starts_at=starts_at,
                ends_at=starts_at + timedelta(hours=24),
                is_24hr=True,
                source="historical_analysis_import",
            )
        )
        generate_empty_template(
            session,
            "2026-05",
            TemplateGenerationOptions(
                duty_keys=["MAIN_1ST_24HR"],
                starts_on=date(2026, 5, 2),
                ends_on=date(2026, 5, 2),
            ),
        )

        result = clear_template_cache(session, "2026-05")
        remaining_slots = session.query(DutySlot).all()

        assert result["cleared_slots"] == 1
        assert result["cleared_runs"] == 1
        assert len(remaining_slots) == 1
        assert remaining_slots[0].source == "historical_analysis_import"


def test_clear_template_cache_blocks_assigned_generated_slots() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        included, excluded, person = seed_template_context(session)
        lock_month_scope(session, included, excluded)
        generate_empty_template(
            session,
            "2026-05",
            TemplateGenerationOptions(
                duty_keys=["MAIN_1ST_24HR"],
                starts_on=date(2026, 5, 1),
                ends_on=date(2026, 5, 1),
            ),
        )
        slot = session.query(DutySlot).filter(DutySlot.source == "phase4_template").one()
        session.add(DutyAssignment(duty_slot=slot, person=person, source="manual_rota_board"))
        session.commit()

        try:
            clear_template_cache(session, "2026-05")
        except ValueError as exc:
            assert "have assignments" in str(exc)
        else:
            raise AssertionError("Assigned template slots should block cache clearing")


def test_clear_template_cache_can_remove_assignments_when_explicitly_requested() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        included, excluded, person = seed_template_context(session)
        lock_month_scope(session, included, excluded)
        generate_empty_template(
            session,
            "2026-05",
            TemplateGenerationOptions(
                duty_keys=["MAIN_1ST_24HR"],
                starts_on=date(2026, 5, 1),
                ends_on=date(2026, 5, 1),
            ),
        )
        slot = session.query(DutySlot).filter(DutySlot.source == "phase4_template").one()
        session.add(DutyAssignment(duty_slot=slot, person=person, source="manual_rota_board"))
        session.commit()

        result = clear_template_cache(session, "2026-05", clear_assignments=True)

        assert result["cleared_slots"] == 1
        assert result["cleared_assignments"] == 1
        assert session.query(DutyAssignment).count() == 0
        assert session.query(DutySlot).filter(DutySlot.source == "phase4_template").count() == 0


@contextmanager
def template_client() -> Generator[TestClient, None, None]:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    testing_session = sessionmaker(bind=engine)

    with testing_session() as seed_session:
        seed_superadmin(seed_session)
        included, excluded, _person = seed_template_context(seed_session)
        lock_month_scope(seed_session, included, excluded)

    def override_get_db() -> Generator[Session, None, None]:
        with testing_session() as session:
            yield session

    app.dependency_overrides[get_db] = override_get_db
    try:
        yield TestClient(app)
    finally:
        app.dependency_overrides.clear()


def sign_in(client: TestClient) -> str:
    response = client.post(
        "/api/v1/auth/sign-in",
        json={"username": "rotachief", "password": "rotateam"},
    )
    assert response.status_code == 200
    return response.json()["token"]


def test_rota_template_api_generates_and_returns_template_month() -> None:
    with template_client() as client:
        token = sign_in(client)
        response = client.post(
            "/api/v1/rota-template/generate?month=2026-05",
            headers={"Authorization": f"Bearer {token}"},
            json={
                "duty_keys": ["MAIN_1ST_24HR", "MAIN_PAC_PG"],
                "starts_on": "2026-05-01",
                "ends_on": "2026-05-01",
            },
        )

        assert response.status_code == 200
        payload = response.json()
        assert payload["latest_run"]["created_slots"] == 2
        assert payload["latest_run"]["blocked_slots"] == 2
        assert payload["summary"]["status_counts"]["needs_review"] == 2
        assert payload["summary"]["forced_review_slots"] == 2
        assert payload["summary"]["status_counts"].get("unresolved", 0) == 0
        assert {slot["unit_code"] for slot in payload["slots"]} == {"UNIT_I"}
        assert all(slot["is_forced_allocation"] for slot in payload["slots"])


def test_rota_template_api_returns_allocation_statistics() -> None:
    with template_client() as client:
        token = sign_in(client)
        headers = {"Authorization": f"Bearer {token}"}
        client.post(
            "/api/v1/rota-template/generate?month=2026-05",
            headers=headers,
            json={
                "duty_keys": ["MAIN_1ST_24HR"],
                "starts_on": "2026-05-02",
                "ends_on": "2026-05-02",
            },
        )

        response = client.get(
            "/api/v1/rota-template/allocation-statistics?month=2026-05",
            headers=headers,
        )

        assert response.status_code == 200
        payload = response.json()
        assert payload["summary"]["total_slots"] == 1
        assert payload["unit_tallies"]
        assert payload["unit_duty_matrix"]
        assert payload["date_distribution"][0]["date"] == "2026-05-02"


def test_rota_template_api_clears_cache() -> None:
    with template_client() as client:
        token = sign_in(client)
        headers = {"Authorization": f"Bearer {token}"}
        client.post(
            "/api/v1/rota-template/generate?month=2026-05",
            headers=headers,
            json={
                "duty_keys": ["MAIN_1ST_24HR"],
                "starts_on": "2026-05-01",
                "ends_on": "2026-05-01",
            },
        )

        response = client.delete("/api/v1/rota-template/cache?month=2026-05", headers=headers)

        assert response.status_code == 200
        assert response.json()["cleared_slots"] == 1
        month = client.get("/api/v1/rota-template/month?month=2026-05", headers=headers).json()
        assert month["summary"]["total_slots"] == 0


def test_rota_template_api_clears_cache_with_assignments_when_requested() -> None:
    with template_client() as client:
        token = sign_in(client)
        headers = {"Authorization": f"Bearer {token}"}
        generated = client.post(
            "/api/v1/rota-template/generate?month=2026-05",
            headers=headers,
            json={
                "duty_keys": ["MAIN_1ST_24HR"],
                "starts_on": "2026-05-01",
                "ends_on": "2026-05-02",
            },
        ).json()
        slot_id = generated["slots"][0]["id"]
        person_id = client.get("/api/v1/admin/members", headers=headers).json()[0]["id"]
        assigned = client.post(
            f"/api/v1/rota-assignments/slots/{slot_id}/assign",
            headers=headers,
            json={"person_id": person_id, "override_reason": "Test setup assignment before clear"},
        )
        assert assigned.status_code == 200

        blocked = client.delete("/api/v1/rota-template/cache?month=2026-05", headers=headers)
        assert blocked.status_code == 400

        cleared = client.delete(
            "/api/v1/rota-template/cache?month=2026-05&clear_assignments=true",
            headers=headers,
        )
        assert cleared.status_code == 200
        assert cleared.json()["cleared_assignments"] == 1


def test_rota_template_eagle_eye_export_downloads_workbook() -> None:
    with template_client() as client:
        token = sign_in(client)
        headers = {"Authorization": f"Bearer {token}"}
        client.post(
            "/api/v1/rota-template/generate?month=2026-05",
            headers=headers,
            json={
                "duty_keys": ["MAIN_1ST_24HR", "RC_12HR", "PB_SHIFT"],
                "starts_on": "2026-05-01",
                "ends_on": "2026-05-02",
            },
        )

        response = client.get("/api/v1/rota-template/eagle-eye-export?month=2026-05", headers=headers)

        assert response.status_code == 200
        assert response.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "eagle-eye-rota-template-2026-05.xlsx" in response.headers["content-disposition"]
        assert response.content.startswith(b"PK")
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook["Eagle Eye"]
        assert sheet.cell(row=1, column=1).value == "Duty"
        assert sheet.cell(row=1, column=2).value == "2026-05-01\nFriday"
        assert sheet.cell(row=1, column=3).value == "2026-05-02\nSaturday"
        assert sheet.cell(row=2, column=1).value == "MAIN CALLS"
        assert sheet.cell(row=3, column=1).value == "Main 1st Call"
        assert sheet.cell(row=2, column=1).fill.fgColor.rgb in {"FFD9EAF7", "00D9EAF7"}
        assert sheet.cell(row=1, column=3).fill.fgColor.rgb in {"FFFFE699", "00FFE699"}


def test_rota_template_call_wise_export_downloads_workbook() -> None:
    with template_client() as client:
        token = sign_in(client)
        headers = {"Authorization": f"Bearer {token}"}
        client.post(
            "/api/v1/rota-template/generate?month=2026-05",
            headers=headers,
            json={
                "duty_keys": ["MAIN_1ST_24HR"],
                "starts_on": "2026-05-01",
                "ends_on": "2026-05-02",
            },
        )

        response = client.get("/api/v1/rota-template/call-wise-export?month=2026-05", headers=headers)

        assert response.status_code == 200
        assert response.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "call-wise-rota-template-2026-05.xlsx" in response.headers["content-disposition"]
        assert response.content.startswith(b"PK")
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook["1st Call"]
        assert sheet.cell(row=1, column=1).value == "Duty"
        assert sheet.cell(row=1, column=2).value == "2026-05-01\nFriday"
        assert sheet.cell(row=1, column=3).value == "2026-05-02\nSaturday"
        assert sheet.cell(row=2, column=1).value == "Main 1st Call"
        assert sheet.cell(row=2, column=2).value == "Unit 1"
        assert sheet.cell(row=2, column=3).value == "Unit 1"
        assert sheet.cell(row=2, column=3).value == "Unit 1"
        assert sheet.cell(row=1, column=3).fill.fgColor.rgb in {"FFFFE699", "00FFE699"}


def test_eagle_eye_export_groups_duties_with_divider_rows() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        unit = Unit(code="UNIT_I", name="Unit I", campus="MAIN")
        period, _scope = monthly_setup(session, "2026-05")
        starts_at = datetime.combine(date(2026, 5, 2), time(hour=8))
        slots = [
            DutySlot(
                rota_period=period,
                unit=unit,
                duty_date=date(2026, 5, 2),
                duty_type="MAIN_1ST_24HR",
                slot_label="unit:main",
                starts_at=starts_at,
                ends_at=starts_at + timedelta(hours=24),
                is_24hr=True,
                source="phase4_template",
            ),
            DutySlot(
                rota_period=period,
                unit=unit,
                duty_date=date(2026, 5, 2),
                duty_type="RC_12HR",
                slot_label="unit:rc",
                starts_at=starts_at,
                ends_at=starts_at + timedelta(hours=12),
                is_24hr=False,
                source="phase4_template",
            ),
            DutySlot(
                rota_period=period,
                unit=unit,
                duty_date=date(2026, 5, 2),
                duty_type="PB_SHIFT",
                slot_label="unit:shift",
                starts_at=starts_at,
                ends_at=starts_at + timedelta(hours=8),
                is_24hr=False,
                source="phase4_template",
            ),
        ]
        session.add_all([unit, *slots])
        session.flush()

        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {"in_memory": True})
        write_eagle_eye_matrix(workbook, slots, default_phase_one_rules())
        workbook.close()

    workbook = load_workbook(BytesIO(output.getvalue()))
    sheet = workbook["Eagle Eye"]
    assert sheet.cell(row=2, column=1).value == "MAIN CALLS"
    assert sheet.cell(row=3, column=1).value == "Main 1st Call"
    assert sheet.cell(row=4, column=1).value == "RC CALLS"
    assert sheet.cell(row=5, column=1).value == "RC 12hr"
    assert sheet.cell(row=6, column=1).value == "SHIFTS"
    assert sheet.cell(row=7, column=1).value == "PB Shift"
    assert sheet.cell(row=2, column=1).fill.fgColor.rgb in {"FFD9EAF7", "00D9EAF7"}
    assert sheet.cell(row=1, column=2).fill.fgColor.rgb in {"FFFFE699", "00FFE699"}
    assert sheet.cell(row=3, column=2).value == "Unit 1"


def test_call_wise_export_splits_slots_by_required_person_call_level() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        unit = Unit(code="UNIT_III", name="Unit III", campus="MAIN")
        period, _scope = monthly_setup(session, "2026-05")
        starts_at = datetime.combine(date(2026, 5, 2), time(hour=8))
        slots = [
            DutySlot(
                rota_period=period,
                unit=unit,
                duty_date=date(2026, 5, 2),
                duty_type="MAIN_1ST_24HR",
                slot_label="unit:main-1",
                starts_at=starts_at,
                ends_at=starts_at + timedelta(hours=24),
                is_24hr=True,
                source="phase4_template",
            ),
            DutySlot(
                rota_period=period,
                unit=unit,
                duty_date=date(2026, 5, 2),
                duty_type="MAIN_3RD_24HR",
                slot_label="unit:main-3",
                starts_at=starts_at,
                ends_at=starts_at + timedelta(hours=24),
                is_24hr=True,
                source="phase4_template",
            ),
        ]
        session.add_all([unit, *slots])
        session.flush()

        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {"in_memory": True})
        write_call_wise_template_export(workbook, slots, default_phase_one_rules())
        workbook.close()

    workbook = load_workbook(BytesIO(output.getvalue()))
    assert "1st Call" in workbook.sheetnames
    assert "3rd Call" in workbook.sheetnames
    assert workbook["1st Call"].cell(row=2, column=1).value == "Main 1st Call"
    assert workbook["3rd Call"].cell(row=2, column=1).value == "Main 3rd Call"
    assert workbook["3rd Call"].cell(row=2, column=2).value == "Unit 3"
    assert workbook["3rd Call"].cell(row=1, column=2).value == "2026-05-02\nSaturday"
    assert workbook["3rd Call"].cell(row=1, column=2).fill.fgColor.rgb in {"FFFFE699", "00FFE699"}
    assert workbook["3rd Call"].cell(row=2, column=2).fill.fgColor.rgb in {"FFFFF2CC", "00FFF2CC"}
