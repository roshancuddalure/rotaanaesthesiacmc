from pathlib import Path

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app import models  # noqa: F401
from app.db.session import Base
from app.models import DutyAssignment, Person, PersonDesignation
from app.services.department_roster_reset import extract_clean_roster_members, reset_department_members_from_roster


def make_roster(path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "ANAESTHESIA"
    workbook["ANAESTHESIA"].append(["ANAESTHESIA DEPARTMENT", "Dr. Asha Devanand", "TOTAL STAFF-1"])
    workbook.create_sheet("CARDIAC ANAESTHESIA")
    workbook["CARDIAC ANAESTHESIA"].append(["S.NO", "PROFESSORS-2", "EMP NO", "DESIGNATION"])
    workbook["CARDIAC ANAESTHESIA"].append([1, "Dr.Balaji.K", 31583, "PROFESSOR"])
    workbook.create_sheet("NEURO ANAESTHESIA")
    workbook["NEURO ANAESTHESIA"].append(["S.No", "Doctor Name", "Emp. No", "Designation"])
    workbook["NEURO ANAESTHESIA"].append([1, "Ramamani M", 50916, "Professor"])
    workbook.save(path)


def test_extract_clean_roster_members_uses_current_roster_sheets(tmp_path: Path) -> None:
    roster = tmp_path / "roster.xlsx"
    make_roster(roster)

    members = extract_clean_roster_members(roster)

    assert [member.name for member in members] == ["Asha Devanand", "Balaji K", "Ramamani M"]
    assert {member.call_position for member in members} == {
        "ANAESTHESIA ROSTER",
        "PROFESSOR",
    }


def test_reset_department_members_from_roster_clears_old_operational_data(tmp_path: Path) -> None:
    roster = tmp_path / "roster.xlsx"
    make_roster(roster)
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        old_person = Person(canonical_name="Old Duplicate")
        session.add(old_person)
        session.commit()

        result = reset_department_members_from_roster(session, roster)

        assert result.deleted_counts["persons"] == 1
        assert result.created_members == 3
        assert session.query(Person).count() == 3
        assert session.query(PersonDesignation).count() == 3
        assert session.query(DutyAssignment).count() == 0
