from collections.abc import Generator
from contextlib import contextmanager
from datetime import date, datetime, time, timedelta
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app import models  # noqa: F401
from app.db.session import Base, get_db
from app.main import app
from app.models import DutyAssignment, DutySlot, Person, PersonPosting, RotaPublishApproval, Unit
from app.services.auth import seed_superadmin
from app.services.rota_publish import final_rota_export, publish_rota_month, rota_publish_checklist
from app.services.rota_setup import monthly_setup


def duty_bounds(day: date) -> tuple[datetime, datetime]:
    starts_at = datetime.combine(day, time(hour=8))
    return starts_at, starts_at + timedelta(hours=24)


def seed_publish_context(session: Session, *, open_slot: bool = False) -> dict[str, str]:
    user = seed_superadmin(session)
    unit = Unit(code="UNIT_I", name="Unit I", campus="MAIN")
    assigned = Person(canonical_name="Assigned Member", call_level="1ST_CALL")
    cover_one = Person(canonical_name="Cover One", call_level="1ST_CALL")
    cover_two = Person(canonical_name="Cover Two", call_level="1ST_CALL")
    session.add_all([unit, assigned, cover_one, cover_two])
    session.flush()
    for person in [assigned, cover_one, cover_two]:
        session.add(
            PersonPosting(
                person=person,
                unit=unit,
                posting_type="1ST_CALL",
                starts_on=date(2026, 5, 1),
                source="unit_board",
            )
        )
    period, scope = monthly_setup(session, "2026-05")
    scope.is_locked = True
    starts_at, ends_at = duty_bounds(date(2026, 5, 2))
    slot = DutySlot(
        rota_period=period,
        unit=unit,
        duty_date=date(2026, 5, 2),
        duty_type="MAIN_1ST_24HR",
        call_level="1ST_CALL",
        slot_label="UNIT_I:primary",
        starts_at=starts_at,
        ends_at=ends_at,
        is_24hr=True,
        source="phase4_template",
    )
    session.add(slot)
    session.flush()
    if not open_slot:
        session.add(DutyAssignment(duty_slot=slot, person=assigned, source="safe_auto_fill_draft"))
    session.commit()
    return {"user_id": str(user.id)}


def test_publish_checklist_blocks_open_slots() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        seed_publish_context(session, open_slot=True)
        checklist = rota_publish_checklist(session, "2026-05")

        assert checklist["can_publish"] is False
        assert any(item["title"] == "Open slots" for item in checklist["blockers"])


def test_publish_records_approval_and_exports_workbook() -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        seed_publish_context(session)
        user = seed_superadmin(session)
        result = publish_rota_month(
            session,
            month="2026-05",
            approved_by=user,
            confirm_warnings=False,
            approval_note="Board approved final rota.",
        )
        assert result["latest_publish"]["approved_by"] == "Rota Chief"
        assert session.query(RotaPublishApproval).count() == 1

        filename, payload = final_rota_export(session, "2026-05")
        workbook = load_workbook(BytesIO(payload), read_only=True)
        assert filename == "final-rota-2026-05.xlsx"
        assert "Final Rota" in workbook.sheetnames
        assert "Duty Counts" in workbook.sheetnames
        assert "Unit Safety" in workbook.sheetnames
        assert "Publish Readiness" in workbook.sheetnames
        assert "Call Fairness" in workbook.sheetnames
        assert "Review Decisions" in workbook.sheetnames
        assert workbook["Review Items"].cell(row=1, column=7).value == "Accepted"
        assert workbook["Exchange Audit"].cell(row=1, column=2).value == "Validation Status"
        assert workbook["Call Fairness"].cell(row=1, column=1).value == "Call Level"


@contextmanager
def publish_client() -> Generator[TestClient, None, None]:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    testing_session = sessionmaker(bind=engine)

    with testing_session() as seed_session:
        seed_publish_context(seed_session)

    def override_get_db() -> Generator[Session, None, None]:
        with testing_session() as session:
            yield session

    app.dependency_overrides[get_db] = override_get_db
    try:
        yield TestClient(app)
    finally:
        app.dependency_overrides.clear()


def sign_in(client: TestClient) -> str:
    response = client.post(
        "/api/v1/auth/sign-in",
        json={"username": "rotachief", "password": "rotateam"},
    )
    assert response.status_code == 200
    return response.json()["token"]


def test_rota_publish_api_publishes_and_downloads_export() -> None:
    with publish_client() as client:
        token = sign_in(client)
        headers = {"Authorization": f"Bearer {token}"}
        checklist = client.get("/api/v1/rota-publish/month?month=2026-05", headers=headers)
        assert checklist.status_code == 200
        assert checklist.json()["can_publish"] is True

        publish = client.post(
            "/api/v1/rota-publish/publish?month=2026-05",
            headers=headers,
            json={"approval_note": "Approved by board.", "confirm_warnings": False},
        )
        assert publish.status_code == 200
        assert publish.json()["latest_publish"]["status"] == "published"

        export = client.get("/api/v1/rota-publish/export?month=2026-05", headers=headers)
        assert export.status_code == 200
        assert export.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
