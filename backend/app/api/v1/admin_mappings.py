from pathlib import Path
from typing import Literal
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException
from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.session import get_db
from app.domain.duty_types import DUTY_TYPES
from app.models.mappings import AdminMapping
from app.services.imports import (
    classify_duty_label,
    clean_cell_value,
    is_blank_assignment_value,
    normalize_label,
)

router = APIRouter()

REPO_DIR = Path(__file__).resolve().parents[4]
HISTORICAL_DIR = REPO_DIR / "data" / "source" / "historical"

MappingType = Literal["duty_label", "unit_label", "posting_label"]


class AdminMappingRead(BaseModel):
    id: UUID
    mapping_type: str
    source_label: str
    target_key: str | None
    target_label: str | None
    status: str
    source: str
    notes: str | None


class AdminMappingUpdate(BaseModel):
    target_key: str | None = None
    target_label: str | None = None
    status: str = "reviewed"
    notes: str | None = None


class AdminMappingCreate(BaseModel):
    mapping_type: MappingType
    source_label: str
    target_key: str | None = None
    target_label: str | None = None
    status: str = "reviewed"
    notes: str | None = None


class MappingScanResult(BaseModel):
    created: int
    existing: int
    total: int


class MappingOptions(BaseModel):
    duty_types: list[dict[str, str]]
    mapping_types: list[str]


def mapping_to_read(mapping: AdminMapping) -> AdminMappingRead:
    return AdminMappingRead(
        id=mapping.id,
        mapping_type=mapping.mapping_type,
        source_label=mapping.source_label,
        target_key=mapping.target_key,
        target_label=mapping.target_label,
        status=mapping.status,
        source=mapping.source,
        notes=mapping.notes,
    )


def default_status(target_key: str | None) -> str:
    return "suggested" if target_key else "needs_review"


def collect_rota_duty_labels() -> set[str]:
    labels: set[str] = set()
    for path in HISTORICAL_DIR.glob("*.xlsx"):
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.worksheets[0]
            for row in worksheet.iter_rows(min_col=1, max_col=1, min_row=3):
                if not row:
                    continue
                label = clean_cell_value(row[0].value)
                if not label:
                    continue
                if normalize_label(label) in {"consultant", "professor"}:
                    break
                labels.add(label)
        finally:
            workbook.close()
    return labels


def collect_unitwise_labels() -> tuple[set[str], set[str]]:
    unit_labels: set[str] = set()
    posting_labels: set[str] = set()
    unitwise_dir = HISTORICAL_DIR / "unitwise"
    for path in unitwise_dir.glob("*.xlsx"):
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.worksheets[0]
            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                continue
            unit_labels.update(
                clean_cell_value(value)
                for value in rows[0][1:]
                if not is_blank_assignment_value(value)
            )
            posting_labels.update(
                clean_cell_value(row[0])
                for row in rows[1:]
                if row and not is_blank_assignment_value(row[0])
            )
        finally:
            workbook.close()
    return unit_labels, posting_labels


def seed_mapping(
    db: Session,
    mapping_type: str,
    source_label: str,
    target_key: str | None,
    target_label: str | None,
) -> bool:
    existing = db.scalar(
        select(AdminMapping).where(
            AdminMapping.mapping_type == mapping_type,
            AdminMapping.source_label == source_label,
        )
    )
    if existing is not None:
        return False

    db.add(
        AdminMapping(
            mapping_type=mapping_type,
            source_label=source_label,
            target_key=target_key,
            target_label=target_label,
            status=default_status(target_key),
            source="historical_scan",
        )
    )
    return True


@router.get("/admin/mappings/options")
def get_mapping_options() -> MappingOptions:
    return MappingOptions(
        duty_types=[
            {"key": duty_type.key, "label": duty_type.label}
            for duty_type in sorted(DUTY_TYPES, key=lambda duty: duty.label)
        ],
        mapping_types=["duty_label", "unit_label", "posting_label"],
    )


@router.get("/admin/mappings")
def list_mappings(
    mapping_type: MappingType | None = None,
    db: Session = Depends(get_db),
) -> list[AdminMappingRead]:
    statement = select(AdminMapping).order_by(
        AdminMapping.mapping_type,
        AdminMapping.status.desc(),
        AdminMapping.source_label,
    )
    if mapping_type is not None:
        statement = statement.where(AdminMapping.mapping_type == mapping_type)

    return [mapping_to_read(mapping) for mapping in db.scalars(statement)]


@router.post("/admin/mappings")
def create_mapping(payload: AdminMappingCreate, db: Session = Depends(get_db)) -> AdminMappingRead:
    existing = db.scalar(
        select(AdminMapping).where(
            AdminMapping.mapping_type == payload.mapping_type,
            AdminMapping.source_label == payload.source_label,
        )
    )
    if existing is not None:
        raise HTTPException(status_code=409, detail="Mapping already exists")

    mapping = AdminMapping(
        mapping_type=payload.mapping_type,
        source_label=payload.source_label,
        target_key=payload.target_key,
        target_label=payload.target_label,
        status=payload.status,
        notes=payload.notes,
    )
    db.add(mapping)
    db.commit()
    db.refresh(mapping)
    return mapping_to_read(mapping)


@router.put("/admin/mappings/{mapping_id}")
def update_mapping(
    mapping_id: UUID,
    payload: AdminMappingUpdate,
    db: Session = Depends(get_db),
) -> AdminMappingRead:
    mapping = db.get(AdminMapping, mapping_id)
    if mapping is None:
        raise HTTPException(status_code=404, detail="Mapping not found")

    mapping.target_key = payload.target_key
    mapping.target_label = payload.target_label
    mapping.status = payload.status
    mapping.notes = payload.notes
    db.commit()
    db.refresh(mapping)
    return mapping_to_read(mapping)


@router.post("/admin/mappings/scan-historical")
def scan_historical_mappings(db: Session = Depends(get_db)) -> MappingScanResult:
    if not HISTORICAL_DIR.exists():
        raise HTTPException(status_code=404, detail="Historical source folder was not found")

    created = 0
    existing = 0

    for label in sorted(collect_rota_duty_labels()):
        target_key = classify_duty_label(label)
        target_label = next((duty.label for duty in DUTY_TYPES if duty.key == target_key), None)
        if seed_mapping(db, "duty_label", label, target_key, target_label):
            created += 1
        else:
            existing += 1

    unit_labels, posting_labels = collect_unitwise_labels()
    for label in sorted(unit_labels):
        target_key = normalize_label(label).upper().replace(" ", "_")
        if seed_mapping(db, "unit_label", label, target_key, label):
            created += 1
        else:
            existing += 1

    for label in sorted(posting_labels):
        target_key = normalize_label(label).upper().replace(" ", "_")
        if seed_mapping(db, "posting_label", label, target_key, label):
            created += 1
        else:
            existing += 1

    db.commit()
    return MappingScanResult(created=created, existing=existing, total=created + existing)
