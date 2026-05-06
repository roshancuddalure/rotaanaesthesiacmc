import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.models import (
    AdminMapping,
    DutyAssignment,
    DutySlot,
    ImportBatch,
    ImportSourceRecord,
    ImportWarning,
    LeaveRequest,
    Person,
    PersonAlias,
    PersonDesignation,
    PersonPosting,
    RotaPeriod,
    RuleSetting,
    RuleVersion,
    Unit,
)
from app.services.imports import is_valid_person_name, normalize_label
from app.services.roster_reconciliation import DEFAULT_ROSTER_PATH, clean_roster_name

CURRENT_SHEETS = ("ANAESTHESIA", "CARDIAC ANAESTHESIA", "NEURO ANAESTHESIA")
RESET_EFFECTIVE_FROM = date(2026, 3, 1)


@dataclass(frozen=True)
class CleanRosterMember:
    name: str
    source_sheet: str
    call_position: str
    raw_name: str


@dataclass(frozen=True)
class DepartmentRosterResetResult:
    deleted_counts: dict[str, int]
    created_members: int
    created_designations: int
    duplicate_names_skipped: int
    source_file: str
    source_sheets: list[str]
    examples: list[str]


def clean_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.casefold())


def is_person_like(value: object, require_dr: bool = False) -> bool:
    if not isinstance(value, str):
        return False
    if require_dr and not re.search(r"\bDr\.?", value, re.IGNORECASE):
        return False
    name = clean_roster_name(value)
    if not name or not is_valid_person_name(name):
        return False
    normalized = normalize_label(name)
    if normalized in {"doctor name", "name"} or normalized.startswith("total"):
        return False
    blocked_parts = (
        "department",
        "anaesthesia",
        "professor",
        "resident",
        "fellowship",
        "batch",
        "year",
        "staff",
        "designation",
        "medical officer",
        "registrar",
    )
    return not any(part in normalized for part in blocked_parts)


def normalize_position(value: object, fallback: str) -> str:
    if not isinstance(value, str) or not value.strip():
        return fallback
    cleaned = value.strip().upper()
    cleaned = cleaned.replace("ASSOC.", "ASSOCIATE")
    cleaned = cleaned.replace("ASST.", "ASSISTANT")
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned


def add_member(
    members: dict[str, CleanRosterMember],
    raw_name: object,
    source_sheet: str,
    call_position: str,
) -> None:
    name = clean_roster_name(raw_name)
    if not name:
        return
    key = clean_key(name)
    if key in members:
        return
    members[key] = CleanRosterMember(
        name=name,
        source_sheet=source_sheet,
        call_position=call_position,
        raw_name=str(raw_name).strip(),
    )


def extract_general_anaesthesia(members: dict[str, CleanRosterMember], workbook) -> None:
    worksheet = workbook["ANAESTHESIA"]
    for row in worksheet.iter_rows(values_only=True):
        for value in row:
            if is_person_like(value, require_dr=True):
                add_member(members, value, "ANAESTHESIA", "ANAESTHESIA ROSTER")


def extract_cardiac_anaesthesia(members: dict[str, CleanRosterMember], workbook) -> None:
    worksheet = workbook["CARDIAC ANAESTHESIA"]
    current_position = "CARDIAC ANAESTHESIA ROSTER"
    for row in worksheet.iter_rows(values_only=True):
        label = row[1] if len(row) > 1 else None
        if isinstance(label, str) and not is_person_like(label, require_dr=True):
            normalized = normalize_label(label)
            if any(word in normalized for word in ("professor", "resident", "fellowship")):
                current_position = normalize_position(label, current_position)
        if len(row) > 1 and is_person_like(row[1], require_dr=True):
            explicit_position = row[3] if len(row) > 3 else None
            add_member(
                members,
                row[1],
                "CARDIAC ANAESTHESIA",
                normalize_position(explicit_position, current_position),
            )


def extract_neuro_anaesthesia(members: dict[str, CleanRosterMember], workbook) -> None:
    worksheet = workbook["NEURO ANAESTHESIA"]
    for row in worksheet.iter_rows(values_only=True):
        for name_index, position_index in ((1, 3), (4, 6)):
            if name_index >= len(row) or not is_person_like(row[name_index]):
                continue
            add_member(
                members,
                row[name_index],
                "NEURO ANAESTHESIA",
                normalize_position(row[position_index] if position_index < len(row) else None, "NEURO ANAESTHESIA ROSTER"),
            )


def extract_clean_roster_members(path: Path = DEFAULT_ROSTER_PATH) -> list[CleanRosterMember]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    members: dict[str, CleanRosterMember] = {}
    extract_general_anaesthesia(members, workbook)
    extract_cardiac_anaesthesia(members, workbook)
    extract_neuro_anaesthesia(members, workbook)
    return sorted(members.values(), key=lambda member: member.name.casefold())


def clear_operational_data(db: Session) -> dict[str, int]:
    deleted_counts: dict[str, int] = {}
    models = (
        DutyAssignment,
        DutySlot,
        PersonPosting,
        LeaveRequest,
        PersonDesignation,
        PersonAlias,
        Person,
        Unit,
        RotaPeriod,
        AdminMapping,
        ImportWarning,
        ImportSourceRecord,
        ImportBatch,
        RuleSetting,
        RuleVersion,
    )
    for model in models:
        deleted_counts[model.__tablename__] = db.query(model).count()
        db.execute(delete(model))
    db.flush()
    return deleted_counts


def reset_department_members_from_roster(
    db: Session,
    path: Path = DEFAULT_ROSTER_PATH,
) -> DepartmentRosterResetResult:
    roster_members = extract_clean_roster_members(path)
    deleted_counts = clear_operational_data(db)
    created_designations = 0

    for roster_member in roster_members:
        person = Person(canonical_name=roster_member.name, active_status="active")
        db.add(person)
        db.flush()
        db.add(
            PersonDesignation(
                person=person,
                designation=roster_member.call_position,
                effective_from=RESET_EFFECTIVE_FROM,
                source="trusted_roster_reset",
                notes=f"{roster_member.source_sheet}: {roster_member.raw_name}",
            )
        )
        created_designations += 1

    db.commit()
    return DepartmentRosterResetResult(
        deleted_counts=deleted_counts,
        created_members=len(roster_members),
        created_designations=created_designations,
        duplicate_names_skipped=max(0, len(roster_members) - len({clean_key(member.name) for member in roster_members})),
        source_file=str(path),
        source_sheets=list(CURRENT_SHEETS),
        examples=[f"{member.name} - {member.call_position}" for member in roster_members[:12]],
    )


def roster_counts(db: Session) -> dict[str, int]:
    return {
        "people": db.query(Person).count(),
        "aliases": db.query(PersonAlias).count(),
        "designations": db.query(PersonDesignation).count(),
        "duplicate_names": len(
            db.scalars(select(Person.canonical_name)).all()
        )
        - len({clean_key(name) for name in db.scalars(select(Person.canonical_name)).all()}),
    }
