import re
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Person
from app.services.department_roster_reset import clean_key
from app.services.roster_reconciliation import clean_roster_name

DEFAULT_UNITWISE_CALL_LEVEL_PATH = (
    Path(__file__).resolve().parents[3]
    / "data"
    / "source"
    / "historical"
    / "unitwise"
    / "May 2026.xlsx"
)

CALL_LEVEL_LABELS = {
    "5th calls": "5TH_CALL",
    "4th calls": "4TH_CALL",
    "3rd call SR/APs": "3RD_CALL",
    "DM/PDF": "3RD_CALL",
    "2nd call SRs": "2ND_CALL",
    "2nd call PGs (2022)": "2ND_CALL",
    "3rd call PGs (2023)": "3RD_CALL",
    "2nd call PGs (2023)": "2ND_CALL",
    "2nd call PGs (2024)": "2ND_CALL",
    "1st call PGs (2025)": "1ST_CALL",
}


@dataclass(frozen=True)
class UnitwiseCallLevelEntry:
    raw_name: str
    cleaned_name: str
    call_level: str


@dataclass(frozen=True)
class CallLevelPrefillResult:
    matched: int
    unmatched: int
    cleared: int
    unmatched_names: list[str]
    examples: list[str]


def strip_unitwise_suffixes(value: object) -> str:
    cleaned = clean_roster_name(value)
    cleaned = re.sub(
        r"\bMay\s+\d{1,2}(st|nd|rd|th)?\s+\d{1,2}(st|nd|rd|th)?\b",
        " ",
        cleaned,
        flags=re.IGNORECASE,
    )
    cleaned = re.sub(r"\bBP\s+\d{1,2}\s+\d{1,2}\b", " ", cleaned, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", cleaned).strip()


def extract_unitwise_call_levels(
    path: Path = DEFAULT_UNITWISE_CALL_LEVEL_PATH,
) -> list[UnitwiseCallLevelEntry]:
    worksheet = load_workbook(path, data_only=True, read_only=True).active
    current_call_level: str | None = None
    entries: list[UnitwiseCallLevelEntry] = []

    for row in worksheet.iter_rows(values_only=True):
        first_cell = row[0] if row else None
        if isinstance(first_cell, str) and first_cell.strip() in CALL_LEVEL_LABELS:
            current_call_level = CALL_LEVEL_LABELS[first_cell.strip()]
        elif isinstance(first_cell, str) and first_cell.strip():
            current_call_level = None

        if current_call_level is None:
            continue

        for value in row[1:]:
            if not isinstance(value, str) or not value.strip():
                continue
            cleaned = strip_unitwise_suffixes(value)
            if not cleaned:
                continue
            entries.append(
                UnitwiseCallLevelEntry(
                    raw_name=value.strip(),
                    cleaned_name=cleaned,
                    call_level=current_call_level,
                )
            )

    deduped: dict[tuple[str, str], UnitwiseCallLevelEntry] = {}
    for entry in entries:
        deduped.setdefault((clean_key(entry.cleaned_name), entry.call_level), entry)
    return list(deduped.values())


def person_match_score(query: str, person: Person) -> float:
    query_key = clean_key(query)
    person_key = clean_key(person.canonical_name)
    query_tokens = [token.casefold() for token in re.findall(r"[A-Za-z0-9]+", query)]
    person_tokens = [token.casefold() for token in re.findall(r"[A-Za-z0-9]+", person.canonical_name)]
    if len(query_tokens) >= 2 and len(query_tokens[-1]) <= 2:
        wanted_initial = query_tokens[-1][0]
        if not any(token.startswith(wanted_initial) for token in person_tokens[1:]):
            return 0.0
    if query_key == person_key:
        return 1.0
    if len(query_key) >= 5 and person_key.startswith(query_key):
        return 0.96

    query_token_set = set(query_tokens)
    person_token_set = set(person_tokens)
    if query_token_set and query_token_set <= person_token_set:
        return 0.95
    return SequenceMatcher(None, query_key, person_key).ratio()


def match_person(query: str, people: list[Person]) -> Person | None:
    scored = sorted(
        ((person_match_score(query, person), person) for person in people),
        key=lambda item: item[0],
        reverse=True,
    )
    if not scored or scored[0][0] < 0.88:
        return None
    if len(scored) > 1 and scored[0][0] == scored[1][0]:
        return None
    return scored[0][1]


def prefill_call_levels_from_unitwise(
    db: Session,
    path: Path = DEFAULT_UNITWISE_CALL_LEVEL_PATH,
) -> CallLevelPrefillResult:
    people = db.scalars(select(Person).order_by(Person.canonical_name)).all()
    entries = extract_unitwise_call_levels(path)
    matched = 0
    unmatched_names: list[str] = []
    examples: list[str] = []

    cleared = 0
    for person in people:
        if person.call_level:
            person.call_level = None
            cleared += 1

    for entry in entries:
        person = match_person(entry.cleaned_name, people)
        if person is None:
            unmatched_names.append(entry.raw_name)
            continue
        if person.call_level and person.call_level != entry.call_level:
            person.call_level = f"{person.call_level}; {entry.call_level}"
        else:
            person.call_level = entry.call_level
        matched += 1
        if len(examples) < 20:
            examples.append(f"{entry.raw_name} -> {person.canonical_name}: {person.call_level}")

    db.commit()
    return CallLevelPrefillResult(
        matched=matched,
        unmatched=len(unmatched_names),
        cleared=cleared,
        unmatched_names=unmatched_names[:50],
        examples=examples,
    )
