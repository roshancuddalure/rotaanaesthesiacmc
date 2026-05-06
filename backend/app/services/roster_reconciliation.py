import re
from dataclasses import dataclass, field
from datetime import date
from difflib import SequenceMatcher
from pathlib import Path
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, selectinload

from app.models import Person, PersonDesignation
from app.services.imports import clean_cell_value, is_valid_person_name, normalize_label
from app.services.members import compact_name_key, ensure_person_alias, merge_people

DEFAULT_ROSTER_PATH = (
    Path(__file__).resolve().parents[3]
    / "Plan"
    / "Data"
    / "ANAESTHESIA department doctors(namelist).xlsx"
)

CURRENT_ROSTER_SHEETS = {
    "OVERALL MARCH 26",
    "ANAESTHESIA",
    "CARDIAC ANAESTHESIA",
    "NEURO ANAESTHESIA",
}

EXPLICIT_DESIGNATION_SHEETS = {"CARDIAC ANAESTHESIA", "NEURO ANAESTHESIA"}

REJECT_LABEL_PARTS = (
    "anaesthesia",
    "department",
    "designation",
    "professor",
    "professors",
    "associate",
    "assistant",
    "resident",
    "fellowship",
    "tutor",
    "batch",
    "year",
    "total",
    "staff",
    "emp",
    "s no",
)

STATUS_SUFFIX_RE = re.compile(
    r"\s*(?:[-–—]\s*)?\b(study leave|maternity leave|senior resident|sr|ml|ap|micu)\b.*$",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class RosterEntry:
    canonical_name: str
    raw_name: str
    sheet_name: str
    designation: str | None = None


@dataclass
class RosterReconciliationResult:
    roster_entries: int = 0
    matched_people: int = 0
    created_people: int = 0
    renamed_people: int = 0
    merged_people: int = 0
    aliases_created: int = 0
    designations_created: int = 0
    unmatched_database_people: int = 0
    examples: list[str] = field(default_factory=list)


def display_name(value: str) -> str:
    words: list[str] = []
    for token in value.split():
        if len(token) <= 2:
            words.append(token.upper())
        else:
            words.append(token[:1].upper() + token[1:].lower())
    return " ".join(words)


def clean_roster_name(value: object) -> str:
    cleaned = clean_cell_value(value)
    cleaned = re.sub(r"^(dr|prof|mr|mrs|ms)\.?\s*", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\([^)]*\)", " ", cleaned)
    cleaned = STATUS_SUFFIX_RE.sub(" ", cleaned)
    cleaned = cleaned.replace(".", " ")
    cleaned = re.sub(r"\s*[-–—]\s*$", " ", cleaned)
    cleaned = re.sub(r"[^A-Za-z0-9]+", " ", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return display_name(cleaned)


def is_roster_person_value(value: object) -> bool:
    cleaned = clean_roster_name(value)
    if not cleaned or not is_valid_person_name(cleaned):
        return False
    normalized = normalize_label(cleaned)
    if re.fullmatch(r"\d+", normalized):
        return False
    if any(part in normalized for part in REJECT_LABEL_PARTS):
        return False
    if normalized.startswith(("md ", "dm ", "pdf ")):
        return False
    return True


def normalize_designation(value: object) -> str | None:
    cleaned = clean_cell_value(value)
    if not cleaned:
        return None
    normalized = normalize_label(cleaned)
    if not normalized or normalized.isdigit() or normalized in {"designation", "emp no"}:
        return None
    return cleaned.upper().replace("ASSOC.", "ASSOCIATE").replace("ASST.", "ASSISTANT")


def extract_department_roster(path: Path = DEFAULT_ROSTER_PATH) -> list[RosterEntry]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    entries: dict[str, RosterEntry] = {}

    for sheet_name in CURRENT_ROSTER_SHEETS:
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows(values_only=True):
            if sheet_name in EXPLICIT_DESIGNATION_SHEETS:
                explicit_cells = ((1, 3), (4, 6))
                for name_index, designation_index in explicit_cells:
                    if name_index >= len(row) or not is_roster_person_value(row[name_index]):
                        continue
                    canonical = clean_roster_name(row[name_index])
                    designation = (
                        normalize_designation(row[designation_index])
                        if designation_index < len(row)
                        else None
                    )
                    entries.setdefault(
                        compact_name_key(canonical),
                        RosterEntry(canonical, clean_cell_value(row[name_index]), sheet_name, designation),
                    )
                continue

            for value in row:
                if not is_roster_person_value(value):
                    continue
                canonical = clean_roster_name(value)
                entries.setdefault(
                    compact_name_key(canonical),
                    RosterEntry(canonical, clean_cell_value(value), sheet_name),
                )

    return sorted(entries.values(), key=lambda entry: entry.canonical_name.casefold())


def person_keys(person: Person) -> set[str]:
    keys = {compact_name_key(clean_roster_name(person.canonical_name))}
    keys.update(compact_name_key(clean_roster_name(alias.alias)) for alias in person.aliases)
    return {key for key in keys if key}


def safely_matches(roster_key: str, person_key: str) -> bool:
    if roster_key == person_key:
        return True
    if min(len(roster_key), len(person_key)) >= 7 and abs(len(roster_key) - len(person_key)) <= 3:
        if roster_key.startswith(person_key) or person_key.startswith(roster_key):
            return True
    return SequenceMatcher(None, roster_key, person_key).ratio() >= 0.94


def find_matching_people(roster_key: str, people: list[Person]) -> list[Person]:
    matches: list[Person] = []
    for person in people:
        keys = person_keys(person)
        if any(safely_matches(roster_key, key) for key in keys):
            matches.append(person)
    return matches


def alias_count(person: Person) -> int:
    return len(person.aliases)


def ensure_alias_counted(
    db: Session,
    person: Person,
    alias: str,
    result: RosterReconciliationResult,
    source: str,
) -> None:
    before = alias_count(person)
    ensure_person_alias(db, person, alias, source=source)
    db.flush()
    db.refresh(person, attribute_names=["aliases"])
    if alias_count(person) > before:
        result.aliases_created += 1


def rename_to_roster_name(
    db: Session,
    person: Person,
    canonical_name: str,
    result: RosterReconciliationResult,
) -> Person:
    if person.canonical_name == canonical_name:
        return person
    existing = db.scalar(select(Person).where(Person.canonical_name == canonical_name))
    if existing is not None and existing.id != person.id:
        merge_people(db, existing.id, [person.id])
        result.merged_people += 1
        return existing
    ensure_alias_counted(db, person, person.canonical_name, result, "roster_previous_name")
    person.canonical_name = canonical_name
    result.renamed_people += 1
    return person


def add_roster_designation(
    db: Session,
    person: Person,
    designation: str | None,
    result: RosterReconciliationResult,
) -> None:
    if not designation:
        return
    exists = db.scalar(
        select(PersonDesignation).where(
            PersonDesignation.person_id == person.id,
            PersonDesignation.designation == designation,
            PersonDesignation.effective_from == date(2026, 3, 1),
        )
    )
    if exists is not None:
        return
    db.add(
        PersonDesignation(
            person=person,
            designation=designation,
            effective_from=date(2026, 3, 1),
            source="trusted_roster",
            notes="Imported from ANAESTHESIA department doctors(namelist).xlsx",
        )
    )
    result.designations_created += 1


def reconcile_department_roster(
    db: Session,
    path: Path = DEFAULT_ROSTER_PATH,
) -> RosterReconciliationResult:
    entries = extract_department_roster(path)
    result = RosterReconciliationResult(roster_entries=len(entries))

    matched_ids: set[UUID] = set()

    for entry in entries:
        people = db.scalars(
            select(Person)
            .options(selectinload(Person.aliases), selectinload(Person.designations))
            .order_by(Person.canonical_name)
        ).all()
        roster_key = compact_name_key(entry.canonical_name)
        matches = find_matching_people(roster_key, people)
        if matches:
            target = next((person for person in matches if person.canonical_name == entry.canonical_name), matches[0])
            target = db.get(Person, target.id)
            if target is None:
                continue
            target = rename_to_roster_name(db, target, entry.canonical_name, result)
            for source in matches:
                if source.id == target.id:
                    continue
                if db.get(Person, source.id) is None or db.get(Person, target.id) is None:
                    continue
                merge_people(db, target.id, [source.id])
                result.merged_people += 1
            matched_ids.add(target.id)
            result.matched_people += 1
        else:
            target = Person(canonical_name=entry.canonical_name, active_status="active")
            db.add(target)
            db.flush()
            db.refresh(target, attribute_names=["aliases", "designations"])
            matched_ids.add(target.id)
            result.created_people += 1

        ensure_alias_counted(db, target, entry.raw_name, result, "trusted_roster")
        add_roster_designation(db, target, entry.designation, result)
        if len(result.examples) < 10:
            result.examples.append(f"{entry.raw_name} -> {entry.canonical_name}")

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise

    all_people = db.scalars(select(Person)).all()
    result.unmatched_database_people = sum(1 for person in all_people if person.id not in matched_ids)
    return result
