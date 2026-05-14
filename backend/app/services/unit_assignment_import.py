from __future__ import annotations

import io
import logging
import re
import unicodedata
from dataclasses import dataclass
from datetime import date
from difflib import SequenceMatcher
from uuid import UUID

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.models import AdminMapping, Person, PersonAlias, PersonPosting, Unit
from app.services.imports import (
    ParseWarning,
    ParsedUnitPosting,
    clean_cell_value,
    is_blank_assignment_value,
    normalize_label,
    split_unitwise_names,
)
from app.services.leave import month_bounds
from app.services.leave_import import match_person, person_lookup
from app.services.leave_import import normalize_name as normalize_person_name
from app.services.unit_management import UNIT_BOARD_SOURCE, is_special_unit_posting, normalize_posting_type

logger = logging.getLogger(__name__)

POSTING_ALIASES = {
    "1STCALL": "1ST_CALL",
    "FIRSTCALL": "1ST_CALL",
    "CO1STCALL": "CO_1ST_CALL",
    "COFIRSTCALL": "CO_1ST_CALL",
    "2NDCALL": "2ND_CALL",
    "SECONDCALL": "2ND_CALL",
    "3RDCALL": "3RD_CALL",
    "THIRDCALL": "3RD_CALL",
    "DMPDF": "3RD_CALL",
    "4THCALL": "4TH_CALL",
    "FOURTHCALL": "4TH_CALL",
    "CO4THCALL": "CO_4TH_CALL",
    "COFOURTHCALL": "CO_4TH_CALL",
    "5THCALL": "5TH_CALL",
    "FIFTHCALL": "5TH_CALL",
    "PAC": "PAC",
    "PAINCALL": "PAIN",
    "PAIN": "PAIN",
    "ICU": "SICU",
    "ICUPOSTING": "SICU",
    "ICUPOSTINGS": "SICU",
    "SICU": "SICU",
    "SICUPOSTING": "SICU",
    "SICUPOSTINGS": "SICU",
    "DRP": "DRP",
    "DRPPOSTING": "DRP",
    "NEUROICU": "NEURO_ICU",
    "NEUROICUPOSTING": "NEURO_ICU",
    "RESERVE": "OTHER_SPECIAL",
}
MEMBER_AUTO_MATCH_THRESHOLD = 0.88
MEMBER_AUTO_MATCH_GAP = 0.04
UNIT_AUTO_MATCH_THRESHOLD = 0.80
UNIT_AUTO_MATCH_GAP = 0.04
UNIT_REVIEW_SUGGESTED_THRESHOLD = 0.92
UNIT_NUMBER_WORDS = {
    "one": 1,
    "two": 2,
    "three": 3,
    "four": 4,
    "five": 5,
    "six": 6,
    "seven": 7,
    "eight": 8,
    "nine": 9,
    "ten": 10,
    "eleven": 11,
    "twelve": 12,
}
UNIT_ROMAN_NUMERALS = {
    "i": 1,
    "ii": 2,
    "iii": 3,
    "iv": 4,
    "v": 5,
    "vi": 6,
    "vii": 7,
    "viii": 8,
    "ix": 9,
    "x": 10,
    "xi": 11,
    "xii": 12,
}


@dataclass(frozen=True)
class ParsedUnitwiseUpload:
    postings: tuple[ParsedUnitPosting, ...]
    warnings: tuple[ParseWarning, ...]
    sheets: tuple[str, ...]
    source_format: str


def compact_token(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", value)
    ascii_text = "".join(character for character in decomposed if not unicodedata.combining(character))
    return re.sub(r"[^A-Za-z0-9]+", "", ascii_text).upper()


def normalize_unit_label(value: str) -> str:
    normalized = normalize_label(value)
    return re.sub(r"\b(unit|ot|campus|main|department|dept)\b", "", normalized).strip()


def unit_number_from_label(value: str) -> int | None:
    label = normalize_label(value)
    if not label:
        return None
    tokens = label.split()
    has_unit_context = any(token in {"unit", "u", "ot"} for token in tokens)
    for token in tokens:
        if token.isdigit() and (has_unit_context or len(tokens) == 1):
            return int(token)
        if token in UNIT_NUMBER_WORDS and (has_unit_context or len(tokens) == 1):
            return UNIT_NUMBER_WORDS[token]
        if has_unit_context and token in UNIT_ROMAN_NUMERALS:
            return UNIT_ROMAN_NUMERALS[token]
    compacted = re.sub(r"[^a-z0-9]+", "", label)
    compact_match = re.fullmatch(r"(?:unit|u)0*(\d{1,2})", compacted)
    if compact_match:
        return int(compact_match.group(1))
    return None


def person_tokens(value: str) -> list[str]:
    return [token for token in normalize_label(value).split() if token]


def relaxed_person_token(value: str) -> str:
    return relaxed_name_key(value)


def initials_for_tokens(tokens: list[str]) -> str:
    return "".join(token[0] for token in tokens if token)


def relaxed_name_key(value: str) -> str:
    key = normalize_person_name(value)
    key = re.sub(r"([bcdfgjklmnpqrstvwxyz])h", r"\1", key)
    key = key.replace("w", "v")
    return key


def normalize_import_posting(value: str) -> str:
    compacted = compact_token(value)
    if compacted in POSTING_ALIASES:
        return POSTING_ALIASES[compacted]
    explicit_call = re.search(
        r"\b(co\s*)?([1-5])(?:st|nd|rd|th)?\s*calls?\b",
        value,
        flags=re.IGNORECASE,
    )
    if explicit_call:
        number = explicit_call.group(2)
        if explicit_call.group(1) and number == "4":
            return "CO_4TH_CALL"
        if explicit_call.group(1) and number == "1":
            return "CO_1ST_CALL"
        return {
            "1": "1ST_CALL",
            "2": "2ND_CALL",
            "3": "3RD_CALL",
            "4": "4TH_CALL",
            "5": "5TH_CALL",
        }[number]
    label = normalize_label(value)
    if "dm pdf" in label:
        return "3RD_CALL"
    if "co" in label and "4" in label and "call" in label:
        return "CO_4TH_CALL"
    if "co" in label and "1" in label and "call" in label:
        return "CO_1ST_CALL"
    if "1" in label and "call" in label:
        return "1ST_CALL"
    if "2" in label and "call" in label:
        return "2ND_CALL"
    if "3" in label and "call" in label:
        return "3RD_CALL"
    if "4" in label and "call" in label:
        return "4TH_CALL"
    if "5" in label and "call" in label:
        return "5TH_CALL"
    if "pac" in label:
        return "PAC"
    if "pain" in label:
        return "PAIN"
    if "neuro" in label and "icu" in label:
        return "NEURO_ICU"
    if "sicu" in label or label in {"icu", "icu posting", "icu postings"}:
        return "SICU"
    if "drp" in label:
        return "DRP"
    if "reserve" in label:
        return "OTHER_SPECIAL"
    return normalize_posting_type(value.replace("calls", "call").replace("CALLS", "CALL"))


def is_unitwise_context_label(value: str) -> bool:
    label = normalize_label(value)
    return label in {
        "main",
        "main campus",
        "ranipet",
        "ranipet campus",
        "rc",
        "unit",
        "date",
        "day",
    }


def posting_allows_context_rows(posting_label: str) -> bool:
    return normalize_import_posting(posting_label) in {"PAIN", "PAC"}


def is_special_section_label(value: str) -> bool:
    return is_special_unit_posting(normalize_import_posting(value))


def is_special_child_posting_label(value: str) -> bool:
    label = normalize_label(value)
    if not label:
        return False
    if date_range_label(value):
        return True
    if re.search(r"\b20\d{2}\b", label) and "call" in label:
        return True
    if "call" in label and any(token in label.split() for token in {"1", "2", "3", "4", "5"}):
        return True
    return label in {"main", "main campus", "ranipet", "ranipet campus", "rc"}


def date_range_label(value: object) -> str | None:
    cleaned = clean_cell_value(value)
    if re.search(
        r"\b(?:\d{4}-\d{2}-\d{2}|\d{1,2}\s*(?:st|nd|rd|th)?)\s*(?:-|to)\s*(?:\d{4}-\d{2}-\d{2}|\d{1,2})\b",
        cleaned,
        re.IGNORECASE,
    ):
        return cleaned
    return None


def looks_like_unit_header(value: object) -> bool:
    label = normalize_label(value)
    if not label:
        return False
    if label in {"date", "day", "name", "names", "posting", "call"}:
        return False
    if unit_number_from_label(label) is not None:
        return True
    if any(token in label.split() for token in {"unit", "main", "cardiac", "neuro", "ranipet", "rc", "ot"}):
        return True
    return False


def header_row_score(row) -> int:
    non_empty_after_first = [cell for cell in row[1:] if clean_cell_value(cell.value)]
    if not non_empty_after_first:
        return 0
    score = sum(2 for cell in non_empty_after_first if looks_like_unit_header(cell.value))
    score += sum(1 for cell in non_empty_after_first if date_range_label(cell.value))
    first_label = normalize_label(row[0].value if row else "")
    if first_label in {"unit", "units", "posting", "call", "duty"}:
        score += 1
    return score


def find_unit_header_row(rows) -> int | None:
    scored = [
        (header_row_score(row), index)
        for index, row in enumerate(rows[: min(len(rows), 12)])
    ]
    scored = [(score, index) for score, index in scored if score > 0]
    if not scored:
        return None
    score, index = max(scored, key=lambda item: (item[0], -item[1]))
    return index if score >= 2 else None


def row_has_date_context(row) -> bool:
    label = normalize_label(row[0].value if row else "")
    if label in {"date", "day"}:
        return True
    return any(date_range_label(cell.value) for cell in row[1:])


def infer_row_date_bounds(*, raw_person_name: str, raw_date_label: str | None, month: str) -> tuple[date, date]:
    month_start, month_end = month_bounds(month)
    for source in (raw_person_name, raw_date_label or ""):
        full_date_match = re.search(
            r"\b(\d{4}-\d{2}-\d{2})\s*(?:to|-)\s*(\d{4}-\d{2}-\d{2})\b",
            source,
            flags=re.IGNORECASE,
        )
        if full_date_match:
            try:
                start = date.fromisoformat(full_date_match.group(1))
                end = date.fromisoformat(full_date_match.group(2))
                if start <= end:
                    return start, end
            except ValueError:
                pass
        match = re.search(
            r"\b(\d{1,2})\s*(?:st|nd|rd|th)?\s*(?:-|to)\s*(\d{1,2})\b",
            source,
            flags=re.IGNORECASE,
        )
        if not match:
            continue
        start_day = int(match.group(1))
        end_day = int(match.group(2))
        if end_day < start_day:
            continue
        try:
            return date(month_start.year, month_start.month, start_day), date(
                month_start.year,
                month_start.month,
                end_day,
            )
        except ValueError:
            continue
    return month_start, month_end


def ranges_overlap(first_start: str, first_end: str, second_start: str, second_end: str) -> bool:
    return first_start <= second_end and second_start <= first_end


def parse_unitwise_excel_upload(filename: str, content: bytes) -> ParsedUnitwiseUpload:
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    postings: list[ParsedUnitPosting] = []
    warnings: list[ParseWarning] = []
    sheets: list[str] = []
    source_file = filename
    try:
        for worksheet in workbook.worksheets:
            sheets.append(worksheet.title)
            rows = list(worksheet.iter_rows(values_only=False))
            if not rows:
                warnings.append(
                    ParseWarning(source_file, worksheet.title, None, None, "EMPTY_UNITWISE", "No rows found.")
                )
                continue
            header_row_index = find_unit_header_row(rows)
            if header_row_index is None:
                header_row_index = next(
                    (
                        index
                        for index, row in enumerate(rows[: min(len(rows), 12)])
                        if any(clean_cell_value(cell.value) for cell in row[1:])
                    ),
                    None,
                )
                if header_row_index is None:
                    warnings.append(
                        ParseWarning(
                            source_file,
                            worksheet.title,
                            None,
                            None,
                            "MISSING_UNIT_HEADER",
                            "No unit/date header row found in the top rows.",
                        )
                    )
                    continue
                warnings.append(
                    ParseWarning(
                        source_file,
                        worksheet.title,
                        header_row_index + 1,
                        None,
                        "LOW_CONFIDENCE_HEADER",
                        "Used first non-empty row as a fallback header.",
                    )
                )
            header_row = rows[header_row_index]
            unit_by_column = {
                column_index: clean_cell_value(cell.value)
                for column_index, cell in enumerate(header_row[1:], start=2)
                if clean_cell_value(cell.value)
            }
            if not unit_by_column:
                warnings.append(
                    ParseWarning(
                        source_file,
                        worksheet.title,
                        1,
                        None,
                        "MISSING_UNIT_HEADER",
                        "No unit headers found in the first row.",
                    )
                )
                continue
            current_posting_label = ""
            current_special_section_label: str | None = None
            date_label_by_column: dict[int, str] = {
                column_index: label
                for column_index, label in unit_by_column.items()
                if date_range_label(label)
            }
            parser_rule = (
                "excel_first_row_headers"
                if header_row_index == 0
                else "excel_header_scan"
            )
            parser_confidence = "high" if header_row_score(header_row) >= 2 else "low"
            for row in rows[header_row_index + 1 :]:
                if not row:
                    continue
                label = clean_cell_value(row[0].value)
                if row_has_date_context(row):
                    for column_index, cell in enumerate(row[1:], start=2):
                        if date_label := date_range_label(cell.value):
                            date_label_by_column[column_index] = date_label
                    if not split_unitwise_names(" ".join(clean_cell_value(cell.value) for cell in row[1:])):
                        continue
                label_is_context = bool(
                    current_posting_label
                    and posting_allows_context_rows(current_posting_label)
                    and is_unitwise_context_label(label)
                )
                label_is_special_child = bool(
                    current_special_section_label
                    and label
                    and is_special_child_posting_label(label)
                    and not is_special_section_label(label)
                )
                child_posting_label = label if label_is_special_child else None
                if label and not label_is_context and not label_is_special_child:
                    current_posting_label = label
                    current_special_section_label = label if is_special_section_label(label) else None
                if not current_posting_label:
                    continue
                for column_index, cell in enumerate(row[1:], start=2):
                    raw_date_label = date_label_by_column.get(column_index)
                    unit_label = label if label_is_context else unit_by_column.get(column_index)
                    if unit_label is None or is_blank_assignment_value(cell.value):
                        continue
                    names = split_unitwise_names(cell.value)
                    if not names:
                        if posting_allows_context_rows(current_posting_label) and (
                            date_label := date_range_label(cell.value)
                        ):
                            date_label_by_column[column_index] = date_label
                            continue
                        warnings.append(
                            ParseWarning(
                                source_file,
                                worksheet.title,
                                cell.row,
                                column_index,
                                "INVALID_PERSON_NAME",
                                f"Discarded non-person unitwise value: {clean_cell_value(cell.value)}",
                            )
                        )
                        continue
                    for person_name in names:
                        postings.append(
                            ParsedUnitPosting(
                                source_file=source_file,
                                sheet_name=worksheet.title,
                                unit_label=unit_label,
                                posting_label=current_posting_label,
                                person_name=person_name,
                                raw_person_name=clean_cell_value(cell.value),
                                row_index=cell.row,
                                column_index=column_index,
                                column_label=get_column_letter(column_index),
                                raw_date_label=raw_date_label,
                                section_posting_label=current_special_section_label,
                                child_posting_label=child_posting_label,
                                parser_rule=parser_rule,
                                parser_confidence=parser_confidence,
                                source_context="; ".join(
                                    part
                                    for part in (
                                        f"header row {header_row_index + 1}",
                                        f"unit/date header {unit_label}",
                                        f"posting {current_posting_label}",
                                        f"child {child_posting_label}" if child_posting_label else "",
                                        f"date {raw_date_label}" if raw_date_label else "",
                                    )
                                    if part
                                ),
                            )
                        )
    finally:
        workbook.close()
    return ParsedUnitwiseUpload(tuple(postings), tuple(warnings), tuple(sheets), "excel_unitwise")


def text_lines(filename: str, content: bytes) -> list[str]:
    last_error: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            return content.decode(encoding).splitlines()
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ValueError(f"Unable to read text unitwise file: {last_error}") from last_error


def line_unit_match(line: str, unit_labels: dict[str, str]) -> str | None:
    key = normalize_unit_label(line)
    if key in unit_labels:
        return unit_labels[key]
    line_key = normalize_unit_label(re.split(r"[:|,\t-]", line, maxsplit=1)[0])
    return unit_labels.get(line_key)


def posting_from_line(line: str) -> tuple[str | None, str]:
    stripped_line = line.strip().rstrip(":").strip()
    compacted_line = compact_token(stripped_line)
    if compacted_line in POSTING_ALIASES:
        key = POSTING_ALIASES[compacted_line]
        return key, ""
    direct_key = stripped_line.upper().replace(" ", "_").replace("-", "_")
    if direct_key in set(POSTING_ALIASES.values()) | {"OTHER_SPECIAL"}:
        return direct_key, ""
    if is_special_unit_posting(direct_key):
        key = direct_key
        return key, ""
    patterns = [
        r"\bco\s*4(?:th)?\s*calls?\b",
        r"\bco\s*1(?:st)?\s*calls?\b",
        r"\b[1-5](?:st|nd|rd|th)?\s*calls?\b",
        r"\bfirst\s*calls?\b",
        r"\bsecond\s*calls?\b",
        r"\bthird\s*calls?\b",
        r"\bfourth\s*calls?\b",
        r"\bfifth\s*calls?\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, line, re.IGNORECASE)
        if match:
            remainder = (line[: match.start()] + " " + line[match.end() :]).strip(" :-|\t")
            return match.group(0), remainder
    return None, line


def parse_unitwise_text_upload(filename: str, content: bytes, units: list[Unit]) -> ParsedUnitwiseUpload:
    unit_labels: dict[str, str] = {}
    for unit in units:
        for label in (unit.name, unit.code):
            unit_labels[normalize_unit_label(label)] = unit.name

    postings: list[ParsedUnitPosting] = []
    warnings: list[ParseWarning] = []
    current_unit = ""
    current_posting = ""
    current_special_section = ""
    for row_number, raw_line in enumerate(text_lines(filename, content), start=1):
        line = clean_cell_value(raw_line)
        if not line or line.startswith("#") or normalize_label(line).startswith("month "):
            continue
        section_match = re.fullmatch(r"\[([^\]]+)\]", line)
        if section_match:
            section = section_match.group(1).strip()
            if section.casefold().startswith("unit:"):
                unit_label = section.split(":", 1)[1].strip()
                detected_unit = line_unit_match(unit_label, unit_labels) or unit_label
                current_unit = detected_unit
                current_posting = ""
                current_special_section = ""
                continue
            posting_type = normalize_import_posting(section)
            if is_special_unit_posting(posting_type):
                current_unit = ""
                current_posting = posting_type
                current_special_section = section
                continue
            warnings.append(
                ParseWarning(filename, "Text", row_number, None, "UNKNOWN_SECTION", f"Unknown section: {line}")
            )
            continue
        detected_unit = line_unit_match(line, unit_labels)
        if detected_unit and normalize_unit_label(line) in unit_labels:
            current_unit = detected_unit
            current_special_section = ""
            continue

        working_line = line
        if detected_unit:
            current_unit = detected_unit
            current_special_section = ""
            working_line = re.split(r"[:|,\t-]", line, maxsplit=1)[-1].strip()
        posting_label, remainder = posting_from_line(working_line)
        if posting_label:
            current_posting = posting_label
            working_line = remainder
            if is_special_unit_posting(normalize_import_posting(posting_label)):
                current_unit = ""
                current_special_section = posting_label
        if not current_posting or (not current_unit and not is_special_unit_posting(current_posting)):
            warnings.append(
                ParseWarning(
                    filename,
                    "Text",
                    row_number,
                    None,
                    "MISSING_CONTEXT",
                    f"Could not determine unit/posting for line: {line}",
                )
            )
            continue
        if not working_line:
            continue
        working_line = re.sub(r"^\s*[-*]\s*", "", working_line).strip()
        if not working_line:
            continue
        names_source = working_line.split("|", 1)[0].strip()
        names = split_unitwise_names(names_source)
        if not names:
            warnings.append(
                ParseWarning(
                    filename,
                    "Text",
                    row_number,
                    None,
                    "INVALID_PERSON_NAME",
                    f"Discarded non-person unitwise value: {line}",
                )
            )
            continue
        for person_name in names:
            postings.append(
                ParsedUnitPosting(
                    source_file=filename,
                    sheet_name="Text",
                    unit_label=current_unit,
                    posting_label=current_posting,
                    person_name=person_name,
                    raw_person_name=working_line,
                    row_index=row_number,
                    column_index=1,
                    column_label="A",
                    section_posting_label=current_special_section or None,
                    parser_rule="text_template" if line.startswith("-") or current_special_section else "text_freeform",
                    parser_confidence="high",
                    source_context=f"unit {current_unit or 'special card'}; posting {current_posting}",
                )
            )
    return ParsedUnitwiseUpload(tuple(postings), tuple(warnings), ("Text",), "text_unitwise")


def read_unitwise_upload(db: Session, filename: str, content: bytes) -> ParsedUnitwiseUpload:
    lowered = filename.lower()
    if lowered.endswith((".xlsx", ".xlsm")):
        return parse_unitwise_excel_upload(filename, content)
    if lowered.endswith((".txt", ".csv")):
        units = list(db.scalars(select(Unit).where(Unit.active_status == "active")))
        return parse_unitwise_text_upload(filename, content, units)
    raise ValueError("Unit assignment import supports XLSX, XLSM, TXT, or CSV files")


def unit_lookup(db: Session) -> dict[str, Unit | None]:
    units = list(db.scalars(select(Unit).where(Unit.active_status == "active")))
    lookup: dict[str, Unit | None] = {}
    for unit in units:
        for label in (unit.code, unit.name):
            for key in {normalize_label(label), normalize_unit_label(label)}:
                existing = lookup.get(key)
                if existing is not None and existing.id != unit.id:
                    lookup[key] = None
                elif key:
                    lookup[key] = unit
        unit_text = f"{unit.code} {unit.name}".casefold()
        extra_labels: list[str] = []
        if "main" in unit_text:
            extra_labels.extend(["main", "main campus"])
        if "ranipet" in unit_text or re.search(r"\brc\b", unit_text):
            extra_labels.extend(["ranipet", "ranipet campus", "rc"])
        if "cardiac" in unit_text:
            extra_labels.extend(["cardiac", "cardiac anaesthesia", "cardiac anesthesia", "ctvs"])
        if "neuro" in unit_text:
            extra_labels.extend(
                [
                    "neuro",
                    "neuro anaesthesia",
                    "neuro anesthesia",
                    "neuroanaesthesia",
                    "neuroanesthesia",
                ]
            )
        for label in extra_labels:
            key = normalize_label(label)
            existing = lookup.get(key)
            if existing is not None and existing.id != unit.id:
                lookup[key] = None
            else:
                lookup[key] = unit
    mappings = db.scalars(
        select(AdminMapping).where(
            AdminMapping.mapping_type == "unit_label",
            AdminMapping.target_key.is_not(None),
        )
    ).all()
    unit_by_code = {unit.code.casefold(): unit for unit in units}
    unit_by_name = {unit.name.casefold(): unit for unit in units}
    unit_by_id = {str(unit.id).casefold(): unit for unit in units}
    for mapping in mappings:
        target_key = str(mapping.target_key or "").casefold()
        target_label = str(mapping.target_label or "").casefold()
        unit = unit_by_code.get(target_key) or unit_by_id.get(target_key) or unit_by_name.get(target_label)
        if unit is None:
            continue
        for label in (mapping.source_label, mapping.target_key or "", mapping.target_label or ""):
            for key in {normalize_label(label), normalize_unit_label(label)}:
                existing = lookup.get(key)
                if existing is not None and existing.id != unit.id:
                    lookup[key] = None
                elif key:
                    lookup[key] = unit
    return lookup


def ranked_unit_candidates(raw_label: str, lookup: dict[str, Unit | None]) -> list[tuple[float, Unit]]:
    keys = {normalize_label(raw_label), normalize_unit_label(raw_label)}
    scores_by_unit: dict[UUID, tuple[float, Unit]] = {}
    for candidate_key, unit in lookup.items():
        if unit is None or len(candidate_key) < 2:
            continue
        score = max(SequenceMatcher(None, key, candidate_key).ratio() for key in keys if key)
        previous = scores_by_unit.get(unit.id)
        if previous is None or score > previous[0]:
            scores_by_unit[unit.id] = (score, unit)
    return sorted(scores_by_unit.values(), key=lambda item: item[0], reverse=True)


def match_numbered_unit(raw_label: str, lookup: dict[str, Unit | None]) -> tuple[Unit | None, str | None, float | None]:
    raw_number = unit_number_from_label(raw_label)
    if raw_number is None:
        return None, None, None
    units_by_number: dict[UUID, Unit] = {}
    ambiguous = False
    for candidate_key, unit in lookup.items():
        if unit is None:
            continue
        candidate_number = unit_number_from_label(candidate_key)
        if candidate_number != raw_number:
            continue
        units_by_number[unit.id] = unit
        if len(units_by_number) > 1:
            ambiguous = True
    if ambiguous:
        return None, "unit_number_ambiguous", 1.0
    if units_by_number:
        return next(iter(units_by_number.values())), "unit_number_exact", 1.0
    return None, None, None


def match_unit(raw_label: str, lookup: dict[str, Unit | None]) -> tuple[Unit | None, str | None, float | None]:
    for key in (normalize_label(raw_label), normalize_unit_label(raw_label)):
        if key in lookup:
            return lookup[key], "unit_exact" if lookup[key] is not None else "unit_ambiguous", 1.0
    unit, method, score = match_numbered_unit(raw_label, lookup)
    if method is not None:
        return unit, method, score
    ranked = ranked_unit_candidates(raw_label, lookup)
    if not ranked:
        return None, None, None
    best_score, best_unit = ranked[0]
    second_score = ranked[1][0] if len(ranked) > 1 else 0.0
    if best_score >= UNIT_AUTO_MATCH_THRESHOLD and best_score - second_score >= UNIT_AUTO_MATCH_GAP:
        return best_unit, "unit_fuzzy_auto", best_score
    return None, "unit_fuzzy_candidate", best_score


def person_similarity(query_key: str, candidate_key: str) -> float:
    ratio = SequenceMatcher(None, query_key, candidate_key).ratio()
    if query_key in candidate_key or candidate_key in query_key:
        ratio = max(ratio, min(len(query_key), len(candidate_key)) / max(len(query_key), len(candidate_key)))
    return ratio


def name_variant_similarity(query_name: str, candidate_name: str) -> float:
    query_tokens = person_tokens(query_name)
    candidate_tokens = person_tokens(candidate_name)
    if not query_tokens or not candidate_tokens:
        return 0.0
    query_key = normalize_person_name(query_name)
    candidate_key = normalize_person_name(candidate_name)
    score = person_similarity(query_key, candidate_key)
    relaxed_query = relaxed_name_key(query_name)
    relaxed_candidate = relaxed_name_key(candidate_name)
    if relaxed_query != query_key or relaxed_candidate != candidate_key:
        score = max(score, person_similarity(relaxed_query, relaxed_candidate))
    query_set = set(query_tokens)
    candidate_set = set(candidate_tokens)
    relaxed_query_set = {relaxed_person_token(token) for token in query_tokens}
    relaxed_candidate_set = {relaxed_person_token(token) for token in candidate_tokens}
    if query_set and query_set <= candidate_set:
        score = max(score, min(0.96, 0.86 + 0.03 * len(query_set)))
    if relaxed_query_set and relaxed_query_set <= relaxed_candidate_set:
        score = max(score, min(0.96, 0.87 + 0.035 * len(relaxed_query_set)))
    if candidate_set and candidate_set <= query_set:
        score = max(score, min(0.95, 0.84 + 0.02 * len(candidate_set)))
    if len(query_tokens) == 1:
        first_candidate = candidate_tokens[0]
        first_score = SequenceMatcher(None, query_tokens[0], first_candidate).ratio()
        if first_score >= 0.88:
            score = max(score, min(0.96, first_score + 0.04))
        if len(query_tokens[0]) >= 4 and (first_candidate.startswith(query_tokens[0]) or candidate_key.startswith(query_key)):
            score = max(score, 0.93)
    query_initials = initials_for_tokens(query_tokens)
    candidate_initials = initials_for_tokens(candidate_tokens)
    if len(query_tokens) > 1 and query_initials == candidate_initials:
        score = max(score, 0.91)
    return score


def ranked_person_candidates(cleaned_name: str, lookup) -> list[tuple[float, Person]]:
    scores_by_person: dict[UUID, tuple[float, Person]] = {}
    for candidate_key, person in lookup.items():
        if person is None or len(candidate_key) < 3:
            continue
        score = max(
            name_variant_similarity(cleaned_name, candidate_key),
            name_variant_similarity(cleaned_name, person.canonical_name),
        )
        previous = scores_by_person.get(person.id)
        if previous is None or score > previous[0]:
            scores_by_person[person.id] = (score, person)
    return sorted(scores_by_person.values(), key=lambda item: item[0], reverse=True)


def match_person_for_unit_import(
    cleaned_name: str,
    lookup,
) -> tuple[Person | None, str | None, str, Person | None, float | None]:
    person, match_method, match_confidence, suggestion = match_person(cleaned_name, lookup)
    if person is not None or match_confidence == "ambiguous":
        return person, match_method, match_confidence, suggestion, 1.0 if person is not None else None
    ranked = ranked_person_candidates(cleaned_name, lookup)
    if not ranked:
        return None, None, "low", suggestion, None
    best_score, best_person = ranked[0]
    second_score = ranked[1][0] if len(ranked) > 1 else 0.0
    if best_score >= MEMBER_AUTO_MATCH_THRESHOLD and best_score - second_score >= MEMBER_AUTO_MATCH_GAP:
        return best_person, "fuzzy_auto", "high", None, best_score
    return None, None, "suggestion" if suggestion else "low", suggestion, best_score


def existing_unit_board_keys(db: Session, month: str) -> set[tuple[str, str, str]]:
    starts_on, ends_on = month_bounds(month)
    postings = db.scalars(
        select(PersonPosting).where(
            PersonPosting.source == UNIT_BOARD_SOURCE,
            PersonPosting.starts_on <= ends_on,
            (PersonPosting.ends_on.is_(None)) | (PersonPosting.ends_on >= starts_on),
        )
    )
    return {
        (
            str(posting.person_id),
            str(posting.unit_id) if posting.unit_id else "",
            normalize_posting_type(posting.posting_type),
        )
        for posting in postings
    }


def warning_to_text(warning: ParseWarning) -> str:
    location = f"{warning.sheet_name}"
    if warning.row_index is not None:
        location += f" row {warning.row_index}"
    if warning.column_index is not None:
        location += f" col {warning.column_index}"
    return f"{location}: {warning.message}"


def import_row_key(posting: ParsedUnitPosting) -> str:
    return "|".join(
        [
            posting.sheet_name,
            str(posting.row_index),
            posting.column_label,
            posting.person_name,
            posting.unit_label,
            posting.posting_label,
            posting.section_posting_label or "",
            posting.child_posting_label or "",
        ]
    )


def make_preview_row(
    posting: ParsedUnitPosting,
    *,
    person_matches,
    units: dict[str, Unit | None],
    people_by_id: dict[str, Person],
    units_by_id: dict[str, Unit],
    existing_keys: set[tuple[str, str, str]],
    month: str,
    resolutions: dict[str, dict[str, object]],
) -> dict[str, object]:
    issues: list[str] = []
    resolution_notes: list[str] = []
    auto_resolved_fields: list[str] = []
    review_suggested = False
    row_key = import_row_key(posting)
    resolution = resolutions.get(row_key, {})
    skip_row = bool(resolution.get("skip"))
    person, match_method, match_confidence, suggestion, match_score = match_person_for_unit_import(
        posting.person_name,
        person_matches,
    )
    unit, unit_match_method, unit_match_score = match_unit(posting.unit_label, units)
    posting_type = normalize_import_posting(posting.posting_label)
    if person_id := resolution.get("person_id"):
        resolved_person = people_by_id.get(str(person_id))
        if resolved_person is None:
            issues.append("Selected department member was not found")
        else:
            person = resolved_person
            suggestion = None
            match_method = "manual_override"
            match_confidence = "high"
            match_score = 1.0
            resolution_notes.append(f"Member manually resolved to {person.canonical_name}")
    if unit_id := resolution.get("unit_id"):
        resolved_unit = units_by_id.get(str(unit_id))
        if resolved_unit is None:
            issues.append("Selected unit was not found")
        else:
            unit = resolved_unit
            unit_match_method = "unit_manual_override"
            unit_match_score = 1.0
            resolution_notes.append(f"Unit manually resolved to {unit.name}")
    if resolved_posting_type := resolution.get("posting_type"):
        posting_type = normalize_import_posting(str(resolved_posting_type))
        resolution_notes.append(f"Posting manually resolved to {posting_type}")
    special_posting = is_special_unit_posting(posting_type)
    if special_posting:
        unit = None
        unit_match_method = "special_posting_no_unit"
        unit_match_score = None
        resolution_notes.append("Special posting uses its own card and does not need a unit")
    starts_on, ends_on = infer_row_date_bounds(
        raw_person_name=" ".join(
            value
            for value in (posting.raw_person_name, posting.child_posting_label or "")
            if value
        ),
        raw_date_label=posting.raw_date_label,
        month=month,
    )
    if skip_row:
        issues.append("Row excluded by reviewer")
    if match_confidence == "ambiguous":
        issues.append("Ambiguous department member match")
    elif person is None:
        issues.append("Unresolved department member")
        if suggestion is not None:
            issues.append(f"Suggested match: {suggestion.canonical_name}")
    if unit_match_method in {"unit_ambiguous", "unit_number_ambiguous"} and not special_posting:
        issues.append("Ambiguous unit match")
    elif unit is None and not special_posting:
        issues.append("Unresolved unit")
    if not posting_type:
        issues.append("Missing posting/call level")
    if person is not None:
        if match_method in {"normalized_exact", "manual_override"}:
            auto_resolved_fields.append("member")
        elif match_method == "fuzzy_auto":
            auto_resolved_fields.append("member")
            review_suggested = True
            resolution_notes.append(
                f"Member auto-resolved by fuzzy match at {round((match_score or 0) * 100)}%"
            )
    if special_posting:
        auto_resolved_fields.append("unit")
    elif unit is not None:
        if unit_match_method in {"unit_exact", "unit_number_exact", "unit_manual_override"}:
            auto_resolved_fields.append("unit")
            if unit_match_method == "unit_number_exact":
                resolution_notes.append(f"Unit auto-resolved by unit number to {unit.name}")
        elif unit_match_method == "unit_fuzzy_auto":
            auto_resolved_fields.append("unit")
            review_suggested = True
            resolution_notes.append(
                f"Unit auto-resolved by fuzzy match at {round((unit_match_score or 0) * 100)}%"
            )
    if person is not None and (unit is not None or special_posting):
        key = (str(person.id), str(unit.id) if unit else "", posting_type)
        if key in existing_keys:
            issues.append("Existing matching unit assignment already recorded")
    return {
        "row_key": row_key,
        "row_number": posting.row_index,
        "sheet_name": posting.sheet_name,
        "column_label": posting.column_label,
        "raw_person_name": posting.raw_person_name,
        "cleaned_person_name": posting.person_name,
        "person_id": str(person.id) if person else None,
        "person_name": person.canonical_name if person else None,
        "suggested_person_id": str(suggestion.id) if suggestion else None,
        "suggested_person_name": suggestion.canonical_name if suggestion else None,
        "match_method": match_method,
        "match_confidence": match_confidence,
        "match_score": round(match_score, 3) if match_score is not None else None,
        "raw_unit_label": posting.unit_label,
        "unit_id": str(unit.id) if unit else None,
        "unit_name": unit.name if unit else None,
        "unit_match_method": unit_match_method,
        "unit_match_score": round(unit_match_score, 3) if unit_match_score is not None else None,
        "raw_posting_label": posting.posting_label,
        "section_posting_label": posting.section_posting_label,
        "child_posting_label": posting.child_posting_label,
        "parser_rule": posting.parser_rule,
        "parser_confidence": posting.parser_confidence,
        "source_context": posting.source_context,
        "posting_type": posting_type,
        "special_posting": special_posting,
        "skip": skip_row,
        "starts_on": starts_on.isoformat(),
        "ends_on": ends_on.isoformat(),
        "preview_status": "matched" if not issues else "needs_review",
        "auto_resolved": bool(auto_resolved_fields) and not issues,
        "auto_resolved_fields": auto_resolved_fields,
        "review_suggested": review_suggested and not issues,
        "resolution_notes": resolution_notes,
        "issues": issues,
    }


def mark_duplicate_import_rows(rows: list[dict[str, object]]) -> None:
    seen: dict[str, list[int]] = {}
    for index, row in enumerate(rows):
        if row.get("skip"):
            continue
        person_id = row.get("person_id")
        if not person_id:
            continue
        overlaps = [
            previous
            for previous in seen.get(str(person_id), [])
            if not (
                (row.get("special_posting") or rows[previous].get("special_posting"))
                and row.get("posting_type") != rows[previous].get("posting_type")
            )
            if ranges_overlap(
                str(row.get("starts_on")),
                str(row.get("ends_on")),
                str(rows[previous].get("starts_on")),
                str(rows[previous].get("ends_on")),
            )
        ]
        if overlaps:
            row["preview_status"] = "needs_review"
            row["issues"].append("Person appears more than once in this import for overlapping dates")  # type: ignore[union-attr]
            for previous in overlaps:
                if rows[previous].get("skip"):
                    continue
                rows[previous]["preview_status"] = "needs_review"
                rows[previous]["issues"].append(  # type: ignore[union-attr]
                    "Person appears more than once in this import for overlapping dates"
                )
        seen.setdefault(str(person_id), []).append(index)


def preview_unit_assignment_import(
    db: Session,
    filename: str,
    content: bytes,
    month: str,
    *,
    replace_existing: bool = False,
    resolutions: dict[str, dict[str, object]] | None = None,
) -> dict[str, object]:
    month_bounds(month)
    parsed = read_unitwise_upload(db, filename, content)
    person_matches = person_lookup(db)
    active_people = list(db.scalars(select(Person).where(Person.active_status == "active")))
    active_units = list(db.scalars(select(Unit).where(Unit.active_status == "active")))
    people_by_id = {str(person.id): person for person in active_people}
    units_by_id = {str(unit.id): unit for unit in active_units}
    units = unit_lookup(db)
    existing_keys = set() if replace_existing else existing_unit_board_keys(db, month)
    resolution_lookup = resolutions or {}
    rows = [
        make_preview_row(
            posting,
            person_matches=person_matches,
            units=units,
            people_by_id=people_by_id,
            units_by_id=units_by_id,
            existing_keys=existing_keys,
            month=month,
            resolutions=resolution_lookup,
        )
        for posting in parsed.postings
    ]
    mark_duplicate_import_rows(rows)
    for row in rows:
        auto_assignable = row_is_safe_to_apply(row)
        row["auto_assignable"] = auto_assignable
        row["row_action"] = "auto_assign" if auto_assignable else "needs_review"
        row["auto_assign_blockers"] = [] if auto_assignable else list(row.get("issues", []))
        row["auto_decision_reason"] = (
            "Ready for auto-assign"
            if auto_assignable
            else "; ".join(str(issue) for issue in row.get("issues", [])) or "Needs review"
        )
    matched = sum(1 for row in rows if row["preview_status"] == "matched")
    auto_resolved = sum(1 for row in rows if row.get("auto_resolved"))
    auto_assignable = sum(1 for row in rows if row.get("auto_assignable"))
    needs_review = len(rows) - auto_assignable
    review_suggested = sum(1 for row in rows if row.get("review_suggested"))
    unresolved = sum(
        1
        for row in rows
        if "Unresolved department member" in row["issues"] or "Unresolved unit" in row["issues"]
    )
    invalid = sum(1 for row in rows if row["preview_status"] != "matched") - unresolved
    logger.info(
        "unit_assignment_import.preview filename=%s month=%s rows=%s matched=%s unresolved=%s invalid=%s "
        "auto_resolved=%s auto_assignable=%s needs_review=%s review_suggested=%s replace=%s warnings=%s sheets=%s",
        filename,
        month,
        len(rows),
        matched,
        unresolved,
        max(0, invalid),
        auto_resolved,
        auto_assignable,
        needs_review,
        review_suggested,
        replace_existing,
        len(parsed.warnings),
        ",".join(parsed.sheets),
    )
    for row in rows:
        if row.get("auto_assignable"):
            logger.info(
                "unit_assignment_import.row_auto_assignable filename=%s sheet=%s row=%s col=%s imported_person=%s "
                "matched_person=%s raw_unit=%s matched_unit=%s posting=%s person_match=%s unit_match=%s "
                "review_suggested=%s notes=%s",
                filename,
                row.get("sheet_name"),
                row.get("row_number"),
                row.get("column_label"),
                row.get("raw_person_name"),
                row.get("person_name"),
                row.get("raw_unit_label"),
                row.get("unit_name") or "special_card",
                row.get("posting_type"),
                row.get("match_method") or row.get("match_confidence"),
                row.get("unit_match_method"),
                row.get("review_suggested"),
                "; ".join(str(note) for note in row.get("resolution_notes", [])),
            )
        else:
            logger.info(
                "unit_assignment_import.row_needs_review filename=%s sheet=%s row=%s col=%s person=%s "
                "unit=%s issues=%s person_match=%s unit_match=%s",
                filename,
                row.get("sheet_name"),
                row.get("row_number"),
                row.get("column_label"),
                row.get("raw_person_name"),
                row.get("raw_unit_label"),
                "; ".join(str(issue) for issue in row.get("issues", [])),
                row.get("match_method") or row.get("match_confidence"),
                row.get("unit_match_method"),
            )
    return {
        "filename": filename,
        "month": month,
        "total_rows": len(rows),
        "matched_rows": matched,
        "auto_resolved_rows": auto_resolved,
        "auto_assignable_rows": auto_assignable,
        "needs_review_rows": needs_review,
        "review_suggested_rows": review_suggested,
        "unresolved_rows": unresolved,
        "invalid_rows": max(0, invalid),
        "sheets": list(parsed.sheets),
        "source_formats": [parsed.source_format],
        "parser_warnings": [warning_to_text(warning) for warning in parsed.warnings],
        "rows": rows,
    }


def row_is_safe_to_apply(row: dict[str, object]) -> bool:
    return (
        row.get("preview_status") == "matched"
        and not row.get("skip")
        and row.get("person_id")
        and (row.get("unit_id") or row.get("special_posting"))
        and row.get("posting_type")
        and not row.get("issues")
    )


def clear_unit_board_month(db: Session, month: str) -> int:
    starts_on, ends_on = month_bounds(month)
    result = db.execute(
        delete(PersonPosting).where(
            PersonPosting.source == UNIT_BOARD_SOURCE,
            PersonPosting.starts_on <= ends_on,
            (PersonPosting.ends_on.is_(None)) | (PersonPosting.ends_on >= starts_on),
        )
    )
    return result.rowcount or 0


def learn_person_alias_from_row(db: Session, row: dict[str, object]) -> bool:
    if row.get("match_method") != "manual_override" or not row.get("person_id"):
        return False
    alias = clean_cell_value(row.get("cleaned_person_name") or row.get("raw_person_name"))
    if not alias or "," in alias or "/" in alias or ";" in alias:
        return False
    existing = db.scalar(select(PersonAlias).where(PersonAlias.alias == alias))
    if existing is not None:
        return False
    db.add(
        PersonAlias(
            person_id=UUID(str(row["person_id"])),
            alias=alias,
            source="unit_import_learning",
        )
    )
    return True


def learn_unit_mapping_from_row(db: Session, row: dict[str, object]) -> bool:
    if row.get("unit_match_method") != "unit_manual_override" or not row.get("unit_id"):
        return False
    source_label = clean_cell_value(row.get("raw_unit_label"))
    if not source_label:
        return False
    existing = db.scalar(
        select(AdminMapping).where(
            AdminMapping.mapping_type == "unit_label",
            AdminMapping.source_label == source_label,
        )
    )
    if existing is not None:
        return False
    db.add(
        AdminMapping(
            mapping_type="unit_label",
            source_label=source_label,
            target_key=str(row.get("unit_id")),
            target_label=clean_cell_value(row.get("unit_name")),
            status="reviewed",
            source="unit_import_learning",
            notes="Learned from applied unitwise import correction",
        )
    )
    return True


def learn_from_applied_import_row(db: Session, row: dict[str, object]) -> int:
    learned = 0
    if learn_person_alias_from_row(db, row):
        learned += 1
    if learn_unit_mapping_from_row(db, row):
        learned += 1
    return learned


def apply_unit_assignment_import(
    db: Session,
    filename: str,
    content: bytes,
    month: str,
    *,
    replace_existing: bool = False,
    resolutions: dict[str, dict[str, object]] | None = None,
) -> dict[str, object]:
    preview = preview_unit_assignment_import(
        db,
        filename,
        content,
        month,
        replace_existing=replace_existing,
        resolutions=resolutions,
    )
    starts_on, ends_on = month_bounds(month)
    deleted_rows = clear_unit_board_month(db, month) if replace_existing else 0
    created = 0
    learned_mappings = 0
    skipped_rows: list[dict[str, object]] = []
    for row in preview["rows"]:
        if not row_is_safe_to_apply(row):
            skipped_rows.append(row)
            continue
        assignment = PersonPosting(
            person_id=UUID(str(row["person_id"])),
            unit_id=UUID(str(row["unit_id"])) if row.get("unit_id") else None,
            posting_type=str(row["posting_type"]),
            starts_on=date.fromisoformat(str(row.get("starts_on") or starts_on.isoformat())),
            ends_on=date.fromisoformat(str(row.get("ends_on") or ends_on.isoformat())),
            source=UNIT_BOARD_SOURCE,
            notes=(
                f"Imported from {filename}, sheet {row.get('sheet_name')}, "
                f"row {row.get('row_number')}"
            ),
        )
        db.add(assignment)
        learned_mappings += learn_from_applied_import_row(db, row)
        logger.info(
            "unit_assignment_import.auto_assigned filename=%s month=%s row=%s person_id=%s person=%s "
            "unit_id=%s unit=%s posting=%s dates=%s..%s member_match=%s unit_match=%s review_suggested=%s",
            filename,
            month,
            row.get("row_number"),
            row.get("person_id"),
            row.get("person_name"),
            row.get("unit_id"),
            row.get("unit_name") or "special_card",
            row.get("posting_type"),
            row.get("starts_on"),
            row.get("ends_on"),
            row.get("match_method") or row.get("match_confidence"),
            row.get("unit_match_method"),
            row.get("review_suggested"),
        )
        created += 1
    db.commit()
    logger.info(
        "unit_assignment_import.apply filename=%s month=%s created=%s skipped=%s deleted_existing=%s replace=%s",
        filename,
        month,
        created,
        len(skipped_rows),
        deleted_rows,
        replace_existing,
    )
    return {
        "filename": filename,
        "month": month,
        "created_rows": created,
        "auto_assigned_rows": created,
        "learned_mappings": learned_mappings,
        "deleted_existing_rows": deleted_rows,
        "skipped_rows": len(skipped_rows),
        "skipped_preview_rows": skipped_rows,
        "preview": preview,
    }
