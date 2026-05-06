import re
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.domain.duty_types import DUTY_TYPES


EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xls"}
OPENPYXL_SUFFIXES = {".xlsx", ".xlsm"}
TEXT_SUFFIXES = {".txt", ".csv"}

MONTH_LOOKUP = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


@dataclass(frozen=True)
class SheetProfile:
    name: str
    max_row: int
    max_column: int
    non_empty_cells: int
    first_non_empty_row: int | None
    first_non_empty_column: int | None


@dataclass(frozen=True)
class WorkbookProfile:
    path: Path
    sheets: tuple[SheetProfile, ...]


@dataclass(frozen=True)
class SourceCell:
    sheet_name: str
    row_index: int
    column_index: int
    column_label: str
    raw_value: Any
    cleaned_value: str


@dataclass(frozen=True)
class ReconstructedMonth:
    year: int
    month: int
    source_text: str


@dataclass(frozen=True)
class ParsedRotaAssignment:
    source_file: Path
    sheet_name: str
    duty_date: date
    weekday_label: str
    row_index: int
    column_index: int
    column_label: str
    duty_label: str
    duty_type: str
    person_name: str
    raw_person_name: str
    starts_at: datetime
    ends_at: datetime
    is_24hr: bool


@dataclass(frozen=True)
class ParsedUnitPosting:
    source_file: Path
    sheet_name: str
    unit_label: str
    posting_label: str
    person_name: str
    raw_person_name: str
    row_index: int
    column_index: int
    column_label: str


@dataclass(frozen=True)
class ParseWarning:
    source_file: Path
    sheet_name: str
    row_index: int | None
    column_index: int | None
    code: str
    message: str


@dataclass(frozen=True)
class ParsedRotaWorkbook:
    source_file: Path
    month: ReconstructedMonth
    assignments: tuple[ParsedRotaAssignment, ...]
    warnings: tuple[ParseWarning, ...]


@dataclass(frozen=True)
class ParsedUnitwiseWorkbook:
    source_file: Path
    month: ReconstructedMonth
    postings: tuple[ParsedUnitPosting, ...]
    warnings: tuple[ParseWarning, ...]


DUTY_LABEL_TO_KEY = {
    duty.label.casefold(): duty.key
    for duty in DUTY_TYPES
} | {
    duty.key.casefold(): duty.key
    for duty in DUTY_TYPES
}

IGNORED_PERSON_VALUES = {
    "",
    "-",
    "x",
    "xx",
    "xxx",
    "xxxx",
    "xxxxx",
    "nil",
    "na",
    "n/a",
    "none",
}

NON_PERSON_LABELS = {
    "date",
    "day",
    "main",
    "main campus",
    "ranipet",
    "ranipet campus",
    "rc",
    "cb",
    "pac",
    "pain",
    "pain call",
    "icu",
    "icu posting",
    "icu postings",
    "sicu",
    "sicu posting",
    "drp",
    "drp posting",
    "neuro icu",
    "unit",
    "unit i",
    "unit ii",
    "unit iii",
    "unit iv",
    "unit v",
    "unit vi",
    "dept of cardiac anaesthesia",
    "dept of neuroanaesthesia",
    "out of department postings for pgs",
    "1st year new pgs",
    "professor",
    "consultant",
    "aadil",
    "harish",
    "prabhu",
}

NAME_NOISE_WORDS = (
    "MICU",
    "SICU",
    "ICU",
    "DRP",
    "BP",
    "ONLY",
    "POSTING",
    "POSTINGS",
)

ROTA_STOP_LABELS = {
    "consultant",
    "professor",
}


def detect_import_kind(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in EXCEL_SUFFIXES:
        return "excel"
    if suffix in TEXT_SUFFIXES:
        return "text"
    return "unknown"


def require_supported_excel(path: Path) -> None:
    if path.suffix.lower() not in OPENPYXL_SUFFIXES:
        msg = f"{path.name} is not a supported Excel workbook. Use .xlsx or .xlsm."
        raise ValueError(msg)


def clean_cell_value(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def clean_person_name(value: Any) -> str:
    cleaned = clean_cell_value(value)
    cleaned = re.sub(r"^(dr|prof|mr|mrs|ms)\.?\s*", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s*\([^)]*\)\s*", " ", cleaned)
    cleaned = re.sub(
        r"\s*[-–—]?\s*(till|until)\s+[a-z]{3,9}\s+\d{1,2}\b.*$",
        " ",
        cleaned,
        flags=re.IGNORECASE,
    )
    cleaned = re.sub(r"\b\d{1,2}\s*(st|nd|rd|th)?\s*(to|-)\s*\d{1,2}\b", " ", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\b\d{1,2}\s*[-/]\s*\d{1,2}\b", " ", cleaned)
    for word in NAME_NOISE_WORDS:
        cleaned = re.sub(rf"\b{word}\b", " ", cleaned, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", cleaned).strip(" ,;")


def is_blank_assignment_value(value: Any) -> bool:
    cleaned = clean_person_name(value).casefold()
    return cleaned in IGNORED_PERSON_VALUES


def is_valid_person_name(value: Any) -> bool:
    cleaned = clean_person_name(value)
    normalized = normalize_label(cleaned)
    if not cleaned or normalized in IGNORED_PERSON_VALUES or normalized in NON_PERSON_LABELS:
        return False
    if re.fullmatch(r"[\d\W_]+", cleaned):
        return False
    if any(token in MONTH_LOOKUP for token in normalized.split()):
        return False
    if re.search(r"\b\d{1,2}\s*(st|nd|rd|th)?\s*(to|-)\s*\d{1,2}\b", cleaned, re.IGNORECASE):
        return False
    if len(re.findall(r"[A-Za-z]", cleaned)) < 3:
        return False
    if normalized.startswith(("unit ", "main pac", "rc pac", "main shift", "rc shift")):
        return False
    return True


def split_unitwise_names(value: Any) -> list[str]:
    cleaned = clean_cell_value(value)
    if not cleaned:
        return []
    parts = re.split(r"\s*[,/]\s*", cleaned)
    names = [clean_person_name(part) for part in parts]
    return [name for name in names if is_valid_person_name(name)]


def normalize_label(value: Any) -> str:
    cleaned = clean_cell_value(value).casefold()
    cleaned = cleaned.replace("&", " and ")
    cleaned = re.sub(r"\([^)]*\)", " ", cleaned)
    cleaned = re.sub(r"[^a-z0-9]+", " ", cleaned)
    replacements = {
        "first": "1",
        "second": "2",
        "third": "3",
        "fourth": "4",
        "fifth": "5",
    }
    for word, number in replacements.items():
        cleaned = re.sub(rf"\b{word}\b", number, cleaned)
    return re.sub(r"\s+", " ", cleaned).strip()


def normalize_classification_label(value: Any) -> str:
    cleaned = clean_cell_value(value).casefold()
    cleaned = cleaned.replace("&", " and ")
    cleaned = re.sub(r"\b20\d{2}\b", " ", cleaned)
    cleaned = re.sub(r"[^a-z0-9]+", " ", cleaned)
    replacements = {
        "first": "1",
        "second": "2",
        "third": "3",
        "fourth": "4",
        "fifth": "5",
    }
    for word, number in replacements.items():
        cleaned = re.sub(rf"\b{word}\b", number, cleaned)
    return re.sub(r"\s+", " ", cleaned).strip()


def classify_duty_label(value: Any) -> str | None:
    cleaned = clean_cell_value(value)
    if not cleaned:
        return None
    normalized = re.sub(r"\s+", " ", cleaned).casefold()
    if normalized in DUTY_LABEL_TO_KEY:
        return DUTY_LABEL_TO_KEY[normalized]

    compact = re.sub(r"[^a-z0-9]+", "", normalized)
    for label, key in DUTY_LABEL_TO_KEY.items():
        if re.sub(r"[^a-z0-9]+", "", label) == compact:
            return key

    label = normalize_classification_label(value)
    if "pac" in label:
        return "PAC"
    if "shift" in label:
        return "SHIFT"
    if "neuro" in label or "stroke" in label:
        return "NEURO_DEPT"
    if "chad" in label:
        return "CHAD"
    if "ruhsa" in label:
        return "RUHSA"
    if "cart" in label:
        return "CART"
    if "c s" in label and "am" in label:
        return "CAESAR_A_12HR"
    if label.startswith("cb") and (
        ("co" in label and "1" in label)
        or "cs" in label
        or "am cal" in label
    ):
        return "CB_CO_12HR"
    if label.startswith("cb") and "paed" in label:
        return "CB_PAEDS"
    if "schell" in label:
        if "floating" in label:
            return "SCHELL_AND_FLOATING"
        return "SCHELL_24HR"
    if "floating" in label:
        return "FLOATING_24HR"
    if "5" in label and "call" in label:
        return "FIFTH_CALL"
    if "caesar" in label or "cesar" in label:
        if re.search(r"\b(b|pm)\b", label):
            return "CAESAR_B_24HR"
        return "CAESAR_A_12HR"
    if label.startswith("cb") and "co" in label and "3" in label:
        return "CB_CO3RD_24HR"
    if label.startswith("cb") and "co" in label and "4" in label:
        return "CB_CO4TH_24HR"
    if label.startswith("cb") and "1" in label:
        return "CB_1ST_24HR"
    if label.startswith("cb") and "3" in label:
        return "CB_3RD_24HR"
    if label.startswith("cb") and "4" in label:
        return "CB_4TH_24HR"
    if label.startswith("rc") or label.startswith("ranipet"):
        if "co" in label and "3" in label:
            return "RC_CO3RD_24HR"
        if "co" in label and "4" in label:
            return "RC_CO4TH_24HR"
        if "12" in label:
            return "RC_12HR"
        if "co" in label and "1" in label:
            return "RC_1ST_A_24HR"
        if "co" in label:
            return "RC_CO_12HR"
        if "1" in label and re.search(r"\ba\b", label):
            return "RC_1ST_A_24HR"
        if "1" in label and re.search(r"\bb\b", label):
            return "RC_1ST_B_24HR"
        if "1" in label and re.search(r"\bc\b", label):
            return "RC_12HR"
        if "3" in label:
            return "RC_3RD_24HR"
        if "4" in label:
            return "RC_4TH_24HR"
        if "2" in label:
            return "RC_2ND_24HR"
        if "1" in label:
            return "RC_1ST_A_24HR"
    if label.startswith("main"):
        if "co" in label and "1" in label:
            return "MAIN_1ST_CO_24HR"
        if "co" in label and "3" in label:
            return "MAIN_CO3RD_24HR"
        if "co" in label and "4" in label:
            return "MAIN_CO4TH_24HR"
        if "1" in label:
            return "MAIN_1ST_24HR"
        if "2" in label:
            return "MAIN_2ND_24HR"
        if "3" in label:
            return "MAIN_3RD_24HR"
        if "4" in label:
            return "MAIN_4TH_24HR"
    if "paed" in label:
        return "PAEDS_CALL"
    return None


def reconstruct_month_from_text(value: str) -> ReconstructedMonth | None:
    cleaned = clean_cell_value(value)
    match = re.search(
        r"\b("
        + "|".join(sorted(MONTH_LOOKUP, key=len, reverse=True))
        + r")\b[\s\-_.,]*(20\d{2})",
        cleaned,
        flags=re.IGNORECASE,
    )
    if match is None:
        match = re.search(
            r"\b(20\d{2})[\s\-_.,]*("
            + "|".join(sorted(MONTH_LOOKUP, key=len, reverse=True))
            + r")\b",
            cleaned,
            flags=re.IGNORECASE,
        )
        if match is None:
            return None
        year = int(match.group(1))
        month = MONTH_LOOKUP[match.group(2).casefold()]
    else:
        month = MONTH_LOOKUP[match.group(1).casefold()]
        year = int(match.group(2))

    return ReconstructedMonth(year=year, month=month, source_text=cleaned)


def reconstruct_month_from_path(path: Path) -> ReconstructedMonth | None:
    reconstructed = reconstruct_month_from_text(path.stem)
    if reconstructed is not None:
        return reconstructed

    cleaned = clean_cell_value(path.stem)
    month_match = re.search(
        r"\b(" + "|".join(sorted(MONTH_LOOKUP, key=len, reverse=True)) + r")\b",
        cleaned,
        flags=re.IGNORECASE,
    )
    year_match = re.search(r"\b(20\d{2}|\d{2})\b", cleaned)
    if month_match is None or year_match is None:
        return None

    year = int(year_match.group(1))
    if year < 100:
        year += 2000
    month = MONTH_LOOKUP[month_match.group(1).casefold()]
    return ReconstructedMonth(year=year, month=month, source_text=cleaned)


def reconstruct_rota_date(value: Any, month: ReconstructedMonth) -> date | None:
    cleaned = clean_cell_value(value)
    if not cleaned:
        return None

    if isinstance(value, datetime):
        if value.year == month.year and value.month == month.month:
            return date(month.year, month.month, value.day)
        if value.year == month.year and value.day == month.month:
            return date(month.year, month.month, value.month)
    if isinstance(value, date):
        if value.year == month.year and value.month == month.month:
            return date(month.year, month.month, value.day)
        if value.year == month.year and value.day == month.month:
            return date(month.year, month.month, value.month)

    number_match = re.search(r"\b([1-9]|[12]\d|3[01])\b", cleaned)
    if number_match is None:
        return None
    day = int(number_match.group(1))
    try:
        return date(month.year, month.month, day)
    except ValueError:
        return None


def reconstruct_rota_date_from_weekday(
    weekday_label: Any,
    occurrence: int,
    month: ReconstructedMonth,
) -> date | None:
    weekday_lookup = {
        "monday": 0,
        "mon": 0,
        "tuesday": 1,
        "tue": 1,
        "tues": 1,
        "wednesday": 2,
        "wed": 2,
        "thursday": 3,
        "thu": 3,
        "thur": 3,
        "thurs": 3,
        "friday": 4,
        "fri": 4,
        "saturday": 5,
        "sat": 5,
        "sunday": 6,
        "sun": 6,
    }
    normalized = normalize_label(weekday_label)
    target_weekday = weekday_lookup.get(normalized)
    if target_weekday is None:
        return None

    matching_days: list[int] = []
    for day in range(1, 32):
        try:
            candidate = date(month.year, month.month, day)
        except ValueError:
            break
        if candidate.weekday() == target_weekday:
            matching_days.append(day)

    if occurrence >= len(matching_days):
        return None
    return date(month.year, month.month, matching_days[occurrence])


def duty_window_for_date(duty_date: date, duty_type: str) -> tuple[datetime, datetime, bool]:
    is_24hr = duty_type.endswith("_24HR") or duty_type in {
        "CART",
        "FIFTH_CALL",
        "SCHELL_24HR",
        "FLOATING_24HR",
    }
    if duty_type in {
        "CAESAR_A_12HR",
        "CB_CO_12HR",
        "RC_12HR",
        "RC_CO_12HR",
        "SHIFT",
        "PAC",
        "NEURO_DEPT",
    }:
        starts_at = datetime.combine(duty_date, time(hour=7, minute=30))
        return starts_at, starts_at + timedelta(hours=12), False

    starts_at = datetime.combine(duty_date, time(hour=8))
    return starts_at, starts_at + timedelta(hours=24 if is_24hr else 12), is_24hr


def profile_excel_workbook(path: Path) -> WorkbookProfile:
    require_supported_excel(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        profiles: list[SheetProfile] = []
        for worksheet in workbook.worksheets:
            non_empty_cells = 0
            first_row: int | None = None
            first_column: int | None = None
            max_row = 0
            max_column = 0
            for row in worksheet.iter_rows():
                for cell in row:
                    if clean_cell_value(cell.value):
                        non_empty_cells += 1
                        first_row = cell.row if first_row is None else min(first_row, cell.row)
                        first_column = (
                            cell.column
                            if first_column is None
                            else min(first_column, cell.column)
                        )
                        max_row = max(max_row, cell.row)
                        max_column = max(max_column, cell.column)

            profiles.append(
                SheetProfile(
                    name=worksheet.title,
                    max_row=max_row,
                    max_column=max_column,
                    non_empty_cells=non_empty_cells,
                    first_non_empty_row=first_row,
                    first_non_empty_column=first_column,
                )
            )
    finally:
        workbook.close()

    return WorkbookProfile(path=path, sheets=tuple(profiles))


def iter_excel_source_cells(path: Path, sheet_name: str | None = None) -> list[SourceCell]:
    require_supported_excel(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheets = [workbook[sheet_name]] if sheet_name else workbook.worksheets
        cells: list[SourceCell] = []
        for worksheet in worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    cleaned = clean_cell_value(cell.value)
                    if not cleaned:
                        continue
                    cells.append(
                        SourceCell(
                            sheet_name=worksheet.title,
                            row_index=cell.row,
                            column_index=cell.column,
                            column_label=get_column_letter(cell.column),
                            raw_value=cell.value,
                            cleaned_value=cleaned,
                        )
                    )
    finally:
        workbook.close()

    return cells


def parse_rota_workbook(path: Path) -> ParsedRotaWorkbook:
    require_supported_excel(path)
    month = reconstruct_month_from_path(path)
    if month is None:
        msg = f"Could not infer month/year from {path.name}."
        raise ValueError(msg)

    workbook = load_workbook(path, read_only=True, data_only=True)
    assignments: list[ParsedRotaAssignment] = []
    warnings: list[ParseWarning] = []
    try:
        worksheet = workbook.worksheets[0]
        rows = list(worksheet.iter_rows(values_only=False))
        if len(rows) < 3:
            warnings.append(
                ParseWarning(path, worksheet.title, None, None, "EMPTY_ROTA", "No rota rows found.")
            )
            return ParsedRotaWorkbook(path, month, tuple(assignments), tuple(warnings))

        date_cells = rows[0]
        weekday_cells = rows[1] if len(rows) > 1 else ()
        column_dates: dict[int, date] = {}
        weekday_occurrences: dict[str, int] = {}
        for column_index, cell in enumerate(date_cells[1:], start=2):
            weekday_label = ""
            if column_index - 1 < len(weekday_cells):
                weekday_label = clean_cell_value(weekday_cells[column_index - 1].value)
            normalized_weekday = normalize_label(weekday_label)
            occurrence = weekday_occurrences.get(normalized_weekday, 0)
            duty_date = reconstruct_rota_date(cell.value, month)
            if duty_date is None:
                duty_date = reconstruct_rota_date_from_weekday(weekday_label, occurrence, month)
            if normalized_weekday:
                weekday_occurrences[normalized_weekday] = occurrence + 1
            if duty_date is not None:
                column_dates[column_index] = duty_date

        schell_row_index: int | None = None
        for row in rows[2:]:
            if row:
                label = normalize_label(row[0].value)
                if "schell" in label:
                    schell_row_index = row[0].row
                    break

        for row in rows[2:]:
            if not row:
                continue
            label_cell = row[0]
            duty_label = clean_cell_value(label_cell.value)
            if not duty_label:
                continue
            if normalize_label(duty_label) in ROTA_STOP_LABELS:
                break

            classified_duty_type = classify_duty_label(duty_label)
            normalized_duty_label = normalize_label(duty_label)
            if (
                (month.year, month.month) < (2025, 11)
                and classified_duty_type in {"CB_CO4TH_24HR", "RC_CO4TH_24HR"}
            ):
                classified_duty_type = (
                    "CB_CO_12HR" if classified_duty_type == "CB_CO4TH_24HR" else "RC_CO_12HR"
                )
            if (
                (month.year, month.month) < (2025, 11)
                and classified_duty_type == "RC_1ST_A_24HR"
                and "co" in normalize_classification_label(duty_label)
            ):
                classified_duty_type = "RC_CO_12HR"
            if (
                schell_row_index is not None
                and label_cell.row > schell_row_index
                and normalized_duty_label.startswith("main")
                and ("1" in normalized_duty_label or "2" in normalized_duty_label)
            ):
                classified_duty_type = "PAC"
            if classified_duty_type is None:
                warnings.append(
                    ParseWarning(
                        path,
                        worksheet.title,
                        label_cell.row,
                        label_cell.column,
                        "UNMAPPED_DUTY_LABEL",
                        f"Could not map duty label: {duty_label}",
                    )
                )
                duty_types = ["UNMAPPED"]
            elif classified_duty_type == "SCHELL_AND_FLOATING":
                duty_types = ["SCHELL_24HR", "FLOATING_24HR"]
            else:
                duty_types = [classified_duty_type]

            starts_at: datetime
            ends_at: datetime
            is_24hr: bool
            for duty_type in duty_types:
                for column_index, cell in enumerate(row[1:], start=2):
                    duty_date = column_dates.get(column_index)
                    if duty_date is None or is_blank_assignment_value(cell.value):
                        continue
                    person_name = clean_person_name(cell.value)
                    if not is_valid_person_name(person_name):
                        warnings.append(
                            ParseWarning(
                                path,
                                worksheet.title,
                                cell.row,
                                column_index,
                                "INVALID_PERSON_NAME",
                                f"Discarded non-person rota value: {clean_cell_value(cell.value)}",
                            )
                        )
                        continue

                    starts_at, ends_at, is_24hr = duty_window_for_date(duty_date, duty_type)
                    weekday_label = ""
                    if column_index - 1 < len(weekday_cells):
                        weekday_label = clean_cell_value(weekday_cells[column_index - 1].value)

                    assignments.append(
                        ParsedRotaAssignment(
                            source_file=path,
                            sheet_name=worksheet.title,
                            duty_date=duty_date,
                            weekday_label=weekday_label,
                            row_index=cell.row,
                            column_index=column_index,
                            column_label=get_column_letter(column_index),
                            duty_label=duty_label,
                            duty_type=duty_type,
                            person_name=person_name,
                            raw_person_name=clean_cell_value(cell.value),
                            starts_at=starts_at,
                            ends_at=ends_at,
                            is_24hr=is_24hr,
                        )
                    )
    finally:
        workbook.close()

    return ParsedRotaWorkbook(path, month, tuple(assignments), tuple(warnings))


def parse_unitwise_workbook(path: Path) -> ParsedUnitwiseWorkbook:
    require_supported_excel(path)
    month = reconstruct_month_from_path(path)
    if month is None:
        msg = f"Could not infer month/year from {path.name}."
        raise ValueError(msg)

    workbook = load_workbook(path, read_only=True, data_only=True)
    postings: list[ParsedUnitPosting] = []
    warnings: list[ParseWarning] = []
    try:
        worksheet = workbook.worksheets[0]
        rows = list(worksheet.iter_rows(values_only=False))
        if not rows:
            warnings.append(
                ParseWarning(path, worksheet.title, None, None, "EMPTY_UNITWISE", "No rows found.")
            )
            return ParsedUnitwiseWorkbook(path, month, tuple(postings), tuple(warnings))

        unit_by_column = {
            column_index: clean_cell_value(cell.value)
            for column_index, cell in enumerate(rows[0][1:], start=2)
            if clean_cell_value(cell.value)
        }
        current_posting_label = ""
        for row in rows[1:]:
            if not row:
                continue
            label = clean_cell_value(row[0].value)
            if label:
                current_posting_label = label
            if not current_posting_label:
                continue

            for column_index, cell in enumerate(row[1:], start=2):
                unit_label = unit_by_column.get(column_index)
                if unit_label is None or is_blank_assignment_value(cell.value):
                    continue
                names = split_unitwise_names(cell.value)
                if not names:
                    warnings.append(
                        ParseWarning(
                            path,
                            worksheet.title,
                            cell.row,
                            column_index,
                            "INVALID_PERSON_NAME",
                            f"Discarded non-person unitwise value: {clean_cell_value(cell.value)}",
                        )
                    )
                    continue
                for person_name in names:
                    postings.append(
                        ParsedUnitPosting(
                            source_file=path,
                            sheet_name=worksheet.title,
                            unit_label=unit_label,
                            posting_label=current_posting_label,
                            person_name=person_name,
                            raw_person_name=clean_cell_value(cell.value),
                            row_index=cell.row,
                            column_index=column_index,
                            column_label=get_column_letter(column_index),
                        )
                    )
    finally:
        workbook.close()

    return ParsedUnitwiseWorkbook(path, month, tuple(postings), tuple(warnings))
