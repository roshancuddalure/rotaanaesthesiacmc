from __future__ import annotations

import io
import re
import unicodedata
from datetime import date, datetime
from difflib import SequenceMatcher
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
import pandas as pd
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.models import LeaveRequest, Person
from app.services.leave import month_bounds

NAME_COLUMNS = {
    "name",
    "person",
    "doctor",
    "doctor_name",
    "member",
    "staff",
    "staff_name",
    "consultant",
    "consultant_name",
    "employee",
    "employee_name",
    "faculty",
    "faculty_name",
}
START_COLUMNS = {"start", "starts", "start_date", "from", "from_date", "date", "leave_from", "leave_date"}
END_COLUMNS = {"end", "ends", "end_date", "to", "to_date", "leave_to"}
TYPE_COLUMNS = {"type", "leave_type", "category", "leave_category", "kind"}
SLOT_COLUMNS = {"slot", "leave_slot", "session", "half", "leave_session"}
STATUS_COLUMNS = {"status", "approval_status"}
NOTES_COLUMNS = {"notes", "note", "reason", "remarks"}
MAX_HEADER_SCAN_ROWS = 25
TITLE_PATTERN = re.compile(
    r"^(dr|prof|professor|doctor|consultant|sr|jr|pg|md|ms|dm|mch)\.?\s+",
    re.IGNORECASE,
)
SLOT_ALIASES = {
    "FULL": "FULL_DAY",
    "FULLDAY": "FULL_DAY",
    "FULL_DAY": "FULL_DAY",
    "WHOLEDAY": "FULL_DAY",
    "WHOLE_DAY": "FULL_DAY",
    "ALLDAY": "FULL_DAY",
    "ALL_DAY": "FULL_DAY",
    "FD": "FULL_DAY",
    "FUL": "FULL_DAY",
    "F": "FULL_DAY",
    "AM": "AM",
    "FN": "AM",
    "FORENOON": "AM",
    "MORNING": "AM",
    "FIRSTHALF": "AM",
    "FIRST_HALF": "AM",
    "1STHALF": "AM",
    "FIRSTSESSION": "AM",
    "SESSION1": "AM",
    "FH": "AM",
    "HALFDAYAM": "AM",
    "PM": "PM",
    "AN": "PM",
    "AFTERNOON": "PM",
    "SECONDHALF": "PM",
    "SECOND_HALF": "PM",
    "2NDHALF": "PM",
    "SECONDSESSION": "PM",
    "SESSION2": "PM",
    "SH": "PM",
    "HALFDAYPM": "PM",
    "NIGHT": "NIGHT",
    "NIGHTSHIFT": "NIGHT",
    "N": "NIGHT",
}
LEAVE_TYPE_ALIASES = {
    "AL": "ANNUAL_LEAVE",
    "ANNUALLEAVE": "ANNUAL_LEAVE",
    "EL": "EARNED_LEAVE",
    "EARNEDLEAVE": "EARNED_LEAVE",
    "CL": "CASUAL_LEAVE",
    "CASUALLEAVE": "CASUAL_LEAVE",
    "SL": "SICK_LEAVE",
    "SICKLEAVE": "SICK_LEAVE",
    "ML": "MEDICAL_LEAVE",
    "MEDICALLEAVE": "MEDICAL_LEAVE",
    "MATERNITYLEAVE": "MATERNITY_LEAVE",
    "PATERNITYLEAVE": "PATERNITY_LEAVE",
    "CCL": "CHILD_CARE_LEAVE",
    "LOP": "LOSS_OF_PAY",
    "LWP": "LEAVE_WITHOUT_PAY",
    "OD": "ON_DUTY",
    "OFF": "OFF",
    "LV": "LEAVE",
    "LEAVE": "LEAVE",
}
STATUS_ALIASES = {
    "APPROVED": "approved",
    "APPROVE": "approved",
    "SANCTIONED": "approved",
    "REQUESTED": "requested",
    "REQUEST": "requested",
    "PENDING": "requested",
    "REJECTED": "rejected",
    "CANCELLED": "cancelled",
    "CANCELED": "cancelled",
}
NOISE_NAME_KEYS = {
    "sno",
    "slno",
    "serialno",
    "no",
    "name",
    "names",
    "person",
    "doctor",
    "staff",
    "consultant",
    "prof",
    "assprof",
    "assistantprof",
    "assistantprofessor",
    "associateprof",
    "associateprofessor",
    "sr",
    "jr",
    "pg",
    "postgraduate",
    "postgraduates",
    "pglist",
    "e2pglist",
    "faculty",
    "employee",
    "unit",
    "department",
    "date",
    "from",
    "to",
    "slot",
    "session",
    "status",
    "remarks",
    "notes",
    "total",
    "subtotal",
    "summary",
    "nil",
    "none",
    "na",
    "notapplicable",
    "noone",
    "noleave",
    "leave",
    "leaves",
    "leavedetails",
    "onleave",
    "holiday",
    "publicholiday",
    "cme",
    "nursepaincme",
    "wtlt",
    "may",
    "january",
    "february",
    "march",
    "april",
    "june",
    "july",
    "august",
    "september",
    "october",
    "november",
    "december",
    "mon",
    "tue",
    "wed",
    "thu",
    "fri",
    "sat",
    "sun",
    "monday",
    "tuesday",
    "wednesday",
    "thursday",
    "friday",
    "saturday",
    "sunday",
}
SLOT_TOKEN_PATTERN = (
    r"FULL\s*DAY|WHOLE\s*DAY|ALL\s*DAY|HALF\s*DAY\s*AM|HALF\s*DAY\s*PM|"
    r"FIRST\s*HALF|SECOND\s*HALF|1ST\s*HALF|2ND\s*HALF|FIRST\s*SESSION|SECOND\s*SESSION|"
    r"SESSION\s*1|SESSION\s*2|FORENOON|AFTERNOON|MORNING|NIGHT\s*SHIFT|NIGHT|FD|FH|SH|AM|PM|FN|AN|N"
)
PLAIN_SUFFIX_SLOT_PATTERN = (
    r"FULL\s*DAY|WHOLE\s*DAY|ALL\s*DAY|HALF\s*DAY\s*AM|HALF\s*DAY\s*PM|"
    r"FIRST\s*HALF|SECOND\s*HALF|1ST\s*HALF|2ND\s*HALF|FORENOON|AFTERNOON|MORNING|"
    r"NIGHT\s*SHIFT|NIGHT|FD|FH|SH|AM|PM|FN|AN"
)
TRAILING_NOISE_PATTERN = re.compile(
    r"(?:\s*[-/:,;]\s*|\s+)"
    r"(?:ANNUAL\s+LEAVE|EARNED\s+LEAVE|CASUAL\s+LEAVE|SICK\s+LEAVE|MEDICAL\s+LEAVE|"
    r"MATERNITY\s+LEAVE|PATERNITY\s+LEAVE|CHILD\s+CARE\s+LEAVE|"
    r"AL|EL|CL|SL|ML|CCL|LOP|LWP|OD|OFF|LV|LEAVE|"
    r"APPROVED|PENDING|REQUESTED|SANCTIONED)\s*$",
    re.IGNORECASE,
)

PersonLookup = dict[str, Person | None]


def normalize_name(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", str(value))
    ascii_text = "".join(character for character in decomposed if not unicodedata.combining(character))
    return re.sub(r"[^a-z0-9]+", "", ascii_text.lower())


def compact_token(value: Any) -> str:
    return normalize_header(clean_cell_text(value)).upper().replace("_", "")


def strip_titles(value: str) -> str:
    cleaned = value.strip()
    while True:
        next_value = TITLE_PATTERN.sub("", cleaned).strip()
        if next_value == cleaned:
            return cleaned
        cleaned = next_value


def slot_from_token(value: Any) -> str | None:
    key = compact_token(value)
    if not key or key == "NAN":
        return None
    return SLOT_ALIASES.get(key)


def normalize_slot(value: Any) -> str:
    key = compact_token(value)
    if not key or key == "NAN":
        return "FULL_DAY"
    return SLOT_ALIASES.get(key, "FULL_DAY")


def normalize_leave_type(value: Any) -> str:
    raw = clean_cell_text(value)
    key = compact_token(raw)
    if not key or key == "NAN":
        return "ANNUAL_LEAVE"
    return LEAVE_TYPE_ALIASES.get(key, normalize_header(raw).upper() or "ANNUAL_LEAVE")


def leading_marker_removed(value: str) -> str:
    return re.sub(r"^\s*(?:[#*-]\s*)?(?:\d+|[ivxlcdm]+|[a-z])[\.)\]:-]\s+", "", value, flags=re.IGNORECASE)


def annotation_is_noise(value: str) -> bool:
    key = compact_token(value)
    normalized = normalize_name(value)
    return (
        not key
        or key in SLOT_ALIASES
        or key in LEAVE_TYPE_ALIASES
        or key in STATUS_ALIASES
        or normalized in NOISE_NAME_KEYS
    )


def remove_bracketed_noise(value: str) -> str:
    def replace_annotation(match: re.Match[str]) -> str:
        inner = match.group(1) or match.group(2) or ""
        return " " if annotation_is_noise(inner) else match.group(0)

    return re.sub(r"\(([^)]*)\)|\[([^\]]*)\]", replace_annotation, value)


def extract_slot_annotations(value: str, fallback_slot: str) -> tuple[str, str]:
    slot = normalize_slot(fallback_slot)

    for match in re.finditer(r"\(([^)]*)\)|\[([^\]]*)\]", value):
        token = match.group(1) or match.group(2) or ""
        parsed_slot = slot_from_token(token)
        if parsed_slot:
            slot = parsed_slot

    suffix_match = re.search(rf"\s*[-/:,;]\s*({SLOT_TOKEN_PATTERN})\s*$", value, re.IGNORECASE)
    if suffix_match:
        return value[: suffix_match.start()].strip(), normalize_slot(suffix_match.group(1))

    plain_suffix_match = re.search(rf"\s+({PLAIN_SUFFIX_SLOT_PATTERN})\s*$", value, re.IGNORECASE)
    if plain_suffix_match:
        return value[: plain_suffix_match.start()].strip(), normalize_slot(plain_suffix_match.group(1))

    prefix_match = re.match(rf"^\s*({SLOT_TOKEN_PATTERN})\s*[-/:,;]\s*(.+)$", value, re.IGNORECASE)
    if prefix_match:
        return prefix_match.group(2).strip(), normalize_slot(prefix_match.group(1))

    return value, slot


def remove_trailing_noise(value: str) -> str:
    cleaned = value
    while True:
        next_value = TRAILING_NOISE_PATTERN.sub("", cleaned).strip()
        if next_value == cleaned:
            return cleaned
        cleaned = next_value


def clean_person_text(value: Any) -> str:
    cleaned = clean_cell_text(value)
    cleaned = cleaned.replace("–", "-").replace("—", "-")
    cleaned = leading_marker_removed(cleaned)
    cleaned = remove_bracketed_noise(cleaned)
    cleaned = remove_trailing_noise(cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" -/:,;")
    return strip_titles(cleaned)


def extract_name_and_slot(raw_name: str, fallback_slot: str = "FULL_DAY") -> tuple[str, str]:
    cleaned, slot = extract_slot_annotations(clean_cell_text(raw_name), fallback_slot)
    cleaned = clean_person_text(cleaned)
    cleaned, slot = extract_slot_annotations(cleaned, slot)
    cleaned = clean_person_text(cleaned)
    return cleaned, slot


def normalize_header(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", str(value))
    ascii_text = "".join(character for character in decomposed if not unicodedata.combining(character))
    cleaned = re.sub(r"[^a-zA-Z0-9]+", "_", ascii_text.strip().lower()).strip("_")
    return cleaned


def add_lookup_key(lookup: PersonLookup, key: str, person: Person) -> None:
    if not key:
        return
    existing = lookup.get(key)
    if existing is None and key in lookup:
        return
    if existing is not None and existing.id != person.id:
        lookup[key] = None
        return
    lookup[key] = person


def person_lookup(db: Session) -> PersonLookup:
    people = list(db.scalars(select(Person).options(selectinload(Person.aliases))))
    lookup: PersonLookup = {}
    for person in people:
        add_lookup_key(lookup, normalize_name(person.canonical_name), person)
        add_lookup_key(lookup, normalize_name(strip_titles(person.canonical_name)), person)
        for alias in person.aliases:
            add_lookup_key(lookup, normalize_name(alias.alias), person)
            add_lookup_key(lookup, normalize_name(strip_titles(alias.alias)), person)
    return lookup


def suggest_person(cleaned_name: str, lookup: PersonLookup) -> Person | None:
    key = normalize_name(cleaned_name)
    if not key:
        return None
    scores_by_person: dict[UUID, tuple[float, Person]] = {}
    for candidate_key, person in lookup.items():
        if person is None or len(candidate_key) < 3:
            continue
        score = SequenceMatcher(None, key, candidate_key).ratio()
        previous = scores_by_person.get(person.id)
        if previous is None or score > previous[0]:
            scores_by_person[person.id] = (score, person)
    ranked = sorted(scores_by_person.values(), key=lambda item: item[0], reverse=True)
    if not ranked:
        return None
    best_score, best_person = ranked[0]
    second_score = ranked[1][0] if len(ranked) > 1 else 0.0
    if best_score >= 0.9 and best_score - second_score >= 0.04:
        return best_person
    return None


def match_person(cleaned_name: str, lookup: PersonLookup) -> tuple[Person | None, str | None, str, Person | None]:
    key = normalize_name(cleaned_name)
    if key in lookup:
        person = lookup[key]
        if person is None:
            return None, "ambiguous_exact", "ambiguous", None
        return person, "normalized_exact", "high", None
    suggestion = suggest_person(cleaned_name, lookup)
    return None, None, "suggestion" if suggestion else "low", suggestion


def pick_column(columns: list[str], candidates: set[str]) -> str | None:
    for column in columns:
        if column in candidates:
            return column
    for column in columns:
        if any(candidate in column for candidate in candidates):
            return column
    return None


def parse_date_value(value: Any) -> date | None:
    if value is None or str(value).strip() == "" or str(value).lower() == "nan":
        return None
    parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def normalize_status(value: Any) -> str:
    key = compact_token(value)
    if not key or key == "NAN":
        return "imported_pending_review"
    return STATUS_ALIASES.get(key, normalize_header(clean_cell_text(value)) or "imported_pending_review")


def read_leave_tables(filename: str, content: bytes) -> list[tuple[str, pd.DataFrame, str]]:
    lowered = filename.lower()
    if lowered.endswith(".csv"):
        last_error: Exception | None = None
        for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
            try:
                return [
                    (
                        "CSV",
                        pd.read_csv(io.BytesIO(content), sep=None, engine="python", encoding=encoding),
                        "table",
                    )
                ]
            except (UnicodeDecodeError, pd.errors.ParserError) as exc:
                last_error = exc
        raise ValueError(f"Unable to parse CSV leave file: {last_error}") from last_error
    if lowered.endswith(".xlsx"):
        workbook = load_workbook(io.BytesIO(content), data_only=True)
        tables: list[tuple[str, pd.DataFrame, str]] = []
        for worksheet in workbook.worksheets:
            values = [[cell.value for cell in row] for row in worksheet.iter_rows()]
            for merged_range in worksheet.merged_cells.ranges:
                value = worksheet.cell(merged_range.min_row, merged_range.min_col).value
                for row_index in range(merged_range.min_row - 1, merged_range.max_row):
                    for column_index in range(merged_range.min_col - 1, merged_range.max_col):
                        values[row_index][column_index] = value
            tables.append((worksheet.title, pd.DataFrame(values), "excel_raw"))
        return tables
    if lowered.endswith(".xls"):
        sheets = pd.read_excel(io.BytesIO(content), sheet_name=None, header=None)
        return [(sheet_name, frame, "excel_raw") for sheet_name, frame in sheets.items()]
    raise ValueError("Leave import supports CSV, XLS, or XLSX files")


def clean_cell_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    cleaned = unicodedata.normalize("NFKC", str(value))
    cleaned = cleaned.replace("\u00a0", " ")
    cleaned = re.sub(r"[\u200b-\u200f\ufeff]", "", cleaned)
    cleaned = re.sub(r"[ \t\r\f\v]+", " ", cleaned)
    return cleaned.strip().strip("\"'")


def is_noise_name(value: str) -> bool:
    cleaned = clean_cell_text(value)
    normalized = normalize_name(cleaned)
    if not normalized:
        return True
    if normalized in NOISE_NAME_KEYS:
        return True
    if compact_token(cleaned) in LEAVE_TYPE_ALIASES or compact_token(cleaned) in STATUS_ALIASES:
        return True
    if normalized.isdigit():
        return True
    if re.fullmatch(r"[\d\s./:-]+", cleaned):
        return True
    if parse_date_value(cleaned):
        return True
    return False


def detect_header_row(raw_frame: pd.DataFrame) -> int | None:
    scan_limit = min(MAX_HEADER_SCAN_ROWS, len(raw_frame))
    for row_index in range(scan_limit):
        values = [normalize_header(clean_cell_text(value)) for value in raw_frame.iloc[row_index].tolist()]
        if pick_column(values, NAME_COLUMNS) and pick_column(values, START_COLUMNS):
            return row_index
    return None


def table_from_header(raw_frame: pd.DataFrame, header_row: int) -> pd.DataFrame:
    frame = raw_frame.iloc[header_row + 1 :].copy()
    frame.columns = [normalize_header(clean_cell_text(value)) for value in raw_frame.iloc[header_row]]
    return frame.dropna(how="all")


def date_columns(raw_frame: pd.DataFrame, starts_on: date, ends_on: date) -> tuple[int | None, dict[int, date]]:
    best_row: int | None = None
    best_dates: dict[int, date] = {}
    scan_limit = min(MAX_HEADER_SCAN_ROWS, len(raw_frame))
    for row_index in range(scan_limit):
        found: dict[int, date] = {}
        for column_index, value in enumerate(raw_frame.iloc[row_index].tolist()):
            parsed = parse_date_value(value)
            if parsed and starts_on <= parsed <= ends_on:
                found[column_index] = parsed
        if len(found) > len(best_dates):
            best_row = row_index
            best_dates = found
    if len(best_dates) < 2:
        return None, {}
    return best_row, best_dates


def split_names(value: str) -> list[str]:
    cleaned = clean_cell_text(value).replace("\n", ",")
    cleaned = re.sub(r"\s+(?:and|&|\+)\s+", ",", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s+/\s+", ",", cleaned)
    parts = re.split(r",|;", cleaned)
    return [leading_marker_removed(part.strip()) for part in parts if part.strip()]


def make_preview_row(
    *,
    row_number: int,
    raw_name: str,
    starts: date | None,
    ends: date | None,
    leave_type: str,
    leave_slot: str,
    status_value: str,
    notes: str,
    month_start: date,
    month_end: date,
    lookup: PersonLookup,
    sheet_name: str,
    source_format: str,
    confidence: str,
) -> dict[str, object]:
    issues: list[str] = []
    cleaned_name, parsed_slot = extract_name_and_slot(raw_name, leave_slot)
    if not cleaned_name:
        issues.append("Missing person name")
    if starts is None:
        issues.append("Missing or invalid start date")
    if ends is None:
        issues.append("Missing or invalid end date")
    if starts and ends and ends < starts:
        issues.append("End date before start date")
    if starts and ends and (starts > month_end or ends < month_start):
        issues.append("Leave is outside selected month")
    person, match_method, match_confidence, suggestion = match_person(cleaned_name, lookup)
    if match_confidence == "ambiguous":
        issues.append("Ambiguous department member match")
    elif cleaned_name and person is None:
        issues.append("Unresolved department member")
        if suggestion is not None:
            issues.append(f"Suggested match: {suggestion.canonical_name}")
    preview_status = "matched" if person and not issues else "needs_review"
    return {
        "row_number": row_number,
        "sheet_name": sheet_name,
        "source_format": source_format,
        "confidence": confidence,
        "match_confidence": match_confidence,
        "raw_person_name": raw_name,
        "cleaned_person_name": cleaned_name,
        "person_id": str(person.id) if person else None,
        "person_name": person.canonical_name if person else None,
        "suggested_person_id": str(suggestion.id) if suggestion else None,
        "suggested_person_name": suggestion.canonical_name if suggestion else None,
        "match_method": match_method,
        "starts_on": starts.isoformat() if starts else None,
        "ends_on": ends.isoformat() if ends else None,
        "leave_type": leave_type,
        "leave_slot": parsed_slot,
        "status": status_value,
        "notes": notes,
        "preview_status": preview_status,
        "issues": issues,
    }


def parse_table_rows(
    frame: pd.DataFrame,
    *,
    month_start: date,
    month_end: date,
    lookup: PersonLookup,
    sheet_name: str,
    source_format: str,
    row_offset: int,
) -> list[dict[str, object]]:
    frame.columns = [normalize_header(str(column)) for column in frame.columns]
    columns = list(frame.columns)
    name_col = pick_column(columns, NAME_COLUMNS)
    start_col = pick_column(columns, START_COLUMNS)
    end_col = pick_column(columns, END_COLUMNS)
    if not name_col or not start_col:
        return []

    type_col = pick_column(columns, TYPE_COLUMNS)
    slot_col = pick_column(columns, SLOT_COLUMNS)
    status_col = pick_column(columns, STATUS_COLUMNS)
    notes_col = pick_column(columns, NOTES_COLUMNS)
    rows: list[dict[str, object]] = []

    for index, row in frame.iterrows():
        raw_name = clean_cell_text(row.get(name_col, ""))
        if is_noise_name(raw_name):
            continue
        starts = parse_date_value(row.get(start_col))
        ends = parse_date_value(row.get(end_col)) if end_col else starts
        rows.append(
            make_preview_row(
                row_number=int(index) + row_offset,
                raw_name=raw_name,
                starts=starts,
                ends=ends,
                leave_type=normalize_leave_type(row.get(type_col)) if type_col else "ANNUAL_LEAVE",
                leave_slot=normalize_slot(row.get(slot_col)) if slot_col else "FULL_DAY",
                status_value=normalize_status(row.get(status_col)) if status_col else "imported_pending_review",
                notes=clean_cell_text(row.get(notes_col, "")) if notes_col else "",
                month_start=month_start,
                month_end=month_end,
                lookup=lookup,
                sheet_name=sheet_name,
                source_format=source_format,
                confidence="high",
            )
        )
    return rows


def parse_wide_calendar_rows(
    raw_frame: pd.DataFrame,
    *,
    month_start: date,
    month_end: date,
    lookup: PersonLookup,
    sheet_name: str,
) -> list[dict[str, object]]:
    header_row, dates_by_column = date_columns(raw_frame, month_start, month_end)
    if header_row is None:
        return []
    daily_entries: list[tuple[str, str, str, int, date]] = []
    for row_index in range(header_row + 1, len(raw_frame)):
        row = raw_frame.iloc[row_index]
        for column_index, day in dates_by_column.items():
            raw_cell = clean_cell_text(row.iloc[column_index])
            for raw_name in split_names(raw_cell):
                cleaned_name, parsed_slot = extract_name_and_slot(raw_name)
                if is_noise_name(raw_name) or is_noise_name(cleaned_name):
                    continue
                daily_entries.append((raw_name, cleaned_name, parsed_slot, row_index + 1, day))

    grouped: dict[tuple[str, str, int], dict[str, object]] = {}
    for raw_name, cleaned_name, parsed_slot, row_number, day in daily_entries:
        group = grouped.setdefault(
            (cleaned_name, parsed_slot, row_number),
            {"raw_name": raw_name, "days": []},
        )
        group["days"].append(day)  # type: ignore[union-attr]

    rows: list[dict[str, object]] = []
    for (cleaned_name, parsed_slot, row_number), group in grouped.items():
        raw_name = str(group["raw_name"])
        days = group["days"]  # type: ignore[assignment]
        sorted_days = sorted(set(days))
        range_start = sorted_days[0]
        previous = sorted_days[0]
        for day in sorted_days[1:] + [date.max]:
            if day.toordinal() == previous.toordinal() + 1:
                previous = day
                continue
            rows.append(
                make_preview_row(
                    row_number=row_number,
                    raw_name=raw_name if raw_name else cleaned_name,
                    starts=range_start,
                    ends=previous,
                    leave_type="ANNUAL_LEAVE",
                    leave_slot=parsed_slot,
                    status_value="imported_pending_review",
                    notes=f"Parsed from date-column leave calendar sheet {sheet_name}",
                    month_start=month_start,
                    month_end=month_end,
                    lookup=lookup,
                    sheet_name=sheet_name,
                    source_format="wide_calendar",
                    confidence="medium",
                )
            )
            range_start = day
            previous = day
    return rows


def mark_duplicates(rows: list[dict[str, object]]) -> None:
    seen: dict[tuple[object, object, object, object], int] = {}
    for index, row in enumerate(rows):
        key = (
            row.get("person_id") or normalize_name(str(row.get("raw_person_name") or "")),
            row.get("starts_on"),
            row.get("ends_on"),
            row.get("leave_slot"),
        )
        if key in seen:
            row["preview_status"] = "needs_review"
            row["issues"].append("Possible duplicate leave row")  # type: ignore[union-attr]
            rows[seen[key]]["preview_status"] = "needs_review"
            rows[seen[key]]["issues"].append("Possible duplicate leave row")  # type: ignore[union-attr]
        else:
            seen[key] = index


def parse_iso_date(value: object) -> date | None:
    if not value:
        return None
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").date()
    except ValueError:
        return None


def mark_existing_duplicates(db: Session, rows: list[dict[str, object]], month: str) -> None:
    starts_on, ends_on = month_bounds(month)
    existing = list(
        db.scalars(
            select(LeaveRequest).where(
                LeaveRequest.starts_on <= ends_on,
                LeaveRequest.ends_on >= starts_on,
                LeaveRequest.status != "cancelled",
            )
        )
    )
    existing_keys = {
        (
            str(leave.person_id),
            leave.starts_on.isoformat(),
            leave.ends_on.isoformat(),
            leave.leave_slot,
        )
        for leave in existing
    }
    for row in rows:
        key = (
            row.get("person_id"),
            row.get("starts_on"),
            row.get("ends_on"),
            row.get("leave_slot"),
        )
        if key in existing_keys:
            row["preview_status"] = "needs_review"
            row["issues"].append("Existing matching leave already recorded")  # type: ignore[union-attr]


def row_is_safe_to_apply(row: dict[str, object]) -> bool:
    return (
        row.get("preview_status") == "matched"
        and row.get("person_id") is not None
        and row.get("starts_on") is not None
        and row.get("ends_on") is not None
        and not row.get("issues")
    )


def apply_leave_import(db: Session, filename: str, content: bytes, month: str) -> dict[str, object]:
    preview = preview_leave_import(db, filename, content, month)
    created = 0
    skipped_rows: list[dict[str, object]] = []
    for row in preview["rows"]:
        if not row_is_safe_to_apply(row):
            skipped_rows.append(row)
            continue
        starts_on = parse_iso_date(row["starts_on"])
        ends_on = parse_iso_date(row["ends_on"])
        if starts_on is None or ends_on is None:
            skipped_rows.append(row)
            continue
        leave = LeaveRequest(
            person_id=UUID(str(row["person_id"])),
            leave_type=str(row["leave_type"]),
            leave_slot=str(row["leave_slot"]),
            starts_on=starts_on,
            ends_on=ends_on,
            status=str(row["status"]),
            source="import",
            raw_person_name=str(row["raw_person_name"]),
            notes=(
                f"{row.get('notes') or ''} "
                f"[Imported from {filename}, sheet {row.get('sheet_name')}, row {row.get('row_number')}]"
            ).strip(),
        )
        db.add(leave)
        created += 1
    db.commit()
    return {
        "filename": filename,
        "month": month,
        "created_rows": created,
        "skipped_rows": len(skipped_rows),
        "skipped_preview_rows": skipped_rows,
        "preview": preview,
    }


def preview_leave_import(db: Session, filename: str, content: bytes, month: str) -> dict[str, object]:
    starts_on, ends_on = month_bounds(month)
    lookup = person_lookup(db)
    rows: list[dict[str, object]] = []
    parser_warnings: list[str] = []
    sheets: list[str] = []
    source_formats: set[str] = set()

    for sheet_name, frame, frame_kind in read_leave_tables(filename, content):
        sheets.append(sheet_name)
        if frame_kind == "table":
            parsed = parse_table_rows(
                frame.dropna(how="all"),
                month_start=starts_on,
                month_end=ends_on,
                lookup=lookup,
                sheet_name=sheet_name,
                source_format="table",
                row_offset=2,
            )
        else:
            header_row = detect_header_row(frame)
            if header_row is not None:
                parsed = parse_table_rows(
                    table_from_header(frame, header_row),
                    month_start=starts_on,
                    month_end=ends_on,
                    lookup=lookup,
                    sheet_name=sheet_name,
                    source_format="table_detected_header",
                    row_offset=header_row + 2,
                )
            else:
                parsed = parse_wide_calendar_rows(
                    frame,
                    month_start=starts_on,
                    month_end=ends_on,
                    lookup=lookup,
                    sheet_name=sheet_name,
                )
        if not parsed:
            parser_warnings.append(f"No leave rows detected in sheet {sheet_name}")
        rows.extend(parsed)
        source_formats.update(str(row["source_format"]) for row in parsed)

    mark_duplicates(rows)
    mark_existing_duplicates(db, rows, month)
    matched = sum(1 for row in rows if row["preview_status"] == "matched")
    unresolved = sum(1 for row in rows if "Unresolved department member" in row["issues"])
    invalid = sum(
        1
        for row in rows
        if row["preview_status"] != "matched" and "Unresolved department member" not in row["issues"]
    )

    return {
        "filename": filename,
        "month": month,
        "total_rows": len(rows),
        "matched_rows": matched,
        "unresolved_rows": unresolved,
        "invalid_rows": invalid,
        "sheets": sheets,
        "source_formats": sorted(source_formats),
        "parser_warnings": parser_warnings,
        "rows": rows,
    }
